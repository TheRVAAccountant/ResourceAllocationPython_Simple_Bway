"""Performance tests for large dataset scenarios (1000+ routes)."""

import gc
import os
import time
from datetime import date, datetime, timedelta
from typing import Any, Dict, List

import numpy as np
import pandas as pd
import psutil
import pytest
from openpyxl import Workbook

from src.core.gas_compatible_allocator import GASCompatibleAllocator
from src.services.daily_details_thick_borders import DailyDetailsThickBorderService
from src.services.duplicate_validator import DuplicateVehicleValidator
from src.services.unassigned_vehicles_writer import UnassignedVehiclesWriter


class TestLargeDatasetPerformance:
    """Performance tests for production-scale datasets."""

    # Performance thresholds
    MAX_DUPLICATE_VALIDATION_TIME = 5.0  # seconds for 1000 allocations
    MAX_UNASSIGNED_PROCESSING_TIME = 10.0  # seconds for 1000 vehicles
    MAX_THICK_BORDERS_TIME = 15.0  # seconds for 1000 rows
    MAX_MEMORY_INCREASE = 200  # MB

    @pytest.fixture
    def performance_monitor(self):
        """Monitor performance metrics during tests."""
        process = psutil.Process(os.getpid())

        class PerformanceMonitor:
            def __init__(self):
                self.start_time = None
                self.start_memory = None
                self.end_time = None
                self.end_memory = None

            def start(self):
                gc.collect()  # Clean up before measurement
                self.start_time = time.time()
                self.start_memory = process.memory_info().rss / 1024 / 1024  # MB

            def stop(self):
                self.end_time = time.time()
                self.end_memory = process.memory_info().rss / 1024 / 1024  # MB

            @property
            def execution_time(self):
                return self.end_time - self.start_time if self.end_time else None

            @property
            def memory_increase(self):
                return self.end_memory - self.start_memory if self.end_memory else None

            def report(self, operation_name: str):
                print(f"\n{operation_name} Performance:")
                print(f"  Execution Time: {self.execution_time:.3f} seconds")
                print(
                    f"  Memory Usage: {self.start_memory:.1f} -> {self.end_memory:.1f} MB (+{self.memory_increase:.1f} MB)"
                )

        return PerformanceMonitor()

    def generate_large_allocation_dataset(
        self, count: int, duplicate_rate: float = 0.1
    ) -> List[Dict[str, Any]]:
        """Generate large allocation dataset for testing.

        Args:
            count: Number of allocation records to generate.
            duplicate_rate: Proportion of duplicates to include (0.0-1.0).

        Returns:
            List of allocation result dictionaries.
        """
        allocations = []

        # Calculate unique vehicles vs duplicates
        unique_vehicles = int(count * (1 - duplicate_rate))
        duplicate_count = count - unique_vehicles

        # Generate unique allocations
        for i in range(unique_vehicles):
            allocations.append(
                {
                    "Van ID": f"BW{i + 1}",
                    "Route Code": f"CX{i + 1}",
                    "Associate Name": f"Driver_{i + 1}",
                    "Service Type": np.random.choice(
                        [
                            "Standard Parcel - Large Van",
                            "Standard Parcel - Extra Large Van - US",
                            "Standard Parcel Step Van - US",
                        ]
                    ),
                    "Wave": f"{np.random.randint(8, 12)}:00 AM",
                    "Staging Location": f"STG.G.{np.random.randint(1, 20)}",
                }
            )

        # Add duplicates by reusing vehicle IDs
        for i in range(duplicate_count):
            van_id = f"BW{np.random.randint(1, max(1, unique_vehicles // 2))}"
            allocations.append(
                {
                    "Van ID": van_id,
                    "Route Code": f"CX{unique_vehicles + i + 1}",
                    "Associate Name": f"Driver_{unique_vehicles + i + 1}",
                    "Service Type": np.random.choice(
                        [
                            "Standard Parcel - Large Van",
                            "Standard Parcel - Extra Large Van - US",
                            "Standard Parcel Step Van - US",
                        ]
                    ),
                    "Wave": f"{np.random.randint(8, 12)}:00 AM",
                    "Staging Location": f"STG.G.{np.random.randint(1, 20)}",
                }
            )

        # Shuffle to distribute duplicates randomly
        np.random.shuffle(allocations)
        return allocations

    def generate_large_unassigned_dataset(self, count: int) -> pd.DataFrame:
        """Generate large unassigned vehicles dataset.

        Args:
            count: Number of unassigned vehicles to generate.

        Returns:
            DataFrame with unassigned vehicle data.
        """
        vehicle_types = ["Large", "Extra Large", "Step Van"]
        locations = [f"Depot_{chr(65 + i)}" for i in range(10)]  # Depot_A to Depot_J

        data = []
        for i in range(count):
            data.append(
                {
                    "Van ID": f"BW{i + 10000}",  # High numbers to avoid conflicts
                    "Type": np.random.choice(vehicle_types),
                    "Opnal? Y/N": "Y" if np.random.random() > 0.2 else "N",  # 80% operational
                    "Location": np.random.choice(locations),
                }
            )

        return pd.DataFrame(data)

    def generate_large_vehicle_log(self, count: int) -> Dict[str, Dict]:
        """Generate large vehicle log dictionary.

        Args:
            count: Number of vehicle log entries to generate.

        Returns:
            Dictionary mapping van IDs to vehicle details.
        """
        vehicle_log = {}
        brands = ["Branded", "Rental"]

        for i in range(count):
            van_id = f"BW{i + 10000}"
            vehicle_log[van_id] = {
                "vin": f"1HGCM82633A{i:06d}",
                "geotab": f"GT{i + 10000}" if np.random.random() > 0.1 else "",  # 90% have geotab
                "brand_or_rental": np.random.choice(brands),
                "vehicle_type": np.random.choice(["Large", "Extra Large", "Step Van"]),
            }

        return vehicle_log

    # ==================== Duplicate Validator Performance Tests ====================

    def test_duplicate_validator_1000_allocations(self, duplicate_validator, performance_monitor):
        """Test duplicate validation performance with 1000 allocations."""
        # Generate large dataset with moderate duplicate rate
        allocations = self.generate_large_allocation_dataset(1000, duplicate_rate=0.15)

        performance_monitor.start()
        result = duplicate_validator.validate_allocations(allocations)
        performance_monitor.stop()

        # Performance assertions
        assert (
            performance_monitor.execution_time < self.MAX_DUPLICATE_VALIDATION_TIME
        ), f"Validation took {performance_monitor.execution_time:.2f}s, expected < {self.MAX_DUPLICATE_VALIDATION_TIME}s"

        assert (
            performance_monitor.memory_increase < self.MAX_MEMORY_INCREASE
        ), f"Memory increased by {performance_monitor.memory_increase:.1f}MB, expected < {self.MAX_MEMORY_INCREASE}MB"

        # Functional assertions
        assert isinstance(result.duplicate_count, int)
        assert result.duplicate_count > 0  # Should find duplicates with 15% duplicate rate

        performance_monitor.report("Duplicate Validation (1000 allocations)")

    def test_duplicate_validator_5000_allocations(self, duplicate_validator, performance_monitor):
        """Test duplicate validation performance with 5000 allocations."""
        # Generate very large dataset
        allocations = self.generate_large_allocation_dataset(5000, duplicate_rate=0.2)

        performance_monitor.start()
        result = duplicate_validator.validate_allocations(allocations)
        performance_monitor.stop()

        # Relaxed threshold for very large dataset
        max_time = self.MAX_DUPLICATE_VALIDATION_TIME * 5  # 25 seconds
        assert (
            performance_monitor.execution_time < max_time
        ), f"Large validation took {performance_monitor.execution_time:.2f}s, expected < {max_time}s"

        # Should handle large datasets without excessive memory
        assert (
            performance_monitor.memory_increase < self.MAX_MEMORY_INCREASE * 2
        ), f"Memory increased by {performance_monitor.memory_increase:.1f}MB for large dataset"

        performance_monitor.report("Duplicate Validation (5000 allocations)")

    def test_duplicate_validator_high_duplicate_rate(
        self, duplicate_validator, performance_monitor
    ):
        """Test performance with high duplicate rate (worst case scenario)."""
        # Generate dataset with 50% duplicates (worst case)
        allocations = self.generate_large_allocation_dataset(1000, duplicate_rate=0.5)

        performance_monitor.start()
        result = duplicate_validator.validate_allocations(allocations)
        performance_monitor.stop()

        # Should still perform well even with many duplicates
        assert (
            performance_monitor.execution_time < self.MAX_DUPLICATE_VALIDATION_TIME * 2
        ), f"High duplicate validation took {performance_monitor.execution_time:.2f}s"

        # Should detect many duplicates
        assert result.duplicate_count > 100, "Should detect significant duplicates with 50% rate"

        performance_monitor.report("High Duplicate Rate Validation (50% duplicates)")

    def test_duplicate_report_generation_performance(
        self, duplicate_validator, performance_monitor
    ):
        """Test performance of duplicate report generation."""
        # Generate dataset with duplicates
        allocations = self.generate_large_allocation_dataset(1000, duplicate_rate=0.3)
        result = duplicate_validator.validate_allocations(allocations)

        performance_monitor.start()
        report = duplicate_validator.generate_duplicate_report(result)
        performance_monitor.stop()

        # Report generation should be fast
        assert (
            performance_monitor.execution_time < 2.0
        ), f"Report generation took {performance_monitor.execution_time:.2f}s"

        # Verify report structure
        assert "duplicate_count" in report
        assert "duplicates" in report
        assert len(report["duplicates"]) == result.duplicate_count

        performance_monitor.report("Duplicate Report Generation")

    # ==================== Unassigned Vehicles Writer Performance Tests ====================

    def test_unassigned_writer_1000_vehicles(self, unassigned_writer, performance_monitor):
        """Test unassigned vehicles writer performance with 1000 vehicles."""
        # Generate large datasets
        unassigned_vehicles = self.generate_large_unassigned_dataset(1000)
        vehicle_log = self.generate_large_vehicle_log(1000)

        workbook = Workbook()
        allocation_date = date(2025, 1, 15)

        performance_monitor.start()
        worksheet = unassigned_writer.create_unassigned_sheet(
            workbook=workbook,
            unassigned_vehicles=unassigned_vehicles,
            vehicle_log_dict=vehicle_log,
            allocation_date=allocation_date,
        )
        performance_monitor.stop()

        # Performance assertions
        assert (
            performance_monitor.execution_time < self.MAX_UNASSIGNED_PROCESSING_TIME
        ), f"Unassigned processing took {performance_monitor.execution_time:.2f}s, expected < {self.MAX_UNASSIGNED_PROCESSING_TIME}s"

        # Functional assertions
        assert worksheet.max_row == 1001  # Header + 1000 data rows
        assert worksheet.max_column == len(unassigned_writer.COLUMNS)

        performance_monitor.report("Unassigned Vehicles Processing (1000 vehicles)")

    def test_unassigned_csv_export_performance(
        self, unassigned_writer, performance_monitor, temp_dir
    ):
        """Test CSV export performance with large dataset."""
        # Generate large dataset
        unassigned_vehicles = self.generate_large_unassigned_dataset(2000)
        vehicle_log = self.generate_large_vehicle_log(2000)

        output_path = temp_dir / "large_unassigned.csv"
        allocation_date = date(2025, 1, 15)

        performance_monitor.start()
        unassigned_writer.export_unassigned_to_csv(
            unassigned_vehicles=unassigned_vehicles,
            vehicle_log_dict=vehicle_log,
            output_path=str(output_path),
            allocation_date=allocation_date,
        )
        performance_monitor.stop()

        # Should complete quickly
        assert (
            performance_monitor.execution_time < 5.0
        ), f"CSV export took {performance_monitor.execution_time:.2f}s"

        # Verify file was created and has correct size
        assert output_path.exists()
        exported_df = pd.read_csv(output_path)
        assert len(exported_df) == 2000

        performance_monitor.report("CSV Export (2000 vehicles)")

    def test_unassigned_summary_performance(self, unassigned_writer, performance_monitor):
        """Test summary generation performance with large dataset."""
        # Generate very large dataset
        unassigned_vehicles = self.generate_large_unassigned_dataset(5000)

        performance_monitor.start()
        summary = unassigned_writer.create_unassigned_summary(unassigned_vehicles)
        performance_monitor.stop()

        # Should be very fast
        assert (
            performance_monitor.execution_time < 1.0
        ), f"Summary generation took {performance_monitor.execution_time:.2f}s"

        # Verify summary accuracy
        assert summary["total_unassigned"] == 5000
        assert summary["operational_unassigned"] + summary["non_operational_unassigned"] == 5000

        performance_monitor.report("Summary Generation (5000 vehicles)")

    # ==================== Thick Borders Performance Tests ====================

    def test_thick_borders_1000_rows_single_section(
        self, thick_border_service, performance_monitor
    ):
        """Test thick borders performance with 1000 rows in single date section."""
        # Create workbook with large single section
        workbook = Workbook()
        worksheet = workbook.create_sheet("Daily Details")

        # Add headers (simplified)
        headers = ["Date", "Route #", "Name", "Asset ID", "Van ID"] + [""] * 19
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)

        # Add 1000 rows of same date
        test_date = date(2025, 1, 15)
        for row in range(2, 1002):  # Rows 2-1001
            worksheet.cell(row=row, column=1, value=test_date)
            worksheet.cell(row=row, column=2, value=f"CX{row-1}")
            worksheet.cell(row=row, column=3, value=f"Driver_{row-1}")
            worksheet.cell(row=row, column=4, value=f"A{row-1}")
            worksheet.cell(row=row, column=5, value=f"BW{row-1}")

        performance_monitor.start()
        thick_border_service.apply_thick_borders_to_daily_details(worksheet)
        performance_monitor.stop()

        # Performance assertion
        assert (
            performance_monitor.execution_time < self.MAX_THICK_BORDERS_TIME
        ), f"Thick borders took {performance_monitor.execution_time:.2f}s, expected < {self.MAX_THICK_BORDERS_TIME}s"

        performance_monitor.report("Thick Borders (1000 rows, single section)")

    def test_thick_borders_many_small_sections(self, thick_border_service, performance_monitor):
        """Test thick borders performance with many small date sections."""
        # Create workbook with 200 sections of 5 rows each (1000 total rows)
        workbook = Workbook()
        worksheet = workbook.create_sheet("Daily Details")

        # Add headers
        headers = ["Date", "Route #", "Name", "Asset ID", "Van ID"] + [""] * 19
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)

        # Add 200 sections of 5 rows each
        base_date = date(2025, 1, 1)
        current_row = 2

        for section in range(200):
            section_date = base_date + timedelta(days=section)
            for row_in_section in range(5):
                worksheet.cell(row=current_row, column=1, value=section_date)
                worksheet.cell(row=current_row, column=2, value=f"CX{current_row}")
                worksheet.cell(row=current_row, column=3, value=f"Driver_{current_row}")
                current_row += 1

        performance_monitor.start()
        thick_border_service.apply_thick_borders_to_daily_details(worksheet)
        performance_monitor.stop()

        # Should handle many sections efficiently
        max_time = self.MAX_THICK_BORDERS_TIME * 1.5  # Slightly more time for complexity
        assert (
            performance_monitor.execution_time < max_time
        ), f"Many sections processing took {performance_monitor.execution_time:.2f}s"

        performance_monitor.report("Thick Borders (200 sections of 5 rows)")

    def test_thick_borders_append_performance(self, thick_border_service, performance_monitor):
        """Test append functionality performance with large additions."""
        # Create initial workbook
        workbook = Workbook()
        worksheet = workbook.create_sheet("Daily Details")

        # Add headers and some initial data
        headers = ["Date", "Route #", "Name"] + [""] * 21
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)

        initial_date = date(2025, 1, 15)
        for row in range(2, 52):  # 50 initial rows
            worksheet.cell(row=row, column=1, value=initial_date)
            worksheet.cell(row=row, column=2, value=f"CX{row}")

        # Apply initial borders
        thick_border_service.apply_thick_borders_to_daily_details(worksheet)

        # Now test appending 500 new rows
        new_date = date(2025, 1, 16)
        start_row = 52
        num_new_rows = 500

        for i in range(num_new_rows):
            row = start_row + i
            worksheet.cell(row=row, column=1, value=new_date)
            worksheet.cell(row=row, column=2, value=f"CX{row}")

        performance_monitor.start()
        thick_border_service.apply_thick_borders_after_append(worksheet, start_row, num_new_rows)
        performance_monitor.stop()

        # Append should be faster than full processing
        assert (
            performance_monitor.execution_time < 5.0
        ), f"Append processing took {performance_monitor.execution_time:.2f}s"

        performance_monitor.report("Thick Borders Append (500 new rows)")

    # ==================== Integration Performance Tests ====================

    def test_full_allocation_pipeline_performance(self, performance_monitor, temp_dir):
        """Test performance of full allocation pipeline with large dataset."""
        # Generate large test datasets
        day_ops_data = pd.DataFrame(
            [
                {
                    "Route Code": f"CX{i}",
                    "Service Type": "Standard Parcel - Large Van",
                    "DSP": "BWAY",
                    "Wave": f"{8 + (i % 4)}:00 AM",
                    "Staging Location": f"STG.G.{i % 10 + 1}",
                }
                for i in range(1, 1001)  # 1000 routes
            ]
        )

        vehicle_status_data = pd.DataFrame(
            [
                {
                    "Van ID": f"BW{i}",
                    "Type": "Large",
                    "Opnal? Y/N": "Y" if i <= 800 else "N",  # 800 operational vehicles
                }
                for i in range(1, 1201)  # 1200 total vehicles
            ]
        )

        daily_routes_data = pd.DataFrame(
            [{"Route Code": f"CX{i}", "Driver Name": f"Driver_{i}"} for i in range(1, 1001)]
        )

        # Create temp files
        day_ops_file = temp_dir / "large_day_ops.xlsx"
        routes_file = temp_dir / "large_routes.xlsx"
        summary_file = temp_dir / "large_summary.xlsx"

        # Save test data
        with pd.ExcelWriter(day_ops_file) as writer:
            day_ops_data.to_excel(writer, sheet_name="Solution", index=False)

        with pd.ExcelWriter(routes_file) as writer:
            daily_routes_data.to_excel(writer, sheet_name="Routes", index=False)

        with pd.ExcelWriter(summary_file) as writer:
            vehicle_status_data.to_excel(writer, sheet_name="Vehicle Status", index=False)

        # Test full allocation pipeline
        allocator = GASCompatibleAllocator()

        performance_monitor.start()

        # Load data
        allocator.load_day_of_ops(str(day_ops_file))
        allocator.load_daily_routes(str(routes_file))
        allocator.load_vehicle_status(str(summary_file))

        # Run allocation
        allocation_results, assigned_van_ids = allocator.allocate_vehicles_to_routes()

        # Validate duplicates
        validation_result = allocator.duplicate_validator.validate_allocations(allocation_results)

        # Identify unassigned
        unassigned = allocator.identify_unassigned_vehicles()

        performance_monitor.stop()

        # Performance assertion - should complete full pipeline in reasonable time
        max_pipeline_time = 30.0  # 30 seconds for full pipeline with 1000 routes
        assert (
            performance_monitor.execution_time < max_pipeline_time
        ), f"Full pipeline took {performance_monitor.execution_time:.2f}s, expected < {max_pipeline_time}s"

        # Functional assertions
        assert len(allocation_results) > 0
        assert isinstance(validation_result.duplicate_count, int)
        assert len(unassigned) > 0  # Should have unassigned vehicles

        performance_monitor.report("Full Allocation Pipeline (1000 routes)")

    # ==================== Memory Stress Tests ====================

    def test_memory_usage_under_stress(self, performance_monitor):
        """Test memory usage under stress conditions."""
        validator = DuplicateVehicleValidator()

        # Generate multiple large datasets and process sequentially
        datasets = [
            self.generate_large_allocation_dataset(1000, 0.2),
            self.generate_large_allocation_dataset(1500, 0.15),
            self.generate_large_allocation_dataset(2000, 0.1),
        ]

        performance_monitor.start()

        results = []
        for i, dataset in enumerate(datasets):
            result = validator.validate_allocations(dataset)
            results.append(result)

            # Force garbage collection between iterations
            gc.collect()

        performance_monitor.stop()

        # Memory should not grow excessively
        assert (
            performance_monitor.memory_increase < self.MAX_MEMORY_INCREASE * 2
        ), f"Memory stress test increased memory by {performance_monitor.memory_increase:.1f}MB"

        # All results should be valid
        assert len(results) == 3
        assert all(isinstance(r.duplicate_count, int) for r in results)

        performance_monitor.report("Memory Stress Test (Sequential Large Datasets)")

    # ==================== Scalability Tests ====================

    @pytest.mark.slow
    def test_scalability_10000_allocations(self, performance_monitor):
        """Scalability test with 10,000 allocations (marked as slow)."""
        validator = DuplicateVehicleValidator()

        # Generate very large dataset
        allocations = self.generate_large_allocation_dataset(10000, duplicate_rate=0.05)

        performance_monitor.start()
        result = validator.validate_allocations(allocations)
        performance_monitor.stop()

        # Should complete in reasonable time even for very large dataset
        max_time = 60.0  # 1 minute for 10K allocations
        assert (
            performance_monitor.execution_time < max_time
        ), f"10K allocation validation took {performance_monitor.execution_time:.2f}s"

        # Should not use excessive memory
        max_memory = 500  # MB
        assert (
            performance_monitor.memory_increase < max_memory
        ), f"10K allocations used {performance_monitor.memory_increase:.1f}MB"

        performance_monitor.report("Scalability Test (10,000 allocations)")

    # ==================== Benchmark Comparison Tests ====================

    def test_performance_regression_benchmark(self, performance_monitor):
        """Benchmark test to detect performance regressions."""
        # Standard benchmark dataset
        allocations = self.generate_large_allocation_dataset(1000, duplicate_rate=0.1)
        validator = DuplicateVehicleValidator()

        # Run multiple iterations for more stable results
        times = []
        for _ in range(5):
            start = time.time()
            result = validator.validate_allocations(allocations)
            end = time.time()
            times.append(end - start)

            # Verify results are consistent
            assert isinstance(result.duplicate_count, int)

        avg_time = sum(times) / len(times)
        std_dev = (sum((t - avg_time) ** 2 for t in times) / len(times)) ** 0.5

        # Performance should be consistent and fast
        assert avg_time < 2.0, f"Average time {avg_time:.3f}s exceeded threshold"
        assert std_dev < 0.5, f"Performance inconsistent (std dev: {std_dev:.3f}s)"

        print(f"\nBenchmark Results:")
        print(f"  Average Time: {avg_time:.3f}s")
        print(f"  Standard Deviation: {std_dev:.3f}s")
        print(f"  Min Time: {min(times):.3f}s")
        print(f"  Max Time: {max(times):.3f}s")
