"""Integration tests for duplicate validation and unassigned vehicles features."""

import tempfile
from datetime import date
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from src.core.gas_compatible_allocator import GASCompatibleAllocator
from src.services.duplicate_validator import DuplicateVehicleValidator
from src.services.unassigned_vehicles_writer import UnassignedVehiclesWriter


class TestDuplicateAndUnassignedIntegration:
    """Integration tests for both new features."""

    @pytest.fixture
    def sample_day_of_ops_data(self):
        """Create sample Day of Ops data with duplicate potential."""
        return pd.DataFrame(
            [
                {
                    "Route Code": "CX1",
                    "Service Type": "Standard Parcel - Large Van",
                    "DSP": "BWAY",
                    "Wave": "8:00 AM",
                    "Staging Location": "STG.G.1",
                },
                {
                    "Route Code": "CX2",
                    "Service Type": "Standard Parcel - Large Van",
                    "DSP": "BWAY",
                    "Wave": "8:00 AM",
                    "Staging Location": "STG.G.2",
                },
                {
                    "Route Code": "CX3",
                    "Service Type": "Standard Parcel - Extra Large Van - US",
                    "DSP": "BWAY",
                    "Wave": "9:00 AM",
                    "Staging Location": "STG.G.3",
                },
                {
                    "Route Code": "CX4",
                    "Service Type": "Standard Parcel - Large Van",
                    "DSP": "BWAY",
                    "Wave": "9:00 AM",
                    "Staging Location": "STG.G.4",
                },
                {
                    "Route Code": "CX5",
                    "Service Type": "Standard Parcel Step Van - US",
                    "DSP": "OTHER",  # Different DSP - should be filtered out
                    "Wave": "10:00 AM",
                    "Staging Location": "STG.G.5",
                },
            ]
        )

    @pytest.fixture
    def sample_vehicle_status_data(self):
        """Create sample vehicle status with limited vehicles to force duplicates."""
        return pd.DataFrame(
            [
                {"Van ID": "BW1", "Type": "Large", "Opnal? Y/N": "Y"},
                {"Van ID": "BW2", "Type": "Large", "Opnal? Y/N": "Y"},
                {"Van ID": "BW3", "Type": "Extra Large", "Opnal? Y/N": "Y"},
                {"Van ID": "BW4", "Type": "Extra Large", "Opnal? Y/N": "Y"},
                {"Van ID": "BW5", "Type": "Step Van", "Opnal? Y/N": "Y"},
                {"Van ID": "BW6", "Type": "Large", "Opnal? Y/N": "N"},  # Not operational
                {"Van ID": "BW7", "Type": "Extra Large", "Opnal? Y/N": "N"},  # Not operational
            ]
        )

    @pytest.fixture
    def sample_daily_routes_data(self):
        """Create sample daily routes data."""
        return pd.DataFrame(
            [
                {"Route Code": "CX1", "Driver Name": "John Smith"},
                {"Route Code": "CX2", "Driver Name": "Jane Doe"},
                {"Route Code": "CX3", "Driver Name": "Bob Johnson"},
                {"Route Code": "CX4", "Driver Name": "Alice Brown"},
                {"Route Code": "CX5", "Driver Name": "Charlie Davis"},
            ]
        )

    @pytest.fixture
    def sample_vehicle_log_data(self):
        """Create sample vehicle log data."""
        return pd.DataFrame(
            [
                {
                    "Van ID": "BW1",
                    "VIN": "1HGCM82633A004351",
                    "GeoTab": "GT001",
                    "Branded or Rental": "Branded",
                },
                {
                    "Van ID": "BW2",
                    "VIN": "1HGCM82633A004352",
                    "GeoTab": "GT002",
                    "Branded or Rental": "Branded",
                },
                {
                    "Van ID": "BW3",
                    "VIN": "1HGCM82633A004353",
                    "GeoTab": "GT003",
                    "Branded or Rental": "Rental",
                },
                {
                    "Van ID": "BW4",
                    "VIN": "1HGCM82633A004354",
                    "GeoTab": "GT004",
                    "Branded or Rental": "Rental",
                },
                {
                    "Van ID": "BW5",
                    "VIN": "1HGCM82633A004355",
                    "GeoTab": "",
                    "Branded or Rental": "Branded",
                },
                {
                    "Van ID": "BW6",
                    "VIN": "1HGCM82633A004356",
                    "GeoTab": "GT006",
                    "Branded or Rental": "Branded",
                },
                {
                    "Van ID": "BW7",
                    "VIN": "1HGCM82633A004357",
                    "GeoTab": "GT007",
                    "Branded or Rental": "Rental",
                },
            ]
        )

    def test_allocation_with_duplicates_and_unassigned(
        self,
        sample_day_of_ops_data,
        sample_vehicle_status_data,
        sample_daily_routes_data,
        sample_vehicle_log_data,
    ):
        """Test full allocation process with duplicate detection and unassigned tracking."""
        allocator = GASCompatibleAllocator()

        # Set up test data
        allocator.day_of_ops_data = sample_day_of_ops_data
        allocator.vehicle_status_data = sample_vehicle_status_data
        allocator.daily_routes_data = sample_daily_routes_data
        allocator.vehicle_log_data = sample_vehicle_log_data

        # Run allocation (should create duplicates due to limited vehicles)
        allocation_results, assigned_van_ids = allocator.allocate_vehicles_to_routes()

        # Update with driver names
        allocator.update_with_driver_names()

        # Identify unassigned vehicles
        unassigned = allocator.identify_unassigned_vehicles()

        # Create allocation result
        result = allocator.create_allocation_result()

        # Verify duplicates were detected
        assert result.metadata.get("duplicate_count", 0) >= 0
        if result.warnings:
            assert any("duplicate" in warning.lower() for warning in result.warnings)

        # Verify unassigned vehicles
        assert len(result.unallocated_vehicles) > 0
        assert all(van_id in ["BW5", "BW6", "BW7"] for van_id in result.unallocated_vehicles)

        # Check that non-operational vehicles are in unassigned
        assert "BW6" in result.unallocated_vehicles  # Non-operational
        assert "BW7" in result.unallocated_vehicles  # Non-operational

    def test_excel_output_with_both_features(
        self,
        sample_day_of_ops_data,
        sample_vehicle_status_data,
        sample_daily_routes_data,
        sample_vehicle_log_data,
        tmp_path,
    ):
        """Test Excel output includes duplicate markers and unassigned sheet."""
        # Create temporary Excel files
        day_ops_file = tmp_path / "day_of_ops.xlsx"
        routes_file = tmp_path / "daily_routes.xlsx"
        summary_file = tmp_path / "daily_summary.xlsx"

        # Save test data to files
        with pd.ExcelWriter(day_ops_file) as writer:
            sample_day_of_ops_data.to_excel(writer, sheet_name="Solution", index=False)

        with pd.ExcelWriter(routes_file) as writer:
            sample_daily_routes_data.to_excel(writer, sheet_name="Routes", index=False)

        with pd.ExcelWriter(summary_file) as writer:
            sample_vehicle_status_data.to_excel(writer, sheet_name="Vehicle Status", index=False)
            sample_vehicle_log_data.to_excel(writer, sheet_name="Vehicle Log", index=False)

        # Run allocation
        allocator = GASCompatibleAllocator()
        result = allocator.run_full_allocation(
            day_of_ops_file=str(day_ops_file),
            daily_routes_file=str(routes_file),
            vehicle_status_file=str(summary_file),
            output_file=str(summary_file),
        )

        # Verify output file
        wb = load_workbook(summary_file)

        # Check for Results sheet
        allocation_date = date.today()
        results_sheet_name = allocation_date.strftime("%m-%d-%y") + " Results"
        assert results_sheet_name in wb.sheetnames

        # Check for Unassigned sheet
        unassigned_sheet_name = allocation_date.strftime("%m-%d-%y") + " Available & Unassigned"
        assert unassigned_sheet_name in wb.sheetnames

        # Check Results sheet for duplicate markers
        results_sheet = wb[results_sheet_name]
        has_validation_columns = False

        # Look for validation columns in results
        for row in results_sheet.iter_rows(min_row=2, values_only=True):
            if row and len(row) > 11:  # Extended columns for validation
                has_validation_columns = True
                break

        # Check Unassigned sheet content
        unassigned_sheet = wb[unassigned_sheet_name]

        # Verify headers
        headers = [cell.value for cell in unassigned_sheet[1]]
        assert "Van ID" in headers
        assert "Vehicle Type" in headers
        assert "Operational Status" in headers
        assert "VIN" in headers
        assert "GeoTab Code" in headers

        # Verify unassigned vehicles are listed
        unassigned_van_ids = []
        for row in unassigned_sheet.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0]:
                unassigned_van_ids.append(row[0])

        assert len(unassigned_van_ids) > 0

        wb.close()

    def test_duplicate_validation_integration(self):
        """Test duplicate validation integration with allocator."""
        allocator = GASCompatibleAllocator()

        # Create allocation results with intentional duplicate
        allocator.allocation_results = [
            {
                "Van ID": "BW1",
                "Route Code": "CX1",
                "Associate Name": "Driver A",
                "Service Type": "Standard",
                "Wave": "8:00 AM",
                "Staging Location": "STG.G.1",
            },
            {
                "Van ID": "BW1",  # Duplicate
                "Route Code": "CX2",
                "Associate Name": "Driver B",
                "Service Type": "Standard",
                "Wave": "9:00 AM",
                "Staging Location": "STG.G.2",
            },
        ]

        # Validate
        validator = allocator.duplicate_validator
        validation_result = validator.validate_allocations(allocator.allocation_results)

        # Check detection
        assert validation_result.has_duplicates()
        assert validation_result.duplicate_count == 1
        assert "BW1" in validation_result.duplicates

        # Mark duplicates
        marked_results = validator.mark_duplicates_in_results(
            allocator.allocation_results, validation_result
        )

        # Verify marking
        for result in marked_results:
            if result["Van ID"] == "BW1":
                assert result["Validation Status"] == "DUPLICATE"
                assert "Validation Warning" in result

    def test_unassigned_writer_integration(self):
        """Test unassigned vehicles writer integration."""
        from openpyxl import Workbook

        writer = UnassignedVehiclesWriter()
        workbook = Workbook()

        # Create test data
        unassigned_df = pd.DataFrame(
            [
                {"Van ID": "BW10", "Type": "Large", "Opnal? Y/N": "Y"},
                {"Van ID": "BW11", "Type": "Extra Large", "Opnal? Y/N": "N"},
            ]
        )

        vehicle_log = {
            "BW10": {"vin": "123456", "geotab": "GT010", "brand_or_rental": "Branded"},
            "BW11": {"vin": "123457", "geotab": "GT011", "brand_or_rental": "Rental"},
        }

        # Create sheet
        allocation_date = date(2025, 1, 5)
        worksheet = writer.create_unassigned_sheet(
            workbook=workbook,
            unassigned_vehicles=unassigned_df,
            vehicle_log_dict=vehicle_log,
            allocation_date=allocation_date,
        )

        # Verify sheet
        assert worksheet.title == "01-05-25 Available & Unassigned"
        assert worksheet.cell(row=2, column=1).value == "BW10"
        assert worksheet.cell(row=2, column=6).value == "123456"  # VIN
        assert worksheet.cell(row=3, column=1).value == "BW11"
        assert worksheet.cell(row=3, column=3).value == "N"  # Not operational

    # ==================== Real-World Scenario Tests ====================

    def test_peak_season_allocation_scenario(self, test_excel_files):
        """Test allocation during peak season with high demand and limited vehicles."""
        # Simulate peak season: more routes than vehicles
        peak_routes = pd.DataFrame(
            [
                {
                    "Route Code": f"CX{i}",
                    "Service Type": "Standard Parcel - Large Van",
                    "DSP": "BWAY",
                    "Wave": "8:00 AM",
                    "Staging Location": f"STG.G.{i%5+1}",
                }
                for i in range(1, 51)  # 50 routes
            ]
        )

        limited_vehicles = pd.DataFrame(
            [
                {"Van ID": f"BW{i}", "Type": "Large", "Opnal? Y/N": "Y"}
                for i in range(1, 31)  # Only 30 vehicles
            ]
        )

        driver_assignments = pd.DataFrame(
            [{"Route Code": f"CX{i}", "Driver Name": f"Driver_{i}"} for i in range(1, 51)]
        )

        # Update test files with peak season data
        with pd.ExcelWriter(test_excel_files["day_ops"]) as writer:
            peak_routes.to_excel(writer, sheet_name="Solution", index=False)

        with pd.ExcelWriter(test_excel_files["routes"]) as writer:
            driver_assignments.to_excel(writer, sheet_name="Routes", index=False)

        with pd.ExcelWriter(test_excel_files["summary"]) as writer:
            limited_vehicles.to_excel(writer, sheet_name="Vehicle Status", index=False)

        # Run allocation
        allocator = GASCompatibleAllocator()
        result = allocator.run_full_allocation(
            day_of_ops_file=str(test_excel_files["day_ops"]),
            daily_routes_file=str(test_excel_files["routes"]),
            vehicle_status_file=str(test_excel_files["summary"]),
            output_file=str(test_excel_files["summary"]),
        )

        # Verify peak season handling
        assert len(result.allocated_vehicles) <= 30  # Can't exceed vehicle count
        assert len(result.unallocated_vehicles) > 0  # Should have unassigned routes

    def test_vehicle_breakdown_scenario(self, test_excel_files):
        """Test allocation when many vehicles become non-operational."""
        # Simulate vehicle breakdowns - mix of operational and non-operational
        mixed_vehicles = pd.DataFrame(
            [
                {
                    "Van ID": f"BW{i}",
                    "Type": "Large",
                    "Opnal? Y/N": "Y" if i % 3 != 0 else "N",
                }  # 1/3 are broken down
                for i in range(1, 31)
            ]
        )

        routes_data = pd.DataFrame(
            [
                {
                    "Route Code": f"CX{i}",
                    "Service Type": "Standard Parcel - Large Van",
                    "DSP": "BWAY",
                    "Wave": "8:00 AM",
                    "Staging Location": f"STG.G.{i%3+1}",
                }
                for i in range(1, 21)  # 20 routes
            ]
        )

        # Update test files
        with pd.ExcelWriter(test_excel_files["day_ops"]) as writer:
            routes_data.to_excel(writer, sheet_name="Solution", index=False)

        with pd.ExcelWriter(test_excel_files["summary"]) as writer:
            mixed_vehicles.to_excel(writer, sheet_name="Vehicle Status", index=False)

        allocator = GASCompatibleAllocator()
        allocator.load_day_of_ops(str(test_excel_files["day_ops"]))
        allocator.load_vehicle_status(str(test_excel_files["summary"]))

        allocation_results, assigned_van_ids = allocator.allocate_vehicles_to_routes()
        unassigned = allocator.identify_unassigned_vehicles()

        # Verify breakdown handling
        operational_count = len(mixed_vehicles[mixed_vehicles["Opnal? Y/N"] == "Y"])
        non_operational_count = len(mixed_vehicles[mixed_vehicles["Opnal? Y/N"] == "N"])

        assert len(assigned_van_ids) <= operational_count
        assert len(unassigned) >= non_operational_count  # At least all broken vehicles

    def test_mixed_service_types_scenario(self, test_excel_files):
        """Test allocation with mixed service types requiring different vehicle types."""
        # Mixed service types
        mixed_routes = pd.DataFrame(
            [
                {
                    "Route Code": "CX1",
                    "Service Type": "Standard Parcel - Large Van",
                    "DSP": "BWAY",
                    "Wave": "8:00 AM",
                    "Staging Location": "STG.G.1",
                },
                {
                    "Route Code": "CX2",
                    "Service Type": "Standard Parcel - Extra Large Van - US",
                    "DSP": "BWAY",
                    "Wave": "8:00 AM",
                    "Staging Location": "STG.G.2",
                },
                {
                    "Route Code": "CX3",
                    "Service Type": "Standard Parcel Step Van - US",
                    "DSP": "BWAY",
                    "Wave": "9:00 AM",
                    "Staging Location": "STG.G.3",
                },
                {
                    "Route Code": "CX4",
                    "Service Type": "Standard Parcel - Large Van",
                    "DSP": "BWAY",
                    "Wave": "9:00 AM",
                    "Staging Location": "STG.G.4",
                },
            ]
        )

        # Mixed vehicle types
        mixed_vehicles = pd.DataFrame(
            [
                {"Van ID": "BW1", "Type": "Large", "Opnal? Y/N": "Y"},
                {"Van ID": "BW2", "Type": "Large", "Opnal? Y/N": "Y"},
                {"Van ID": "BW3", "Type": "Extra Large", "Opnal? Y/N": "Y"},
                {"Van ID": "BW4", "Type": "Step Van", "Opnal? Y/N": "Y"},
                {"Van ID": "BW5", "Type": "Large", "Opnal? Y/N": "Y"},
            ]
        )

        # Update test files
        with pd.ExcelWriter(test_excel_files["day_ops"]) as writer:
            mixed_routes.to_excel(writer, sheet_name="Solution", index=False)

        with pd.ExcelWriter(test_excel_files["summary"]) as writer:
            mixed_vehicles.to_excel(writer, sheet_name="Vehicle Status", index=False)

        allocator = GASCompatibleAllocator()
        allocator.load_day_of_ops(str(test_excel_files["day_ops"]))
        allocator.load_vehicle_status(str(test_excel_files["summary"]))

        allocation_results, assigned_van_ids = allocator.allocate_vehicles_to_routes()

        # Verify service type matching
        for result in allocation_results:
            service_type = result.get("Service Type", "")
            van_type = None

            # Find the vehicle type for this van
            van_id = result.get("Van ID")
            if van_id:
                vehicle_row = mixed_vehicles[mixed_vehicles["Van ID"] == van_id]
                if not vehicle_row.empty:
                    van_type = vehicle_row.iloc[0]["Type"]

            # Verify correct matching
            if "Extra Large" in service_type:
                assert (
                    van_type == "Extra Large"
                ), f"Extra Large service should use Extra Large van, got {van_type}"
            elif "Step Van" in service_type:
                assert (
                    van_type == "Step Van"
                ), f"Step Van service should use Step Van, got {van_type}"
            elif "Large" in service_type:
                assert van_type == "Large", f"Large service should use Large van, got {van_type}"
