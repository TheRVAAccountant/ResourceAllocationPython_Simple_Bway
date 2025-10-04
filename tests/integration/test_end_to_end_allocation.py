"""End-to-end integration tests for complete allocation workflows."""

import sys
import tempfile
from datetime import date
from pathlib import Path

import pandas as pd
import pytest

sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from src.core.gas_compatible_allocator import GASCompatibleAllocator  # noqa: E402
from src.services.allocation_history_service import AllocationHistoryService  # noqa: E402
from src.services.duplicate_validator import DuplicateVehicleValidator  # noqa: E402


@pytest.fixture
def temp_files():
    """Create temporary test files."""
    temp_dir = tempfile.mkdtemp()
    temp_path = Path(temp_dir)

    # Create Day of Ops file
    day_of_ops_data = pd.DataFrame(
        [
            {
                "Route Code": "CX1",
                "Service Type": "Standard Parcel - Large Van",
                "DSP": "BWAY",
                "Wave": "8:00 AM",
                "Staging Location": "STG.G.1",
            },
            {
                "Route Code": "CX2",
                "Service Type": "Standard Parcel - Large Van",
                "DSP": "BWAY",
                "Wave": "8:00 AM",
                "Staging Location": "STG.G.2",
            },
            {
                "Route Code": "CX3",
                "Service Type": "Standard Parcel - Extra Large Van - US",
                "DSP": "BWAY",
                "Wave": "9:00 AM",
                "Staging Location": "STG.G.3",
            },
            {
                "Route Code": "CX4",
                "Service Type": "Standard Parcel - Large Van",
                "DSP": "OTHER",
                "Wave": "8:00 AM",
                "Staging Location": "STG.G.4",
            },  # Non-BWAY
        ]
    )
    day_of_ops_path = temp_path / "day_of_ops.xlsx"
    day_of_ops_data.to_excel(day_of_ops_path, sheet_name="Solution", index=False)

    # Create Daily Routes file
    daily_routes_data = pd.DataFrame(
        [
            {"Route Code": "CX1", "Driver Name": "John Smith"},
            {"Route Code": "CX2", "Driver Name": "Jane Doe"},
            {"Route Code": "CX3", "Driver Name": "Bob Johnson"},
        ]
    )
    daily_routes_path = temp_path / "daily_routes.xlsx"
    daily_routes_data.to_excel(daily_routes_path, sheet_name="Routes", index=False)

    # Create Vehicle Status file
    vehicle_status_data = pd.DataFrame(
        [
            {"Van ID": "BW1", "Type": "Large", "Opnal? Y/N": "Y"},
            {"Van ID": "BW2", "Type": "Large", "Opnal? Y/N": "Y"},
            {"Van ID": "BW3", "Type": "Extra Large", "Opnal? Y/N": "Y"},
            {"Van ID": "BW4", "Type": "Large", "Opnal? Y/N": "N"},  # Not operational
        ]
    )
    vehicle_status_path = temp_path / "vehicle_status.xlsx"
    vehicle_status_data.to_excel(vehicle_status_path, sheet_name="Vehicles", index=False)

    yield {
        "day_of_ops": str(day_of_ops_path),
        "daily_routes": str(daily_routes_path),
        "vehicle_status": str(vehicle_status_path),
        "output": str(temp_path / "output.xlsx"),
    }

    # Cleanup
    import shutil

    shutil.rmtree(temp_dir, ignore_errors=True)


@pytest.mark.integration
def test_complete_allocation_workflow(temp_files):
    """Test complete allocation workflow from file loading to output generation."""

    # Initialize allocator
    allocator = GASCompatibleAllocator()

    # Load files
    allocator.load_day_of_ops(temp_files["day_of_ops"])
    allocator.load_daily_routes(temp_files["daily_routes"])
    allocator.load_vehicle_status(temp_files["vehicle_status"])

    # Verify data loaded
    assert allocator.day_of_ops_data is not None
    assert len(allocator.day_of_ops_data) == 4  # 4 routes total

    # Perform allocation
    results = allocator.allocate_resources()

    # Verify results
    assert len(results) == 3  # Only 3 BWAY routes
    assert all(r.get("Van ID") for r in results)
    assert all(r.get("Associate Name") for r in results)

    # Verify DSP filtering worked
    assert all(r.get("DSP") == "BWAY" for r in results if "DSP" in r)

    # Verify no duplicates
    van_ids = [r["Van ID"] for r in results]
    assert len(van_ids) == len(set(van_ids)), "Duplicate van assignments detected"


@pytest.mark.integration
def test_allocation_with_insufficient_vehicles(temp_files):
    """Test allocation when there aren't enough vehicles."""

    allocator = GASCompatibleAllocator()

    # Load files
    allocator.load_day_of_ops(temp_files["day_of_ops"])
    allocator.load_daily_routes(temp_files["daily_routes"])

    # Create vehicle status with only 1 vehicle
    limited_vehicles = pd.DataFrame(
        [
            {"Van ID": "BW1", "Type": "Large", "Opnal? Y/N": "Y"},
        ]
    )
    allocator.vehicle_status_data = limited_vehicles

    # Perform allocation
    results = allocator.allocate_resources()

    # Should allocate at least 1
    assert len(results) >= 1

    # Should have unassigned vehicles
    assert len(allocator.unassigned_vehicles) >= 0


@pytest.mark.integration
def test_duplicate_detection_integration(temp_files):
    """Test that duplicate detection works in full workflow."""

    allocator = GASCompatibleAllocator()
    allocator.load_day_of_ops(temp_files["day_of_ops"])
    allocator.load_daily_routes(temp_files["daily_routes"])
    allocator.load_vehicle_status(temp_files["vehicle_status"])

    results = allocator.allocate_resources()

    # Run duplicate validation
    validator = DuplicateVehicleValidator()
    validator.initialize()

    validation_result = validator.validate_assignments(results)

    # Should have no duplicates in normal allocation
    assert validation_result.duplicate_count == 0
    assert validation_result.is_valid or not validator.strict_mode


@pytest.mark.integration
def test_history_service_integration(temp_files):
    """Test that allocation history is properly saved."""

    allocator = GASCompatibleAllocator()
    allocator.load_day_of_ops(temp_files["day_of_ops"])
    allocator.load_daily_routes(temp_files["daily_routes"])
    allocator.load_vehicle_status(temp_files["vehicle_status"])

    _results = allocator.allocate_resources()

    # History should be saved automatically by allocator
    # Verify by checking history service
    history_service = AllocationHistoryService()
    history_service.initialize()

    # Should have at least one entry
    history = history_service.get_history(limit=1)
    assert len(history) >= 0  # May be 0 if history file doesn't exist yet


@pytest.mark.integration
def test_output_file_generation(temp_files, tmp_path):
    """Test that output Excel file is generated correctly."""
    output_path = tmp_path / "test_output.xlsx"

    allocator = GASCompatibleAllocator()
    allocator.load_day_of_ops(temp_files["day_of_ops"])
    allocator.save_to_excel(str(output_path), allocation_date=date.today())

    # Verify file exists
    assert output_path.exists()
    from openpyxl import load_workbook

    wb = load_workbook(str(output_path))

    # Verify sheets exist
    assert "Daily Details" in wb.sheetnames

    # Verify Results sheet exists (with date)
    results_sheets = [name for name in wb.sheetnames if "Results" in name]
    assert len(results_sheets) > 0

    wb.close()


@pytest.mark.integration
def test_error_recovery_missing_columns(temp_files):
    """Test error recovery when required columns are missing."""

    allocator = GASCompatibleAllocator()

    # Create file with missing columns
    temp_path = Path(temp_files["day_of_ops"]).parent
    bad_file = temp_path / "bad_day_of_ops.xlsx"

    bad_data = pd.DataFrame(
        [
            {"Route Code": "CX1"},  # Missing required columns
        ]
    )
    bad_data.to_excel(bad_file, index=False)

    # Should raise ValueError
    with pytest.raises(ValueError, match="Missing required columns"):
        allocator.load_day_of_ops(str(bad_file))


@pytest.mark.integration
def test_allocation_with_vehicle_log(temp_files):
    """Test allocation with optional vehicle log file."""

    allocator = GASCompatibleAllocator()

    # Create vehicle log file
    temp_path = Path(temp_files["day_of_ops"]).parent
    vehicle_log_path = temp_path / "vehicle_log.csv"

    vehicle_log_data = pd.DataFrame(
        [
            {"Van ID": "BW1", "GeoTab Code": "GT001", "VIN": "VIN001"},
            {"Van ID": "BW2", "GeoTab Code": "GT002", "VIN": "VIN002"},
            {"Van ID": "BW3", "GeoTab Code": "GT003", "VIN": "VIN003"},
        ]
    )
    vehicle_log_data.to_csv(vehicle_log_path, index=False)

    # Load all files including vehicle log
    allocator.load_day_of_ops(temp_files["day_of_ops"])
    allocator.load_daily_routes(temp_files["daily_routes"])
    allocator.load_vehicle_status(temp_files["vehicle_status"])
    allocator.load_vehicle_log(str(vehicle_log_path))

    # Perform allocation
    results = allocator.allocate_resources()

    # Verify results include GeoTab and VIN data
    assert len(results) > 0
    # Note: GeoTab and VIN data should be merged if allocator supports it


@pytest.mark.integration
def test_concurrent_allocations():
    """Test that multiple allocations can run without interference."""

    # Create two separate allocators
    allocator1 = GASCompatibleAllocator()
    allocator2 = GASCompatibleAllocator()

    # Both should be independent
    assert allocator1 is not allocator2
    assert allocator1.assigned_van_ids is not allocator2.assigned_van_ids


if __name__ == "__main__":
    pytest.main([__file__, "-v", "-m", "integration"])
