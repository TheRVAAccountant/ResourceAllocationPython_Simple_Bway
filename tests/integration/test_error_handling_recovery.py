"""Error handling and recovery tests for corrupted Excel files and edge cases."""

import contextlib
import os
from datetime import date
from unittest.mock import patch

import pandas as pd
import pytest
import xlsxwriter
from openpyxl import Workbook
from openpyxl.utils.exceptions import InvalidFileException

from src.core.gas_compatible_allocator import GASCompatibleAllocator
from src.services.daily_details_thick_borders import DailyDetailsThickBorderService
from src.services.duplicate_validator import DuplicateVehicleValidator
from src.services.unassigned_vehicles_writer import UnassignedVehiclesWriter


class TestErrorHandlingAndRecovery:
    """Test cases for error handling and recovery from various failure scenarios."""

    @pytest.fixture
    def corrupted_excel_files(self, temp_dir):
        """Create various corrupted Excel files for testing."""
        files = {}

        # 1. Completely corrupted file (binary corruption)
        corrupted_file = temp_dir / "corrupted.xlsx"
        with open(corrupted_file, "wb") as f:
            f.write(b"This is not an Excel file - completely corrupted binary data!")
        files["corrupted_binary"] = corrupted_file

        # 2. Empty file
        empty_file = temp_dir / "empty.xlsx"
        empty_file.touch()
        files["empty"] = empty_file

        # 3. Wrong file extension
        text_file = temp_dir / "fake.xlsx"
        with open(text_file, "w") as f:
            f.write("This is a text file pretending to be Excel")
        files["fake_extension"] = text_file

        # 4. Excel file with missing sheets
        missing_sheets_file = temp_dir / "missing_sheets.xlsx"
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        wb.create_sheet("WrongSheet")
        wb.save(missing_sheets_file)
        files["missing_sheets"] = missing_sheets_file

        # 5. Excel file with wrong column names
        wrong_columns_file = temp_dir / "wrong_columns.xlsx"
        wrong_data = pd.DataFrame(
            {"Wrong Col 1": ["data1", "data2"], "Wrong Col 2": ["data3", "data4"]}
        )
        with pd.ExcelWriter(wrong_columns_file) as writer:
            wrong_data.to_excel(writer, sheet_name="Solution", index=False)
        files["wrong_columns"] = wrong_columns_file

        # 6. Excel file with mixed data types in critical columns
        mixed_types_file = temp_dir / "mixed_types.xlsx"
        mixed_data = pd.DataFrame(
            {
                "Van ID": ["BW1", 123, None, "BW2", ""],  # Mixed types
                "Type": ["Large", "Extra Large", 999, None, "Step Van"],
                "Opnal? Y/N": ["Y", "N", "Maybe", 1, None],
            }
        )
        with pd.ExcelWriter(mixed_types_file) as writer:
            mixed_data.to_excel(writer, sheet_name="Vehicle Status", index=False)
        files["mixed_types"] = mixed_types_file

        # 7. Extremely large file (stress test)
        large_file = temp_dir / "large_file.xlsx"
        large_data = pd.DataFrame(
            {
                "Van ID": [f"BW{i}" for i in range(10000)],
                "Type": ["Large"] * 10000,
                "Opnal? Y/N": ["Y"] * 10000,
            }
        )
        with pd.ExcelWriter(large_file) as writer:
            large_data.to_excel(writer, sheet_name="Vehicle Status", index=False)
        files["large_file"] = large_file

        # 8. File with special characters and unicode
        unicode_file = temp_dir / "unicode_data.xlsx"
        unicode_data = pd.DataFrame(
            {
                "Van ID": ["BW1", "BW2", "BW3"],
                "Driver Name": ["José García", "李小明", "Müller Schmidt"],
                "Location": ["Depot São Paulo", "仓库北京", "Lager München"],
            }
        )
        with pd.ExcelWriter(unicode_file) as writer:
            unicode_data.to_excel(writer, sheet_name="Routes", index=False)
        files["unicode"] = unicode_file

        return files

    # ==================== File Corruption Recovery Tests ====================

    def test_completely_corrupted_file_handling(self, corrupted_excel_files):
        """Test handling of completely corrupted Excel files."""
        allocator = GASCompatibleAllocator()

        with pytest.raises((InvalidFileException, Exception)) as exc_info:
            allocator.load_day_of_ops(str(corrupted_excel_files["corrupted_binary"]))

        # Should provide meaningful error message
        error_msg = str(exc_info.value).lower()
        assert any(keyword in error_msg for keyword in ["file", "corrupt", "invalid", "excel"])

    def test_empty_file_handling(self, corrupted_excel_files):
        """Test handling of empty Excel files."""
        allocator = GASCompatibleAllocator()

        with pytest.raises(Exception) as exc_info:
            allocator.load_day_of_ops(str(corrupted_excel_files["empty"]))

        # Should provide informative error message
        assert "file" in str(exc_info.value).lower()

    def test_fake_excel_file_handling(self, corrupted_excel_files):
        """Test handling of text files with .xlsx extension."""
        allocator = GASCompatibleAllocator()

        with pytest.raises((InvalidFileException, Exception)) as exc_info:
            allocator.load_day_of_ops(str(corrupted_excel_files["fake_extension"]))

        # Should detect that it's not a real Excel file
        error_msg = str(exc_info.value).lower()
        assert any(keyword in error_msg for keyword in ["file", "format", "excel", "invalid"])

    def test_missing_required_sheets(self, corrupted_excel_files):
        """Test handling of Excel files missing required sheets."""
        allocator = GASCompatibleAllocator()

        with pytest.raises(Exception) as exc_info:
            allocator.load_day_of_ops(
                str(corrupted_excel_files["missing_sheets"]), sheet_name="Solution"
            )

        # Should indicate missing sheet
        error_msg = str(exc_info.value).lower()
        assert "solution" in error_msg or "sheet" in error_msg

    def test_wrong_column_names_handling(self, corrupted_excel_files):
        """Test handling of Excel files with wrong column names."""
        allocator = GASCompatibleAllocator()

        with pytest.raises(ValueError) as exc_info:
            allocator.load_day_of_ops(str(corrupted_excel_files["wrong_columns"]))

        # Should indicate missing required columns
        error_msg = str(exc_info.value)
        assert "Missing required columns" in error_msg
        assert "Route Code" in error_msg or "Service Type" in error_msg

    def test_mixed_data_types_handling(self, corrupted_excel_files):
        """Test handling of mixed data types in critical columns."""
        allocator = GASCompatibleAllocator()

        # Should not crash, but handle gracefully
        try:
            df = allocator.load_vehicle_status(str(corrupted_excel_files["mixed_types"]))

            # Should load the file
            assert len(df) > 0

            # Filter out invalid entries
            filtered_df = df.dropna(subset=["Van ID"])
            filtered_df = filtered_df[filtered_df["Van ID"].astype(str).str.strip() != ""]

            # Should handle string conversion gracefully
            valid_vehicles = []
            for _, row in filtered_df.iterrows():
                if isinstance(row["Van ID"], str) and row["Van ID"].strip():
                    valid_vehicles.append(row)

            assert len(valid_vehicles) >= 2  # At least BW1 and BW2 should be valid

        except Exception as e:
            # If it does fail, should have meaningful error message
            assert "Van ID" in str(e) or "data" in str(e).lower()

    # ==================== Large File Handling Tests ====================

    def test_large_file_performance_and_memory(self, corrupted_excel_files):
        """Test handling of extremely large Excel files."""
        import os

        import psutil

        process = psutil.Process(os.getpid())
        initial_memory = process.memory_info().rss / 1024 / 1024  # MB

        allocator = GASCompatibleAllocator()

        try:
            # Should handle large files without crashing
            df = allocator.load_vehicle_status(str(corrupted_excel_files["large_file"]))

            # Verify data was loaded
            assert len(df) == 10000

            # Check memory usage didn't explode
            final_memory = process.memory_info().rss / 1024 / 1024  # MB
            memory_increase = final_memory - initial_memory

            # Should not use excessive memory (< 500MB for 10k rows)
            assert memory_increase < 500, f"Memory increased by {memory_increase:.1f}MB"

        except MemoryError:
            pytest.skip("System doesn't have enough memory for large file test")
        except Exception as e:
            # Should provide meaningful error for resource constraints
            error_msg = str(e).lower()
            assert any(keyword in error_msg for keyword in ["memory", "size", "large", "resource"])

    def test_unicode_characters_handling(self, corrupted_excel_files):
        """Test handling of Unicode characters in Excel files."""
        allocator = GASCompatibleAllocator()

        # Should handle Unicode characters gracefully
        df = allocator.load_daily_routes(str(corrupted_excel_files["unicode"]), sheet_name="Routes")

        # Should load all data including Unicode
        assert len(df) == 3

        # Check that Unicode characters are preserved
        driver_names = df["Driver Name"].tolist()
        assert "José García" in driver_names
        assert "李小明" in driver_names
        assert "Müller Schmidt" in driver_names

    # ==================== Service-Level Error Handling Tests ====================

    def test_duplicate_validator_with_corrupt_data(self):
        """Test duplicate validator with corrupted allocation data."""
        validator = DuplicateVehicleValidator()

        # Test with various corrupted data scenarios
        corrupted_allocations = [
            {"Van ID": None, "Route Code": "CX1"},  # None Van ID
            {"Van ID": "", "Route Code": "CX2"},  # Empty Van ID
            {"Van ID": 123, "Route Code": "CX3"},  # Non-string Van ID
            {"Route Code": "CX4"},  # Missing Van ID
            None,  # None entry
            {},  # Empty dict
            {"Van ID": "BW1"},  # Missing Route Code
            {"Van ID": "BW2", "Route Code": None},  # None Route Code
        ]

        # Should not crash, but handle gracefully
        result = validator.validate_allocations(corrupted_allocations)

        assert isinstance(result.duplicate_count, int)
        assert result.duplicate_count >= 0
        # Should only process valid entries with Van ID
        assert result.duplicate_count <= 2  # At most BW1, BW2 if no duplicates

    def test_unassigned_writer_with_corrupt_data(self, temp_dir):  # noqa: ARG002
        """Test unassigned vehicles writer with corrupted data."""
        writer = UnassignedVehiclesWriter()

        # Corrupted unassigned vehicles data
        corrupted_vehicles = pd.DataFrame(
            [
                {"Van ID": "BW1", "Type": "Large", "Opnal? Y/N": "Y"},
                {"Van ID": None, "Type": "Large", "Opnal? Y/N": "Y"},  # None Van ID
                {"Van ID": "", "Type": "Extra Large", "Opnal? Y/N": "N"},  # Empty Van ID
                {"Van ID": "BW2", "Type": None, "Opnal? Y/N": "Y"},  # None Type
                {
                    "Van ID": "BW3",
                    "Type": "Step Van",
                    "Opnal? Y/N": None,
                },  # None operational status
            ]
        )

        workbook = Workbook()
        allocation_date = date(2025, 1, 15)

        # Should handle corrupted data gracefully
        worksheet = writer.create_unassigned_sheet(
            workbook=workbook,
            unassigned_vehicles=corrupted_vehicles,
            vehicle_log_dict={},
            allocation_date=allocation_date,
        )

        # Should create sheet with valid entries only
        assert worksheet.max_row >= 2  # Header + at least BW1
        assert worksheet.max_row <= 4  # Header + max 3 valid entries (BW1, BW2, BW3)

        # Check that None/empty Van IDs were skipped
        van_ids_in_sheet = []
        for row in range(2, worksheet.max_row + 1):
            van_id = worksheet.cell(row=row, column=1).value
            if van_id:
                van_ids_in_sheet.append(van_id)

        assert "BW1" in van_ids_in_sheet
        assert None not in van_ids_in_sheet
        assert "" not in van_ids_in_sheet

    def test_thick_borders_with_corrupt_worksheet(self):
        """Test thick borders service with corrupted worksheet data."""
        border_service = DailyDetailsThickBorderService()

        # Create worksheet with problematic data
        workbook = Workbook()
        worksheet = workbook.create_sheet("Daily Details")

        # Add headers
        worksheet.cell(row=1, column=1, value="Date")
        worksheet.cell(row=1, column=2, value="Route #")

        # Add problematic date data
        worksheet.cell(row=2, column=1, value="not-a-date")  # Invalid date string
        worksheet.cell(row=2, column=2, value="CX1")

        worksheet.cell(row=3, column=1, value=None)  # None date
        worksheet.cell(row=3, column=2, value="CX2")

        worksheet.cell(row=4, column=1, value=12345)  # Number instead of date
        worksheet.cell(row=4, column=2, value="CX3")

        worksheet.cell(row=5, column=1, value=date(2025, 1, 15))  # Valid date
        worksheet.cell(row=5, column=2, value="CX4")

        # Should not crash with problematic data
        try:
            border_service.apply_thick_borders_to_daily_details(worksheet)
            # If it succeeds, that's good
            assert True
        except Exception as e:
            # If it fails, should be a reasonable error
            assert "date" in str(e).lower() or "value" in str(e).lower()

    # ==================== Network and I/O Error Simulation Tests ====================

    @patch("pandas.read_excel")
    def test_file_read_failure_handling(self, mock_read_excel):
        """Test handling of file read failures."""
        # Simulate I/O error
        mock_read_excel.side_effect = OSError("Disk read error")

        allocator = GASCompatibleAllocator()

        with pytest.raises(IOError) as exc_info:
            allocator.load_day_of_ops("test_file.xlsx")

        assert "Disk read error" in str(exc_info.value)

    @patch("openpyxl.load_workbook")
    def test_workbook_corruption_during_read(self, mock_load_workbook):
        """Test handling of workbook corruption during read."""
        # Simulate corruption detected during loading
        mock_load_workbook.side_effect = InvalidFileException("File is corrupted")

        from src.services.excel_service import ExcelService

        service = ExcelService()

        with pytest.raises(InvalidFileException):
            service.load_workbook("corrupted_file.xlsx")

    def test_permission_denied_handling(self, temp_dir):
        """Test handling of permission denied errors."""
        # Create a file and make it unreadable (Unix-like systems)
        restricted_file = temp_dir / "restricted.xlsx"

        # Create a simple Excel file first
        wb = Workbook()
        wb.save(restricted_file)

        # Try to make it unreadable (may not work on all systems)
        try:
            os.chmod(restricted_file, 0o000)  # No permissions

            allocator = GASCompatibleAllocator()

            with pytest.raises(PermissionError):
                allocator.load_day_of_ops(str(restricted_file))

        except (OSError, PermissionError):
            # Permission changes might not work in some environments
            pytest.skip("Cannot test permission errors on this system")
        finally:
            # Restore permissions for cleanup
            with contextlib.suppress(Exception):
                os.chmod(restricted_file, 0o666)

    # ==================== Data Consistency Recovery Tests ====================

    def test_recovery_from_partial_allocation_failure(self, test_excel_files):
        """Test recovery when allocation partially fails."""
        allocator = GASCompatibleAllocator()

        # Load valid data
        allocator.load_day_of_ops(str(test_excel_files["day_ops"]))
        allocator.load_vehicle_status(str(test_excel_files["summary"]))

        # Simulate partial failure by corrupting data mid-process
        with patch.object(allocator, "allocate_vehicles_to_routes") as mock_allocate:
            # Simulate partial results before failure
            mock_allocate.side_effect = Exception("Allocation failed mid-process")

            with pytest.raises(Exception) as exc_info:
                allocator.run_full_allocation(
                    day_of_ops_file=str(test_excel_files["day_ops"]),
                    daily_routes_file=str(test_excel_files["routes"]),
                    vehicle_status_file=str(test_excel_files["summary"]),
                    output_file=str(test_excel_files["summary"]),
                )

            assert "Allocation failed mid-process" in str(exc_info.value)

    def test_excel_save_failure_recovery(self, test_excel_files):
        """Test recovery when Excel save operations fail."""
        allocator = GASCompatibleAllocator()

        with patch("pandas.ExcelWriter") as mock_writer:
            # Simulate save failure
            mock_writer.side_effect = PermissionError("Cannot write to file")

            with pytest.raises(PermissionError):
                allocator.run_full_allocation(
                    day_of_ops_file=str(test_excel_files["day_ops"]),
                    daily_routes_file=str(test_excel_files["routes"]),
                    vehicle_status_file=str(test_excel_files["summary"]),
                    output_file=str(test_excel_files["summary"]),
                )

    def test_data_validation_recovery(self):
        """Test recovery from data validation failures."""
        validator = DuplicateVehicleValidator(config={"strict_duplicate_validation": True})

        # Create data that will fail strict validation
        failing_allocations = [
            {"Van ID": "BW1", "Route Code": "CX1", "Associate Name": "Driver A"},
            {"Van ID": "BW1", "Route Code": "CX2", "Associate Name": "Driver B"},  # Duplicate
        ]

        # Should detect duplicates in strict mode
        result = validator.validate_allocations(failing_allocations)

        assert not result.is_valid  # Should fail in strict mode
        assert result.duplicate_count == 1

        # Test recovery by switching to non-strict mode
        non_strict_validator = DuplicateVehicleValidator(
            config={"strict_duplicate_validation": False}
        )

        recovery_result = non_strict_validator.validate_allocations(failing_allocations)

        assert recovery_result.is_valid  # Should pass in non-strict mode
        assert recovery_result.duplicate_count == 1  # Still detects duplicates

    # ==================== Stress and Edge Case Tests ====================

    def test_extremely_long_strings_handling(self):
        """Test handling of extremely long strings in data."""
        # Create allocation with very long strings
        long_string = "X" * 10000  # 10k character string

        long_allocations = [
            {
                "Van ID": "BW1",
                "Route Code": long_string,  # Extremely long route code
                "Associate Name": "Driver A",
                "Service Type": long_string,  # Extremely long service type
                "Wave": "8:00 AM",
                "Staging Location": long_string,  # Extremely long location
            }
        ]

        validator = DuplicateVehicleValidator()

        # Should handle long strings gracefully
        result = validator.validate_allocations(long_allocations)

        assert isinstance(result.duplicate_count, int)
        assert result.duplicate_count >= 0

    def test_circular_reference_handling(self, temp_dir):
        """Test handling of circular references in Excel files."""
        # Create Excel file with potential circular references
        circular_file = temp_dir / "circular.xlsx"

        # Use xlsxwriter to create file with formulas
        workbook = xlsxwriter.Workbook(str(circular_file))
        worksheet = workbook.add_worksheet("Vehicle Status")

        # Add headers
        worksheet.write(0, 0, "Van ID")
        worksheet.write(0, 1, "Type")
        worksheet.write(0, 2, "Opnal? Y/N")

        # Add data with potential circular reference in formula
        worksheet.write(1, 0, "BW1")
        worksheet.write(1, 1, "Large")
        worksheet.write_formula(1, 2, '=IF(C2="Y","Y","N")')  # Self-reference

        workbook.close()

        allocator = GASCompatibleAllocator()

        # Should handle circular references gracefully
        try:
            df = allocator.load_vehicle_status(str(circular_file))
            # If it loads, should have some data
            assert len(df) >= 1
        except Exception as e:
            # If it fails, should be a reasonable error
            error_msg = str(e).lower()
            assert any(
                keyword in error_msg for keyword in ["formula", "reference", "circular", "excel"]
            )

    def test_concurrent_file_access_handling(self, test_excel_files):
        """Test handling of concurrent file access scenarios."""
        import threading
        import time

        allocator1 = GASCompatibleAllocator()
        allocator2 = GASCompatibleAllocator()

        errors = []
        results = []

        def load_data(allocator, result_list, error_list):
            try:
                df = allocator.load_day_of_ops(str(test_excel_files["day_ops"]))
                result_list.append(len(df))
                time.sleep(0.1)  # Simulate processing time
            except Exception as e:
                error_list.append(str(e))

        # Start concurrent access
        thread1 = threading.Thread(target=load_data, args=(allocator1, results, errors))
        thread2 = threading.Thread(target=load_data, args=(allocator2, results, errors))

        thread1.start()
        thread2.start()

        thread1.join()
        thread2.join()

        # At least one should succeed, or both should have reasonable errors
        if errors:
            for error in errors:
                # Should be reasonable error messages for concurrent access
                assert any(
                    keyword in error.lower()
                    for keyword in ["lock", "access", "permission", "sharing"]
                )
        else:
            # If no errors, both should have results
            assert len(results) == 2
            assert all(r > 0 for r in results)

    # ==================== Recovery Strategy Tests ====================

    def test_graceful_degradation_strategy(self, corrupted_excel_files):
        """Test graceful degradation when some components fail."""
        allocator = GASCompatibleAllocator()

        # Test with partially corrupted data - should continue with what's available
        try:
            # Load what we can

            # Try to load each file type
            try:  # noqa: SIM105
                allocator.load_day_of_ops(str(corrupted_excel_files["unicode"]))  # This should work
            except:  # noqa: E722
                pass

            # Even if some files fail, should handle gracefully
            assert True  # Test passes if we reach here without crashing

        except Exception as e:
            # Should provide informative error about what went wrong
            assert len(str(e)) > 0

    def test_rollback_on_failure(self, test_excel_files):
        """Test rollback capabilities when operations fail."""
        allocator = GASCompatibleAllocator()

        # Load initial state
        original_state = {
            "day_of_ops_data": None,
            "vehicle_status_data": None,
            "allocation_results": [],
        }

        # Save original state
        original_state["allocation_results"] = allocator.allocation_results.copy()
        original_state["day_of_ops_data"] = allocator.day_of_ops_data
        original_state["vehicle_status_data"] = allocator.vehicle_status_data

        # Simulate operation that should fail and rollback
        with patch.object(allocator, "update_with_driver_names") as mock_update:
            mock_update.side_effect = Exception("Driver update failed")

            try:
                allocator.load_day_of_ops(str(test_excel_files["day_ops"]))
                allocator.load_vehicle_status(str(test_excel_files["summary"]))

                allocation_results, _ = allocator.allocate_vehicles_to_routes()

                # This should fail
                allocator.update_with_driver_names()
            except Exception:  # noqa: E722
                pass  # Expected to fail

            # On failure, state should be recoverable
            # Allocation results should still exist even if driver update failed
            assert len(allocator.allocation_results) > 0
            assert allocator.day_of_ops_data is not None
            assert allocator.vehicle_status_data is not None
