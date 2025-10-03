"""Test script for Daily Details thick border functionality."""

from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from src.services.daily_details_thick_borders import DailyDetailsThickBorderService
from src.services.daily_details_writer import DailyDetailsWriter


def create_test_daily_details():
    """Create a test Daily Details sheet with sample data."""
    wb = Workbook()

    # Create Daily Details sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    ws = wb.create_sheet("Daily Details")

    # Set up headers using DailyDetailsWriter
    writer = DailyDetailsWriter()
    writer._setup_daily_details_headers(ws)

    # Add sample data for three different dates
    sample_data = []
    base_date = date.today() - timedelta(days=2)

    # Day 1: 5 rows
    for i in range(5):
        sample_data.append(
            [
                base_date,  # Date
                f"CX{i+1}",  # Route #
                f"Driver {i+1}",  # Name
                f"A{i+1}",  # Asset ID
                f"BW{i+1}",  # Van ID
                f"1HGCM82633A{i:06d}",  # VIN
                f"GT{i+1}",  # GeoTab Code
                "Branded",  # Type
                "Large",  # Vehicle Type
                "Standard Parcel - Large Van",  # Route Type
                "",  # Rescue
                "",  # Delivery Pace 1:40pm
                "",  # Delivery Pace 3:40pm
                "",  # Delivery Pace 5:40pm
                "",  # Delivery Pace 7:40pm
                "",  # Delivery Pace 9:40pm
                "",  # RTS TIME
                "",  # Pkg. Delivered
                "",  # Pkg. Returned
                "",  # Route Notes
                "1",  # Week Number
                f"{base_date}|CX{i+1}|Driver {i+1}|BW{i+1}",  # Unique Identifier
                "",  # Vehicle Inspection
                "",  # Route Completion
            ]
        )

    # Day 2: 3 rows
    next_date = base_date + timedelta(days=1)
    for i in range(3):
        sample_data.append(
            [
                next_date,  # Date
                f"CX{i+10}",  # Route #
                f"Driver {i+10}",  # Name
                f"A{i+10}",  # Asset ID
                f"BW{i+10}",  # Van ID
                f"1HGCM82633A{i+10:06d}",  # VIN
                f"GT{i+10}",  # GeoTab Code
                "Rental",  # Type
                "Extra Large",  # Vehicle Type
                "Standard Parcel - Extra Large Van - US",  # Route Type
                "",  # Rescue
                "",  # Delivery Pace columns...
                "",
                "",
                "",
                "",
                "",
                "",  # RTS TIME
                "",  # Pkg. Delivered
                "",  # Pkg. Returned
                "",  # Route Notes
                "1",  # Week Number
                f"{next_date}|CX{i+10}|Driver {i+10}|BW{i+10}",  # Unique Identifier
                "",  # Vehicle Inspection
                "",  # Route Completion
            ]
        )

    # Day 3: 4 rows
    final_date = base_date + timedelta(days=2)
    for i in range(4):
        sample_data.append(
            [
                final_date,  # Date
                f"CX{i+20}",  # Route #
                f"Driver {i+20}",  # Name
                f"A{i+20}",  # Asset ID
                f"BW{i+20}",  # Van ID
                f"1HGCM82633A{i+20:06d}",  # VIN
                f"GT{i+20}",  # GeoTab Code
                "Branded",  # Type
                "Step Van",  # Vehicle Type
                "Standard Parcel Step Van - US",  # Route Type
                "",  # Rescue
                "",  # Delivery Pace columns...
                "",
                "",
                "",
                "",
                "",
                "",  # RTS TIME
                "",  # Pkg. Delivered
                "",  # Pkg. Returned
                "",  # Route Notes
                "1",  # Week Number
                f"{final_date}|CX{i+20}|Driver {i+20}|BW{i+20}",  # Unique Identifier
                "",  # Vehicle Inspection
                "",  # Route Completion
            ]
        )

    # Write data to sheet
    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Apply basic formatting
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for row in range(2, len(sample_data) + 2):
        for col in range(1, 25):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            if col == 1:  # Date column
                cell.number_format = "mm/dd/yyyy"

    return wb, ws


def test_thick_borders():
    """Test the thick border functionality."""
    print("Creating test Daily Details sheet...")
    wb, ws = create_test_daily_details()

    print("Applying thick borders...")
    border_service = DailyDetailsThickBorderService()
    border_service.apply_thick_borders_to_daily_details(ws)

    # Save the test file
    output_path = Path("tests/outputs/test_daily_details_thick_borders.xlsx")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb.save(output_path)
    print(f"Test file saved to: {output_path}")

    # Test the append functionality
    print("\nTesting append with thick borders...")

    # Add new rows for today
    today = date.today()
    new_rows = []
    for i in range(3):
        new_rows.append(
            [
                today,  # Date
                f"CX{i+30}",  # Route #
                f"Driver {i+30}",  # Name
                f"A{i+30}",  # Asset ID
                f"BW{i+30}",  # Van ID
                f"1HGCM82633A{i+30:06d}",  # VIN
                f"GT{i+30}",  # GeoTab Code
                "Branded",  # Type
                "Large",  # Vehicle Type
                "Standard Parcel - Large Van",  # Route Type
                "",  # Rescue
                "",  # Delivery Pace columns...
                "",
                "",
                "",
                "",
                "",
                "",  # RTS TIME
                "",  # Pkg. Delivered
                "",  # Pkg. Returned
                "",  # Route Notes
                "1",  # Week Number
                f"{today}|CX{i+30}|Driver {i+30}|BW{i+30}",  # Unique Identifier
                "",  # Vehicle Inspection
                "",  # Route Completion
            ]
        )

    # Append new rows
    start_row = ws.max_row + 1
    for row_idx, row_data in enumerate(new_rows, start=start_row):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Apply thick borders to new section
    border_service.apply_thick_borders_after_append(ws, start_row, len(new_rows))

    # Save the updated file
    output_path2 = Path("tests/outputs/test_daily_details_thick_borders_appended.xlsx")
    wb.save(output_path2)
    print(f"Updated file saved to: {output_path2}")

    print("\nTest completed successfully!")
    print("The files demonstrate:")
    print("1. Thick borders around each date section")
    print("2. Proper handling of multiple date groups")
    print("3. Correct border application when appending new data")


if __name__ == "__main__":
    test_thick_borders()
