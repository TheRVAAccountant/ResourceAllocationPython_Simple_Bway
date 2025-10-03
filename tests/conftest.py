"""Shared pytest fixtures for Resource Allocation Python test suite."""

import pytest
from datetime import date, datetime, timedelta
from pathlib import Path
import tempfile
import pandas as pd
from openpyxl import Workbook
from typing import Dict, List, Any
import numpy as np

from src.services.duplicate_validator import DuplicateVehicleValidator
from src.services.unassigned_vehicles_writer import UnassignedVehiclesWriter
from src.services.daily_details_thick_borders import DailyDetailsThickBorderService
from src.core.gas_compatible_allocator import GASCompatibleAllocator
from src.services.daily_details_writer import DailyDetailsWriter


# ==================== Test Data Fixtures ====================

@pytest.fixture
def sample_allocation_results():
    """Create sample allocation results with no duplicates."""
    return [
        {
            "Van ID": "BW1",
            "Route Code": "CX1",
            "Associate Name": "John Smith",
            "Service Type": "Standard Parcel - Large Van",
            "Wave": "8:00 AM",
            "Staging Location": "STG.G.1"
        },
        {
            "Van ID": "BW2",
            "Route Code": "CX2",
            "Associate Name": "Jane Doe",
            "Service Type": "Standard Parcel - Large Van",
            "Wave": "8:00 AM",
            "Staging Location": "STG.G.2"
        },
        {
            "Van ID": "BW3",
            "Route Code": "CX3",
            "Associate Name": "Bob Johnson",
            "Service Type": "Standard Parcel - Extra Large Van - US",
            "Wave": "9:00 AM",
            "Staging Location": "STG.G.3"
        }
    ]


@pytest.fixture
def duplicate_allocation_results():
    """Create allocation results with intentional duplicates."""
    return [
        {
            "Van ID": "BW1",
            "Route Code": "CX1",
            "Associate Name": "John Smith",
            "Service Type": "Standard Parcel - Large Van",
            "Wave": "8:00 AM",
            "Staging Location": "STG.G.1"
        },
        {
            "Van ID": "BW1",  # Duplicate
            "Route Code": "CX2",
            "Associate Name": "Jane Doe",
            "Service Type": "Standard Parcel - Large Van",
            "Wave": "9:00 AM",
            "Staging Location": "STG.G.2"
        },
        {
            "Van ID": "BW2",
            "Route Code": "CX3",
            "Associate Name": "Bob Johnson",
            "Service Type": "Standard Parcel - Extra Large Van - US",
            "Wave": "8:00 AM",
            "Staging Location": "STG.G.3"
        },
        {
            "Van ID": "BW2",  # Another duplicate
            "Route Code": "CX4",
            "Associate Name": "Alice Brown",
            "Service Type": "Standard Parcel - Large Van",
            "Wave": "10:00 AM",
            "Staging Location": "STG.G.4"
        }
    ]


@pytest.fixture
def sample_unassigned_vehicles_df():
    """Create sample unassigned vehicles DataFrame."""
    return pd.DataFrame([
        {
            "Van ID": "BW10",
            "Type": "Large",
            "Opnal? Y/N": "Y",
            "Location": "Depot A"
        },
        {
            "Van ID": "BW11",
            "Type": "Extra Large",
            "Opnal? Y/N": "Y",
            "Location": "Depot B"
        },
        {
            "Van ID": "BW12",
            "Type": "Step Van",
            "Opnal? Y/N": "N",
            "Location": "Depot A"
        },
        {
            "Van ID": "BW13",
            "Type": "Large",
            "Opnal? Y/N": "Y",
            "Location": "Depot C"
        },
        {
            "Van ID": "BW14",
            "Type": "Extra Large",
            "Opnal? Y/N": "N",
            "Location": "Depot B"
        }
    ])


@pytest.fixture
def vehicle_log_dict():
    """Create comprehensive vehicle log dictionary."""
    return {
        "BW1": {
            "vin": "1HGCM82633A004351",
            "geotab": "GT001",
            "brand_or_rental": "Branded",
            "vehicle_type": "Large"
        },
        "BW2": {
            "vin": "1HGCM82633A004352",
            "geotab": "GT002", 
            "brand_or_rental": "Rental",
            "vehicle_type": "Large"
        },
        "BW3": {
            "vin": "1HGCM82633A004353",
            "geotab": "GT003",
            "brand_or_rental": "Branded",
            "vehicle_type": "Extra Large"
        },
        "BW10": {
            "vin": "1HGCM82633A004360",
            "geotab": "GT010",
            "brand_or_rental": "Branded",
            "vehicle_type": "Large"  
        },
        "BW11": {
            "vin": "1HGCM82633A004361",
            "geotab": "GT011",
            "brand_or_rental": "Rental",
            "vehicle_type": "Extra Large"
        },
        "BW12": {
            "vin": "1HGCM82633A004362",
            "geotab": "",
            "brand_or_rental": "Branded",
            "vehicle_type": "Step Van"
        },
        "BW13": {
            "vin": "1HGCM82633A004363",
            "geotab": "GT013",
            "brand_or_rental": "Rental",
            "vehicle_type": "Large"
        },
        "BW14": {
            "vin": "1HGCM82633A004364",
            "geotab": "GT014",
            "brand_or_rental": "Branded",
            "vehicle_type": "Extra Large"
        }
    }


@pytest.fixture
def historical_assignments_df():
    """Create historical assignment data for testing."""
    base_date = datetime.now() - timedelta(days=10)
    data = []
    
    for i in range(15):  # 15 days of history
        assignment_date = base_date + timedelta(days=i)
        
        # Create assignments for various vehicles
        for van_id in ["BW1", "BW2", "BW3", "BW10", "BW11"]:
            if np.random.random() > 0.3:  # 70% chance of assignment each day
                data.append({
                    "Van ID": van_id,
                    "Date": assignment_date,
                    "Route Code": f"CX{np.random.randint(1, 20)}",
                    "Driver Name": f"Driver_{np.random.randint(1, 10)}"
                })
    
    return pd.DataFrame(data)


@pytest.fixture
def sample_day_of_ops_df():
    """Create comprehensive Day of Ops data."""
    return pd.DataFrame([
        {
            "Route Code": "CX1",
            "Service Type": "Standard Parcel - Large Van",
            "DSP": "BWAY",
            "Wave": "8:00 AM",
            "Staging Location": "STG.G.1"
        },
        {
            "Route Code": "CX2", 
            "Service Type": "Standard Parcel - Large Van",
            "DSP": "BWAY",
            "Wave": "8:00 AM",
            "Staging Location": "STG.G.2"
        },
        {
            "Route Code": "CX3",
            "Service Type": "Standard Parcel - Extra Large Van - US",
            "DSP": "BWAY",
            "Wave": "9:00 AM",
            "Staging Location": "STG.G.3"
        },
        {
            "Route Code": "CX4",
            "Service Type": "Standard Parcel Step Van - US",
            "DSP": "BWAY",
            "Wave": "9:00 AM",
            "Staging Location": "STG.G.4"
        },
        {
            "Route Code": "CX5",
            "Service Type": "Standard Parcel - Large Van",
            "DSP": "OTHER",  # Different DSP - should be filtered
            "Wave": "10:00 AM",
            "Staging Location": "STG.G.5"
        }
    ])


@pytest.fixture
def sample_vehicle_status_df():
    """Create vehicle status data."""
    return pd.DataFrame([
        {"Van ID": "BW1", "Type": "Large", "Opnal? Y/N": "Y"},
        {"Van ID": "BW2", "Type": "Large", "Opnal? Y/N": "Y"},
        {"Van ID": "BW3", "Type": "Extra Large", "Opnal? Y/N": "Y"},
        {"Van ID": "BW4", "Type": "Extra Large", "Opnal? Y/N": "Y"},
        {"Van ID": "BW5", "Type": "Step Van", "Opnal? Y/N": "Y"},
        {"Van ID": "BW6", "Type": "Large", "Opnal? Y/N": "N"},  # Not operational
        {"Van ID": "BW7", "Type": "Extra Large", "Opnal? Y/N": "N"},  # Not operational
        {"Van ID": "BW8", "Type": "Step Van", "Opnal? Y/N": "N"},  # Not operational
    ])


@pytest.fixture
def sample_daily_routes_df():
    """Create daily routes data."""
    return pd.DataFrame([
        {"Route Code": "CX1", "Driver Name": "John Smith"},
        {"Route Code": "CX2", "Driver Name": "Jane Doe"},
        {"Route Code": "CX3", "Driver Name": "Bob Johnson"},
        {"Route Code": "CX4", "Driver Name": "Alice Brown"},
        {"Route Code": "CX5", "Driver Name": "Charlie Davis"}
    ])


# ==================== Large Dataset Fixtures ====================

@pytest.fixture
def large_allocation_results():
    """Create large allocation results for performance testing."""
    results = []
    for i in range(1000):
        results.append({
            "Van ID": f"BW{i % 100 + 1}",  # Create duplicates by cycling van IDs
            "Route Code": f"CX{i + 1}",
            "Associate Name": f"Driver_{i + 1}",
            "Service Type": "Standard Parcel - Large Van",
            "Wave": f"{8 + (i % 4)}:00 AM",
            "Staging Location": f"STG.G.{i % 10 + 1}"
        })
    return results


@pytest.fixture
def large_unassigned_vehicles_df():
    """Create large unassigned vehicles dataset for performance testing."""
    data = []
    vehicle_types = ["Large", "Extra Large", "Step Van"]
    
    for i in range(500):
        data.append({
            "Van ID": f"BW{i + 1000}",
            "Type": vehicle_types[i % 3],
            "Opnal? Y/N": "Y" if i % 4 != 0 else "N",  # 75% operational
            "Location": f"Depot {chr(65 + (i % 5))}"  # Depot A-E
        })
    
    return pd.DataFrame(data)


# ==================== Service Instance Fixtures ====================

@pytest.fixture
def duplicate_validator():
    """Create DuplicateVehicleValidator instance."""
    validator = DuplicateVehicleValidator()
    validator.initialize()
    return validator


@pytest.fixture
def strict_duplicate_validator():
    """Create DuplicateVehicleValidator in strict mode."""
    validator = DuplicateVehicleValidator(config={"strict_duplicate_validation": True})
    validator.initialize()
    return validator


@pytest.fixture
def unassigned_writer():
    """Create UnassignedVehiclesWriter instance."""
    writer = UnassignedVehiclesWriter()
    writer.initialize()
    return writer


@pytest.fixture
def unassigned_writer_no_alternating():
    """Create UnassignedVehiclesWriter with alternating rows disabled."""
    return UnassignedVehiclesWriter(config={"enable_alternating_rows": False})


@pytest.fixture
def thick_border_service():
    """Create DailyDetailsThickBorderService instance."""
    return DailyDetailsThickBorderService()


@pytest.fixture
def gas_allocator():
    """Create GASCompatibleAllocator instance."""  
    return GASCompatibleAllocator()


@pytest.fixture
def daily_details_writer():
    """Create DailyDetailsWriter instance."""
    return DailyDetailsWriter()


# ==================== Excel Fixtures ====================

@pytest.fixture
def empty_workbook():
    """Create empty Excel workbook."""
    return Workbook()


@pytest.fixture
def daily_details_workbook():
    """Create workbook with Daily Details sheet structure."""
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    
    # Create Daily Details sheet
    ws = wb.create_sheet("Daily Details")
    
    # Set up headers using DailyDetailsWriter
    writer = DailyDetailsWriter()
    writer._setup_daily_details_headers(ws)
    
    return wb


@pytest.fixture
def daily_details_with_data():
    """Create Daily Details workbook with sample data across multiple dates."""
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
        
    ws = wb.create_sheet("Daily Details")
    
    # Set up headers
    writer = DailyDetailsWriter()
    writer._setup_daily_details_headers(ws)
    
    # Add sample data for three dates
    sample_data = []
    base_date = date.today() - timedelta(days=2)
    
    # Day 1: 3 rows
    for i in range(3):
        row_data = [
            base_date,  # Date
            f"CX{i+1}",  # Route #
            f"Driver {i+1}",  # Name
            f"A{i+1}",  # Asset ID
            f"BW{i+1}",  # Van ID
        ] + [""] * 19  # Fill remaining columns
        sample_data.append(row_data)
    
    # Day 2: 2 rows  
    next_date = base_date + timedelta(days=1)
    for i in range(2):
        row_data = [
            next_date,  # Date
            f"CX{i+10}",  # Route #
            f"Driver {i+10}",  # Name
            f"A{i+10}",  # Asset ID
            f"BW{i+10}",  # Van ID
        ] + [""] * 19  # Fill remaining columns
        sample_data.append(row_data)
    
    # Day 3: 4 rows
    final_date = base_date + timedelta(days=2)
    for i in range(4):
        row_data = [
            final_date,  # Date
            f"CX{i+20}",  # Route #
            f"Driver {i+20}",  # Name
            f"A{i+20}",  # Asset ID
            f"BW{i+20}",  # Van ID
        ] + [""] * 19  # Fill remaining columns
        sample_data.append(row_data)
    
    # Write data to worksheet
    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    return wb


# ==================== File System Fixtures ====================

@pytest.fixture
def temp_dir():
    """Create temporary directory for test files."""
    with tempfile.TemporaryDirectory() as temp_dir:
        yield Path(temp_dir)


@pytest.fixture
def test_excel_files(temp_dir, sample_day_of_ops_df, sample_daily_routes_df, 
                     sample_vehicle_status_df, vehicle_log_dict):
    """Create temporary Excel files for integration testing."""
    files = {}
    
    # Day of Ops file
    day_ops_file = temp_dir / "day_of_ops.xlsx"
    with pd.ExcelWriter(day_ops_file) as writer:
        sample_day_of_ops_df.to_excel(writer, sheet_name="Solution", index=False)
    files["day_ops"] = day_ops_file
    
    # Daily Routes file
    routes_file = temp_dir / "daily_routes.xlsx"
    with pd.ExcelWriter(routes_file) as writer:
        sample_daily_routes_df.to_excel(writer, sheet_name="Routes", index=False)
    files["routes"] = routes_file
    
    # Daily Summary file with multiple sheets
    summary_file = temp_dir / "daily_summary.xlsx"
    with pd.ExcelWriter(summary_file) as writer:
        sample_vehicle_status_df.to_excel(writer, sheet_name="Vehicle Status", index=False)
        
        # Convert vehicle_log_dict to DataFrame
        vehicle_log_df = pd.DataFrame.from_dict(vehicle_log_dict, orient="index")
        vehicle_log_df.reset_index(names=["Van ID"], inplace=True)
        vehicle_log_df.to_excel(writer, sheet_name="Vehicle Log", index=False)
    
    files["summary"] = summary_file
    
    return files


# ==================== Date Fixtures ====================

@pytest.fixture
def test_dates():
    """Provide various test dates."""
    today = date.today()
    return {
        "today": today,
        "yesterday": today - timedelta(days=1),
        "tomorrow": today + timedelta(days=1),
        "last_week": today - timedelta(days=7),
        "next_week": today + timedelta(days=7),
        "base_date": date(2025, 1, 15),  # Fixed date for consistent testing
    }


# ==================== Mock Data Generators ====================

def generate_allocation_results(count: int, duplicate_rate: float = 0.0) -> List[Dict[str, Any]]:
    """Generate allocation results with optional duplicates.
    
    Args:
        count: Number of allocation results to generate.
        duplicate_rate: Proportion of results that should be duplicates (0.0-1.0).
        
    Returns:
        List of allocation result dictionaries.
    """
    results = []
    van_ids = [f"BW{i+1}" for i in range(max(1, int(count * (1 - duplicate_rate))))]
    
    for i in range(count):
        if i < len(van_ids):
            van_id = van_ids[i]
        else:
            # Create duplicates by reusing van IDs
            van_id = van_ids[i % len(van_ids)]
        
        results.append({
            "Van ID": van_id,
            "Route Code": f"CX{i+1}",
            "Associate Name": f"Driver_{i+1}",
            "Service Type": "Standard Parcel - Large Van",
            "Wave": f"{8 + (i % 4)}:00 AM",
            "Staging Location": f"STG.G.{i % 10 + 1}"
        })
    
    return results


@pytest.fixture
def allocation_result_generator():
    """Provide function to generate allocation results."""
    return generate_allocation_results


# ==================== Validation Fixtures ====================

@pytest.fixture
def expected_unassigned_columns():
    """Expected columns in unassigned vehicles sheet."""
    return [
        "Van ID",
        "Vehicle Type", 
        "Operational Status",
        "Last Known Location",
        "Days Since Last Assignment",
        "VIN",
        "GeoTab Code",
        "Branded or Rental",
        "Notes",
        "Unassigned Date",
        "Unassigned Time"
    ]


@pytest.fixture
def expected_results_columns():
    """Expected columns in allocation results."""
    return [
        "Van ID",
        "Route Code",
        "Associate Name", 
        "Service Type",
        "Wave",
        "Staging Location",
        "Validation Status",
        "Validation Warning",
        "Conflict Level"
    ]