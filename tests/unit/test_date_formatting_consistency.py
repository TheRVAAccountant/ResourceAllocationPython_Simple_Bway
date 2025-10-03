"""Unit tests for date formatting consistency in Daily Details."""

from datetime import date

import pytest
from openpyxl import Workbook
from openpyxl.styles import Font

from src.services.daily_details_thick_borders import DailyDetailsThickBorderService
from src.services.daily_details_writer import DailyDetailsWriter


class TestDateFormattingConsistency:
    """Test that date formatting is consistent within date groups."""

    @pytest.fixture
    def daily_details_sheet(self):
        """Create a test Daily Details sheet with sample data."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Daily Details"

        # Add headers
        writer = DailyDetailsWriter()
        for col_idx, header in enumerate(writer.DAILY_DETAILS_COLUMNS, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        # Add test data
        test_data = [
            (date(2025, 1, 15), "CX1", "Driver A", "BW101"),
            (date(2025, 1, 15), "CX2", "Driver B", "BW102"),
            (date(2025, 1, 15), "CX3", "Driver C", "BW103"),
            (date(2025, 1, 16), "CX4", "Driver D", "BW104"),
            (date(2025, 1, 16), "CX5", "Driver E", "BW105"),
        ]

        for row_idx, (test_date, route, driver, van) in enumerate(test_data, start=2):
            ws.cell(row=row_idx, column=1, value=test_date)
            ws.cell(row=row_idx, column=2, value=route)
            ws.cell(row=row_idx, column=3, value=driver)
            ws.cell(row=row_idx, column=4, value=van)

        return ws

    def test_no_bold_dates_after_thick_borders(self, daily_details_sheet):
        """Test that no date cells are bold after applying thick borders."""
        # Apply thick borders
        border_service = DailyDetailsThickBorderService()
        border_service.apply_thick_borders_to_daily_details(daily_details_sheet)

        # Check all date cells
        bold_cells = []
        for row in range(2, daily_details_sheet.max_row + 1):
            date_cell = daily_details_sheet.cell(row=row, column=1)
            if date_cell.value and date_cell.font and date_cell.font.bold:
                bold_cells.append(row)

        assert len(bold_cells) == 0, f"Found bold date cells in rows: {bold_cells}"

    def test_consistent_formatting_within_groups(self, daily_details_sheet):
        """Test that all date cells within a group have the same formatting."""
        # Apply thick borders
        border_service = DailyDetailsThickBorderService()
        border_service.apply_thick_borders_to_daily_details(daily_details_sheet)

        # Group rows by date
        date_groups = {}
        for row in range(2, daily_details_sheet.max_row + 1):
            date_cell = daily_details_sheet.cell(row=row, column=1)
            if date_cell.value:
                date_val = date_cell.value
                if date_val not in date_groups:
                    date_groups[date_val] = []
                date_groups[date_val].append(row)

        # Check formatting consistency within each group
        for date_val, rows in date_groups.items():
            # Get formatting from first row
            first_cell = daily_details_sheet.cell(row=rows[0], column=1)
            first_bold = first_cell.font.bold if first_cell.font else False
            first_fill = (
                first_cell.fill.start_color.rgb
                if first_cell.fill and first_cell.fill.start_color
                else None
            )

            # Check all rows in group have same formatting
            for row in rows:
                cell = daily_details_sheet.cell(row=row, column=1)
                cell_bold = cell.font.bold if cell.font else False
                cell_fill = (
                    cell.fill.start_color.rgb if cell.fill and cell.fill.start_color else None
                )

                assert (
                    cell_bold == first_bold
                ), f"Inconsistent bold formatting in group {date_val}: row {row} differs from row {rows[0]}"
                assert (
                    cell_fill == first_fill
                ), f"Inconsistent fill formatting in group {date_val}: row {row} differs from row {rows[0]}"

    def test_gray_fill_applied_to_all_dates(self, daily_details_sheet):
        """Test that gray fill is applied to all date cells."""
        # Apply thick borders
        border_service = DailyDetailsThickBorderService()
        border_service.apply_thick_borders_to_daily_details(daily_details_sheet)

        # Check all date cells have gray fill
        missing_fill = []
        for row in range(2, daily_details_sheet.max_row + 1):
            date_cell = daily_details_sheet.cell(row=row, column=1)
            if date_cell.value:
                has_gray_fill = (
                    date_cell.fill
                    and date_cell.fill.start_color
                    and date_cell.fill.start_color.rgb in ["FFE6E6E6", "00E6E6E6"]
                )
                if not has_gray_fill:
                    missing_fill.append(row)

        assert len(missing_fill) == 0, f"Date cells without gray fill in rows: {missing_fill}"

    def test_append_maintains_consistency(self, daily_details_sheet):
        """Test that appending new data maintains formatting consistency."""
        # Apply initial borders
        border_service = DailyDetailsThickBorderService()
        border_service.apply_thick_borders_to_daily_details(daily_details_sheet)

        # Add new data
        new_start_row = daily_details_sheet.max_row + 1
        new_data = [
            (date(2025, 1, 17), "CX6", "Driver F", "BW106"),
            (date(2025, 1, 17), "CX7", "Driver G", "BW107"),
        ]

        for i, (test_date, route, driver, van) in enumerate(new_data):
            row = new_start_row + i
            daily_details_sheet.cell(row=row, column=1, value=test_date)
            daily_details_sheet.cell(row=row, column=2, value=route)
            daily_details_sheet.cell(row=row, column=3, value=driver)
            daily_details_sheet.cell(row=row, column=4, value=van)

        # Apply borders to new section
        border_service.apply_thick_borders_after_append(
            daily_details_sheet, new_start_row, len(new_data)
        )

        # Check all date cells (including new ones) are not bold
        bold_cells = []
        for row in range(2, daily_details_sheet.max_row + 1):
            date_cell = daily_details_sheet.cell(row=row, column=1)
            if date_cell.value and date_cell.font and date_cell.font.bold:
                bold_cells.append(row)

        assert len(bold_cells) == 0, f"Found bold date cells after append in rows: {bold_cells}"
