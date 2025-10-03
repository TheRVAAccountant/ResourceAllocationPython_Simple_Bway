"""Unit tests for unassigned vehicles writer."""

import time
from datetime import date, datetime
from unittest.mock import patch

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from src.services.unassigned_vehicles_writer import UnassignedVehiclesWriter


class TestUnassignedVehiclesWriter:
    """Test cases for UnassignedVehiclesWriter."""

    def test_create_unassigned_sheet(
        self, unassigned_writer, sample_unassigned_vehicles_df, vehicle_log_dict
    ):
        """Test creating unassigned vehicles sheet."""
        workbook = Workbook()
        allocation_date = date(2025, 1, 5)

        worksheet = unassigned_writer.create_unassigned_sheet(
            workbook=workbook,
            unassigned_vehicles=sample_unassigned_vehicles_df,
            vehicle_log_dict=vehicle_log_dict,
            allocation_date=allocation_date,
        )

        # Check sheet name
        expected_name = "01-05-25 Available & Unassigned"
        assert expected_name in workbook.sheetnames
        assert worksheet.title == expected_name

        # Check headers
        for col_idx, (header, _) in enumerate(unassigned_writer.COLUMNS, start=1):
            assert worksheet.cell(row=1, column=col_idx).value == header

        # Check data rows
        assert worksheet.cell(row=2, column=1).value == "BW10"  # Van ID
        assert worksheet.cell(row=2, column=2).value == "Large"  # Vehicle Type
        assert worksheet.cell(row=2, column=3).value == "Y"  # Operational Status
        assert worksheet.cell(row=2, column=6).value == "1HGCM82633A004352"  # VIN
        assert worksheet.cell(row=2, column=7).value == "GT001"  # GeoTab Code
        assert worksheet.cell(row=2, column=8).value == "Branded"  # Branded or Rental

        # Check date/time formatting
        assert worksheet.cell(row=2, column=10).value == "01/05/2025"
        assert ":" in str(worksheet.cell(row=2, column=11).value)  # Time format

    def test_column_widths(self, unassigned_writer):
        """Test column width settings."""
        workbook = Workbook()
        worksheet = workbook.create_sheet("Test")

        unassigned_writer._setup_headers(worksheet)

        # Check column widths are set
        assert worksheet.column_dimensions["A"].width == 12  # Van ID
        assert worksheet.column_dimensions["B"].width == 15  # Vehicle Type
        assert worksheet.column_dimensions["E"].width == 20  # Days Since Last Assignment

    def test_header_formatting(self, unassigned_writer):
        """Test header formatting."""
        workbook = Workbook()
        worksheet = workbook.create_sheet("Test")

        unassigned_writer._setup_headers(worksheet)

        # Check first header cell formatting
        header_cell = worksheet.cell(row=1, column=1)
        assert header_cell.font.bold
        assert header_cell.font.color == "FFFFFF"
        assert header_cell.fill.start_color.rgb == "FF002060"  # Dark blue
        assert header_cell.alignment.horizontal == "center"

    def test_calculate_days_since_assignment(self, unassigned_writer):
        """Test days since assignment calculation."""
        # Create historical data
        historical_data = pd.DataFrame(
            [
                {"Van ID": "BW10", "Date": pd.Timestamp("2025-01-01")},
                {"Van ID": "BW11", "Date": pd.Timestamp("2024-12-25")},
                {"Van ID": "BW10", "Date": pd.Timestamp("2025-01-03")},  # More recent
            ]
        )

        # Test with vehicle that has history
        days = unassigned_writer.calculate_days_since_assignment("BW10", historical_data)
        expected_days = (datetime.now().date() - date(2025, 1, 3)).days
        assert days == expected_days

        # Test with vehicle that has no history
        days = unassigned_writer.calculate_days_since_assignment("BW99", historical_data)
        assert days == 0

        # Test with None historical data
        days = unassigned_writer.calculate_days_since_assignment("BW10", None)
        assert days == 0

    def test_create_unassigned_summary(self, unassigned_writer, sample_unassigned_vehicles_df):
        """Test unassigned vehicles summary creation."""
        summary = unassigned_writer.create_unassigned_summary(sample_unassigned_vehicles_df)

        assert summary["total_unassigned"] == 3
        assert summary["operational_unassigned"] == 2
        assert summary["non_operational_unassigned"] == 1
        assert summary["by_type"]["Large"] == 1
        assert summary["by_type"]["Extra Large"] == 1
        assert summary["by_type"]["Step Van"] == 1

    def test_format_unassigned_sheet(self, unassigned_writer):
        """Test sheet formatting."""
        workbook = Workbook()
        date(2025, 1, 5)
        worksheet = workbook.create_sheet("Test")

        # Add some test data
        unassigned_writer._setup_headers(worksheet)
        for row in range(2, 5):
            for col in range(1, 12):
                worksheet.cell(row=row, column=col, value=f"Data{row}{col}")

        unassigned_writer.format_unassigned_sheet(worksheet)

        # Check AutoFilter is applied
        assert worksheet.auto_filter.ref == "A1:K4"

        # Check freeze panes
        assert worksheet.freeze_panes == "A2"

        # Check print settings
        assert worksheet.page_setup.orientation == "landscape"
        assert worksheet.page_setup.fitToWidth == 1

    def test_alternating_row_colors(
        self, unassigned_writer, sample_unassigned_vehicles_df, vehicle_log_dict
    ):
        """Test alternating row colors."""
        workbook = Workbook()
        allocation_date = date(2025, 1, 5)

        worksheet = unassigned_writer.create_unassigned_sheet(
            workbook=workbook,
            unassigned_vehicles=sample_unassigned_vehicles_df,
            vehicle_log_dict=vehicle_log_dict,
            allocation_date=allocation_date,
        )

        # Check even rows have gray background
        assert worksheet.cell(row=2, column=1).fill.start_color.rgb == "FFF2F2F2"
        # Check odd rows have no fill (or white)
        assert (
            worksheet.cell(row=3, column=1).fill.start_color.rgb is None
            or worksheet.cell(row=3, column=1).fill.start_color.rgb == "00000000"
        )

    def test_export_unassigned_to_csv(
        self, unassigned_writer, sample_unassigned_vehicles_df, vehicle_log_dict, tmp_path
    ):
        """Test CSV export functionality."""
        output_path = tmp_path / "unassigned.csv"
        allocation_date = date(2025, 1, 5)

        unassigned_writer.export_unassigned_to_csv(
            unassigned_vehicles=sample_unassigned_vehicles_df,
            vehicle_log_dict=vehicle_log_dict,
            output_path=str(output_path),
            allocation_date=allocation_date,
        )

        # Check file exists
        assert output_path.exists()

        # Read and verify CSV
        df = pd.read_csv(output_path)
        assert len(df) == 3
        assert "Van ID" in df.columns
        assert "VIN" in df.columns
        assert "GeoTab Code" in df.columns
        assert df.iloc[0]["Van ID"] == "BW10"
        assert df.iloc[0]["VIN"] == "1HGCM82633A004352"

    def test_empty_unassigned_vehicles(self, unassigned_writer, vehicle_log_dict):
        """Test handling empty unassigned vehicles."""
        workbook = Workbook()
        allocation_date = date(2025, 1, 5)
        empty_df = pd.DataFrame()

        worksheet = unassigned_writer.create_unassigned_sheet(
            workbook=workbook,
            unassigned_vehicles=empty_df,
            vehicle_log_dict=vehicle_log_dict,
            allocation_date=allocation_date,
        )

        # Should still create sheet with headers
        assert worksheet.max_row == 1  # Only headers
        assert worksheet.cell(row=1, column=1).value == "Van ID"

    def test_missing_vehicle_log_data(self, unassigned_writer, sample_unassigned_vehicles_df):
        """Test handling missing vehicle log data."""
        workbook = Workbook()
        allocation_date = date(2025, 1, 5)
        empty_vehicle_log = {}

        worksheet = unassigned_writer.create_unassigned_sheet(
            workbook=workbook,
            unassigned_vehicles=sample_unassigned_vehicles_df,
            vehicle_log_dict=empty_vehicle_log,
            allocation_date=allocation_date,
        )

        # Should still create sheet with available data
        assert worksheet.cell(row=2, column=1).value == "BW10"  # Van ID
        assert worksheet.cell(row=2, column=6).value == ""  # VIN (empty)
        assert worksheet.cell(row=2, column=7).value == ""  # GeoTab (empty)

    # ==================== Edge Cases ====================

    def test_empty_unassigned_vehicles_comprehensive(self, unassigned_writer, vehicle_log_dict):
        """Comprehensive test for empty unassigned vehicles handling."""
        workbook = Workbook()
        allocation_date = date(2025, 1, 5)

        # Test with truly empty DataFrame
        empty_df = pd.DataFrame()
        worksheet = unassigned_writer.create_unassigned_sheet(
            workbook=workbook,
            unassigned_vehicles=empty_df,
            vehicle_log_dict=vehicle_log_dict,
            allocation_date=allocation_date,
        )

        # Should create sheet with headers only
        assert worksheet.max_row == 1
        assert worksheet.cell(row=1, column=1).value == "Van ID"

        # Check all headers are present
        for col_idx, (header, _) in enumerate(unassigned_writer.COLUMNS, start=1):
            assert worksheet.cell(row=1, column=col_idx).value == header

        # Check formatting is still applied
        unassigned_writer.format_unassigned_sheet(worksheet)
        assert worksheet.freeze_panes == "A2"

    def test_vehicles_with_missing_columns(self, unassigned_writer, vehicle_log_dict):
        """Test handling of vehicles with missing required columns."""
        # DataFrame missing some columns
        incomplete_vehicles = pd.DataFrame(
            [
                {"Van ID": "BW20"},  # Missing Type, Opnal? Y/N, Location
                {"Van ID": "BW21", "Type": "Large"},  # Missing Opnal? Y/N, Location
                {"Van ID": "BW22", "Opnal? Y/N": "Y"},  # Missing Type, Location
            ]
        )

        workbook = Workbook()
        allocation_date = date(2025, 1, 5)

        worksheet = unassigned_writer.create_unassigned_sheet(
            workbook=workbook,
            unassigned_vehicles=incomplete_vehicles,
            vehicle_log_dict=vehicle_log_dict,
            allocation_date=allocation_date,
        )

        # Should handle missing data gracefully
        assert worksheet.max_row == 4  # Header + 3 data rows

        # Check first vehicle (BW20)
        assert worksheet.cell(row=2, column=1).value == "BW20"
        assert worksheet.cell(row=2, column=2).value == "Unknown"  # Default Type
        assert worksheet.cell(row=2, column=3).value == "N"  # Default Opnal? Y/N

        # Check second vehicle (BW21)
        assert worksheet.cell(row=3, column=1).value == "BW21"
        assert worksheet.cell(row=3, column=2).value == "Large"
        assert worksheet.cell(row=3, column=3).value == "N"  # Default Opnal? Y/N

    def test_vehicles_with_none_values(self, unassigned_writer):
        """Test handling of vehicles with None values in fields."""
        vehicles_with_nones = pd.DataFrame(
            [
                {"Van ID": "BW30", "Type": None, "Opnal? Y/N": None, "Location": None},
                {
                    "Van ID": None,  # This should be skipped
                    "Type": "Large",
                    "Opnal? Y/N": "Y",
                    "Location": "Depot A",
                },
            ]
        )

        workbook = Workbook()
        allocation_date = date(2025, 1, 5)

        worksheet = unassigned_writer.create_unassigned_sheet(
            workbook=workbook,
            unassigned_vehicles=vehicles_with_nones,
            vehicle_log_dict={},
            allocation_date=allocation_date,
        )

        # Should only have header + 1 data row (Van ID None should be skipped)
        assert worksheet.max_row == 2
        assert worksheet.cell(row=2, column=1).value == "BW30"

    def test_duplicate_sheet_replacement(
        self, unassigned_writer, sample_unassigned_vehicles_df, vehicle_log_dict
    ):
        """Test that existing sheets are properly replaced."""
        workbook = Workbook()
        allocation_date = date(2025, 1, 5)
        sheet_name = allocation_date.strftime("%m-%d-%y") + " Available & Unassigned"

        # Create an existing sheet with same name
        existing_sheet = workbook.create_sheet(sheet_name)
        existing_sheet.cell(row=1, column=1, value="Old Data")

        assert sheet_name in workbook.sheetnames

        # Create new sheet (should replace existing)
        new_worksheet = unassigned_writer.create_unassigned_sheet(
            workbook=workbook,
            unassigned_vehicles=sample_unassigned_vehicles_df,
            vehicle_log_dict=vehicle_log_dict,
            allocation_date=allocation_date,
        )

        # Should have replaced the old sheet
        assert sheet_name in workbook.sheetnames
        assert new_worksheet.cell(row=1, column=1).value == "Van ID"  # New header
        assert new_worksheet.cell(row=1, column=1).value != "Old Data"  # Not old data

    def test_large_dataset_performance(self, unassigned_writer, large_unassigned_vehicles_df):
        """Test performance with large unassigned vehicles dataset."""
        workbook = Workbook()
        allocation_date = date(2025, 1, 5)

        # Create vehicle log for large dataset
        large_vehicle_log = {}
        for i in range(500):
            large_vehicle_log[f"BW{i + 1000}"] = {
                "vin": f"1HGCM82633A{i:06d}",
                "geotab": f"GT{i + 1000}",
                "brand_or_rental": "Branded" if i % 2 == 0 else "Rental",
            }

        start_time = time.time()
        worksheet = unassigned_writer.create_unassigned_sheet(
            workbook=workbook,
            unassigned_vehicles=large_unassigned_vehicles_df,
            vehicle_log_dict=large_vehicle_log,
            allocation_date=allocation_date,
        )
        end_time = time.time()

        execution_time = end_time - start_time

        # Should complete in reasonable time (< 5 seconds for 500 vehicles)
        assert execution_time < 5.0, f"Large dataset processing took {execution_time:.2f} seconds"

        # Verify all data was written
        assert worksheet.max_row == 501  # Header + 500 data rows

        print(
            f"Large dataset processing: {execution_time:.3f}s for {len(large_unassigned_vehicles_df)} vehicles"
        )

    # ==================== Formatting Tests ====================

    def test_header_formatting_comprehensive(self, unassigned_writer):
        """Comprehensive test of header formatting."""
        workbook = Workbook()
        worksheet = workbook.create_sheet("Test")

        unassigned_writer._setup_headers(worksheet)

        # Test all header cells
        for col_idx, (header, expected_width) in enumerate(unassigned_writer.COLUMNS, start=1):
            cell = worksheet.cell(row=1, column=col_idx)

            # Check content
            assert cell.value == header

            # Check font formatting
            assert cell.font.bold is True
            assert cell.font.color.rgb == "FFFFFF"  # White text
            assert cell.font.name == "Calibri"
            assert cell.font.size == 11

            # Check fill formatting
            assert cell.fill.start_color.rgb == "FF002060"  # Dark blue background
            assert cell.fill.fill_type == "solid"

            # Check alignment
            assert cell.alignment.horizontal == "center"
            assert cell.alignment.vertical == "center"
            assert cell.alignment.wrap_text is True

            # Check border
            assert cell.border.left.style == "thin"
            assert cell.border.right.style == "thin"
            assert cell.border.top.style == "thin"
            assert cell.border.bottom.style == "thin"

            # Check column width
            column_letter = get_column_letter(col_idx)
            assert worksheet.column_dimensions[column_letter].width == expected_width

    def test_data_cell_formatting(
        self, unassigned_writer, sample_unassigned_vehicles_df, vehicle_log_dict
    ):
        """Test formatting of data cells."""
        workbook = Workbook()
        allocation_date = date(2025, 1, 5)

        worksheet = unassigned_writer.create_unassigned_sheet(
            workbook=workbook,
            unassigned_vehicles=sample_unassigned_vehicles_df,
            vehicle_log_dict=vehicle_log_dict,
            allocation_date=allocation_date,
        )

        # Check border formatting on data cells
        for row in range(2, worksheet.max_row + 1):
            for col in range(1, len(unassigned_writer.COLUMNS) + 1):
                cell = worksheet.cell(row=row, column=col)
                assert cell.border.left.style == "thin"
                assert cell.border.right.style == "thin"
                assert cell.border.top.style == "thin"
                assert cell.border.bottom.style == "thin"

        # Check alternating row colors (if enabled)
        if unassigned_writer.enable_alternating_rows:
            # Even rows should have gray background
            even_row_cell = worksheet.cell(row=2, column=1)
            assert even_row_cell.fill.start_color.rgb == "FFF2F2F2"

            # Odd rows should have no special fill
            odd_row_cell = worksheet.cell(row=3, column=1)
            # Default fill should be None or white
            assert (
                odd_row_cell.fill.start_color.rgb is None
                or odd_row_cell.fill.start_color.rgb == "00000000"
            )

    def test_alternating_rows_disabled(self, sample_unassigned_vehicles_df, vehicle_log_dict):
        """Test behavior when alternating rows are disabled."""
        writer = UnassignedVehiclesWriter(config={"enable_alternating_rows": False})
        workbook = Workbook()
        allocation_date = date(2025, 1, 5)

        worksheet = writer.create_unassigned_sheet(
            workbook=workbook,
            unassigned_vehicles=sample_unassigned_vehicles_df,
            vehicle_log_dict=vehicle_log_dict,
            allocation_date=allocation_date,
        )

        # No rows should have alternating colors
        for row in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=1)
            assert cell.fill.start_color.rgb is None or cell.fill.start_color.rgb == "00000000"

    def test_specific_column_alignment(
        self, unassigned_writer, sample_unassigned_vehicles_df, vehicle_log_dict
    ):
        """Test specific column alignment formatting."""
        workbook = Workbook()
        allocation_date = date(2025, 1, 5)

        worksheet = unassigned_writer.create_unassigned_sheet(
            workbook=workbook,
            unassigned_vehicles=sample_unassigned_vehicles_df,
            vehicle_log_dict=vehicle_log_dict,
            allocation_date=allocation_date,
        )

        # Check specific column alignments
        for row in range(2, worksheet.max_row + 1):
            # Column 5: Days Since Last Assignment (center aligned)
            days_cell = worksheet.cell(row=row, column=5)
            assert days_cell.alignment.horizontal == "center"

            # Column 10: Unassigned Date (center aligned)
            date_cell = worksheet.cell(row=row, column=10)
            assert date_cell.alignment.horizontal == "center"

            # Column 11: Unassigned Time (center aligned)
            time_cell = worksheet.cell(row=row, column=11)
            assert time_cell.alignment.horizontal == "center"

    def test_print_settings(self, unassigned_writer):
        """Test print settings configuration."""
        workbook = Workbook()
        worksheet = workbook.create_sheet("Test")

        # Add some data and format
        unassigned_writer._setup_headers(worksheet)
        worksheet.cell(row=2, column=1, value="Test Data")

        unassigned_writer.format_unassigned_sheet(worksheet)

        # Check print settings
        assert worksheet.page_setup.orientation == "landscape"
        assert worksheet.page_setup.fitToWidth == 1
        assert worksheet.page_setup.fitToHeight is False

        # Check margins
        assert worksheet.page_margins.left == 0.5
        assert worksheet.page_margins.right == 0.5
        assert worksheet.page_margins.top == 0.75
        assert worksheet.page_margins.bottom == 0.75
        assert worksheet.page_margins.header == 0.3
        assert worksheet.page_margins.footer == 0.3

    # ==================== Data Processing Tests ====================

    def test_calculate_days_since_assignment_edge_cases(self, unassigned_writer):
        """Test edge cases for days since assignment calculation."""
        # Empty historical data
        empty_df = pd.DataFrame()
        days = unassigned_writer.calculate_days_since_assignment("BW1", empty_df)
        assert days == 0

        # Historical data without Date column
        no_date_df = pd.DataFrame([{"Van ID": "BW1", "Route": "CX1"}])
        days = unassigned_writer.calculate_days_since_assignment("BW1", no_date_df)
        assert days == 0

        # Historical data with invalid dates
        invalid_date_df = pd.DataFrame(
            [{"Van ID": "BW1", "Date": "invalid-date"}, {"Van ID": "BW1", "Date": None}]
        )
        days = unassigned_writer.calculate_days_since_assignment("BW1", invalid_date_df)
        assert days == 0

        # Vehicle not in historical data
        historical_df = pd.DataFrame([{"Van ID": "BW2", "Date": pd.Timestamp("2025-01-01")}])
        days = unassigned_writer.calculate_days_since_assignment("BW1", historical_df)
        assert days == 0

    def test_calculate_days_since_assignment_multiple_dates(self, unassigned_writer):
        """Test days calculation with multiple historical dates."""
        historical_data = pd.DataFrame(
            [
                {"Van ID": "BW1", "Date": pd.Timestamp("2025-01-01")},
                {"Van ID": "BW1", "Date": pd.Timestamp("2025-01-05")},  # More recent
                {"Van ID": "BW1", "Date": pd.Timestamp("2024-12-25")},  # Older
                {"Van ID": "BW2", "Date": pd.Timestamp("2025-01-03")},  # Different vehicle
            ]
        )

        days = unassigned_writer.calculate_days_since_assignment("BW1", historical_data)

        # Should use most recent date (2025-01-05)
        expected_days = (datetime.now().date() - date(2025, 1, 5)).days
        assert days == expected_days

    def test_create_unassigned_summary_comprehensive(self, unassigned_writer):
        """Comprehensive test of unassigned summary creation."""
        # Create diverse unassigned vehicles data
        diverse_vehicles = pd.DataFrame(
            [
                {"Van ID": "BW1", "Type": "Large", "Opnal? Y/N": "Y"},
                {"Van ID": "BW2", "Type": "Large", "Opnal? Y/N": "Y"},
                {"Van ID": "BW3", "Type": "Extra Large", "Opnal? Y/N": "Y"},
                {"Van ID": "BW4", "Type": "Extra Large", "Opnal? Y/N": "N"},
                {"Van ID": "BW5", "Type": "Step Van", "Opnal? Y/N": "Y"},
                {"Van ID": "BW6", "Type": "Step Van", "Opnal? Y/N": "N"},
                {"Van ID": "BW7", "Type": "Large", "Opnal? Y/N": "N"},
            ]
        )

        summary = unassigned_writer.create_unassigned_summary(diverse_vehicles)

        # Check totals
        assert summary["total_unassigned"] == 7
        assert summary["operational_unassigned"] == 3  # BW1, BW2, BW3, BW5
        assert summary["non_operational_unassigned"] == 4  # BW4, BW6, BW7

        # Check by type
        assert summary["by_type"]["Large"] == 3
        assert summary["by_type"]["Extra Large"] == 2
        assert summary["by_type"]["Step Van"] == 2

        # Check timestamp exists
        assert "timestamp" in summary
        assert isinstance(summary["timestamp"], str)

    def test_create_unassigned_summary_edge_cases(self, unassigned_writer):
        """Test summary creation with edge cases."""
        # Empty DataFrame
        empty_df = pd.DataFrame()
        summary = unassigned_writer.create_unassigned_summary(empty_df)
        assert summary["total_unassigned"] == 0
        assert summary["operational_unassigned"] == 0
        assert summary["non_operational_unassigned"] == 0
        assert summary["by_type"] == {}

        # DataFrame without required columns
        incomplete_df = pd.DataFrame([{"Van ID": "BW1"}])  # Missing Type and Opnal? Y/N
        summary = unassigned_writer.create_unassigned_summary(incomplete_df)
        assert summary["total_unassigned"] == 1
        assert summary["operational_unassigned"] == 0  # No Opnal? Y/N column
        assert summary["by_type"] == {}  # No Type column

    # ==================== CSV Export Tests ====================

    def test_csv_export_functionality(
        self, unassigned_writer, sample_unassigned_vehicles_df, vehicle_log_dict, temp_dir
    ):
        """Test CSV export functionality."""
        output_path = temp_dir / "unassigned_export.csv"
        allocation_date = date(2025, 1, 5)

        unassigned_writer.export_unassigned_to_csv(
            unassigned_vehicles=sample_unassigned_vehicles_df,
            vehicle_log_dict=vehicle_log_dict,
            output_path=str(output_path),
            allocation_date=allocation_date,
        )

        # Verify file exists
        assert output_path.exists()

        # Read and verify CSV content
        df = pd.read_csv(output_path)

        # Check structure
        expected_columns = [
            "Van ID",
            "Vehicle Type",
            "Operational Status",
            "VIN",
            "GeoTab Code",
            "Branded or Rental",
            "Unassigned Date",
        ]
        for col in expected_columns:
            assert col in df.columns

        # Check data content
        assert len(df) == len(sample_unassigned_vehicles_df)
        assert "BW10" in df["Van ID"].values
        assert "01/05/2025" in df["Unassigned Date"].values

    def test_csv_export_missing_vehicle_log(
        self, unassigned_writer, sample_unassigned_vehicles_df, temp_dir
    ):
        """Test CSV export with missing vehicle log data."""
        output_path = temp_dir / "unassigned_no_log.csv"
        allocation_date = date(2025, 1, 5)
        empty_vehicle_log = {}

        unassigned_writer.export_unassigned_to_csv(
            unassigned_vehicles=sample_unassigned_vehicles_df,
            vehicle_log_dict=empty_vehicle_log,
            output_path=str(output_path),
            allocation_date=allocation_date,
        )

        # Should still create file
        assert output_path.exists()

        df = pd.read_csv(output_path)
        assert len(df) == len(sample_unassigned_vehicles_df)

        # VIN, GeoTab Code, Branded or Rental should be empty
        assert all(pd.isna(df["VIN"]) or df["VIN"] == "")
        assert all(pd.isna(df["GeoTab Code"]) or df["GeoTab Code"] == "")
        assert all(pd.isna(df["Branded or Rental"]) or df["Branded or Rental"] == "")

    # ==================== Error Handling Tests ====================

    @patch("src.services.unassigned_vehicles_writer.logger")
    def test_logging_behavior(
        self, mock_logger, unassigned_writer, sample_unassigned_vehicles_df, vehicle_log_dict
    ):
        """Test that appropriate logging occurs."""
        workbook = Workbook()
        allocation_date = date(2025, 1, 5)

        unassigned_writer.create_unassigned_sheet(
            workbook=workbook,
            unassigned_vehicles=sample_unassigned_vehicles_df,
            vehicle_log_dict=vehicle_log_dict,
            allocation_date=allocation_date,
        )

        # Should log sheet creation
        sheet_name = allocation_date.strftime("%m-%d-%y") + " Available & Unassigned"
        mock_logger.info.assert_any_call(f"Creating unassigned vehicles sheet: {sheet_name}")

        # Should log data writing completion
        expected_count = len(sample_unassigned_vehicles_df)
        mock_logger.info.assert_any_call(f"Wrote {expected_count} unassigned vehicles to sheet")

    def test_error_handling_in_days_calculation(self, unassigned_writer):
        """Test error handling in days since assignment calculation."""
        # Malformed historical data that might cause errors
        problematic_data = pd.DataFrame(
            [
                {"Van ID": "BW1", "Date": "not-a-date"},
                {"Van ID": "BW1", "Date": 12345},  # Number instead of date
                {"Van ID": "BW1", "Date": ["list", "instead", "of", "date"]},
            ]
        )

        # Should not crash, should return 0
        days = unassigned_writer.calculate_days_since_assignment("BW1", problematic_data)
        assert days == 0

    def test_worksheet_name_edge_cases(
        self, unassigned_writer, sample_unassigned_vehicles_df, vehicle_log_dict
    ):
        """Test worksheet naming with various date formats."""
        workbook = Workbook()

        test_dates = [
            date(2025, 1, 1),  # New Year
            date(2025, 12, 31),  # End of year
            date(2025, 2, 29),  # Leap year (if applicable)
            date(2025, 10, 10),  # Double digits
        ]

        for test_date in test_dates:
            try:
                worksheet = unassigned_writer.create_unassigned_sheet(
                    workbook=workbook,
                    unassigned_vehicles=sample_unassigned_vehicles_df,
                    vehicle_log_dict=vehicle_log_dict,
                    allocation_date=test_date,
                )

                expected_name = test_date.strftime("%m-%d-%y") + " Available & Unassigned"
                assert worksheet.title == expected_name
                assert expected_name in workbook.sheetnames
            except ValueError as e:
                # February 29 might not exist in non-leap years
                if test_date.month == 2 and test_date.day == 29:
                    continue  # Skip invalid leap year date
                else:
                    raise e
