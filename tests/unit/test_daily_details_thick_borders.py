"""Comprehensive unit tests for Daily Details thick borders service."""

import time
from datetime import date, datetime, timedelta
from unittest.mock import MagicMock, patch

import pytest
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from src.services.daily_details_thick_borders import DailyDetailsThickBorderService


class TestDailyDetailsThickBorderService:
    """Test cases for DailyDetailsThickBorderService."""

    # ==================== Basic Functionality Tests ====================

    def test_service_initialization(self, thick_border_service):
        """Test service initialization with default settings."""
        assert thick_border_service.thick_side.style == "thick"
        assert thick_border_service.thick_side.color == "000000"
        assert thick_border_service.thin_side.style == "thin"
        assert thick_border_service.thin_side.color == "000000"

        # Check date header styling
        assert thick_border_service.date_header_fill.start_color == "E6E6E6"
        assert thick_border_service.date_header_font.bold == True
        assert thick_border_service.date_header_font.size == 11
        assert thick_border_service.date_header_alignment.horizontal == "left"

    def test_apply_thick_borders_empty_sheet(self, thick_border_service, empty_workbook):
        """Test applying borders to empty sheet."""
        worksheet = empty_workbook.create_sheet("Daily Details")

        # Should handle empty sheet gracefully
        thick_border_service.apply_thick_borders_to_daily_details(worksheet)

        # No changes should be made to empty sheet
        assert worksheet.max_row == 1
        assert worksheet.max_column == 1

    def test_apply_thick_borders_single_date_section(
        self, thick_border_service, daily_details_with_data
    ):
        """Test borders for single date section."""
        worksheet = daily_details_with_data["Daily Details"]

        # Clear existing data and add single date section
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.value = None

        # Add single date section (3 rows)
        test_date = date(2025, 1, 15)
        for i in range(3):
            worksheet.cell(row=i + 2, column=1, value=test_date)
            worksheet.cell(row=i + 2, column=2, value=f"CX{i+1}")
            worksheet.cell(row=i + 2, column=3, value=f"Driver {i+1}")

        thick_border_service.apply_thick_borders_to_daily_details(worksheet)

        # Check that thick borders are applied around the section
        for row in range(2, 5):  # Rows 2-4
            for col in range(1, 25):  # Columns 1-24
                cell = worksheet.cell(row=row, column=col)

                # Check top border
                if row == 2:
                    assert cell.border.top.style == "thick"
                else:
                    assert cell.border.top.style == "thin"

                # Check bottom border
                if row == 4:
                    assert cell.border.bottom.style == "thick"
                else:
                    assert cell.border.bottom.style == "thin"

                # Check left border
                if col == 1:
                    assert cell.border.left.style == "thick"
                else:
                    assert cell.border.left.style == "thin"

                # Check right border
                if col == 24:
                    assert cell.border.right.style == "thick"
                else:
                    assert cell.border.right.style == "thin"

        # Check date header formatting
        date_cell = worksheet.cell(row=2, column=1)
        assert date_cell.fill.start_color.rgb == "FFE6E6E6"
        assert date_cell.font.bold == True
        assert date_cell.alignment.horizontal == "left"

    def test_apply_thick_borders_multiple_date_sections(
        self, thick_border_service, daily_details_with_data
    ):
        """Test borders for multiple date sections."""
        worksheet = daily_details_with_data["Daily Details"]

        # The fixture already has multiple date sections
        thick_border_service.apply_thick_borders_to_daily_details(worksheet)

        # Verify borders exist (detailed verification in single section test)
        # Just check that processing completed without errors
        assert worksheet.max_row > 1

        # Check that date cells have special formatting
        date_cells = []
        current_date = None
        for row in range(2, worksheet.max_row + 1):
            date_cell = worksheet.cell(row=row, column=1)
            if date_cell.value and date_cell.value != current_date:
                current_date = date_cell.value
                date_cells.append(date_cell)

        # Each unique date should have header formatting
        for date_cell in date_cells:
            assert date_cell.fill.start_color.rgb == "FFE6E6E6"
            assert date_cell.font.bold == True

    # ==================== Date Parsing Tests ====================

    def test_parse_date_value_datetime_objects(self, thick_border_service):
        """Test parsing datetime objects."""
        # datetime object
        dt = datetime(2025, 1, 15, 10, 30, 0)
        parsed = thick_border_service._parse_date_value(dt)
        assert parsed == date(2025, 1, 15)

        # date object
        d = date(2025, 1, 15)
        parsed = thick_border_service._parse_date_value(d)
        assert parsed == date(2025, 1, 15)

    def test_parse_date_value_strings(self, thick_border_service):
        """Test parsing various string date formats."""
        test_cases = [
            ("01/15/2025", date(2025, 1, 15)),
            ("2025-01-15", date(2025, 1, 15)),
            ("15-01-2025", date(2025, 1, 15)),
            ("01-15-2025", date(2025, 1, 15)),
        ]

        for date_str, expected in test_cases:
            parsed = thick_border_service._parse_date_value(date_str)
            assert parsed == expected, f"Failed to parse {date_str}"

    def test_parse_date_value_invalid_inputs(self, thick_border_service):
        """Test parsing invalid date inputs."""
        invalid_inputs = [None, "", "invalid-date", "not-a-date", 123, [], {}]

        for invalid_input in invalid_inputs:
            parsed = thick_border_service._parse_date_value(invalid_input)
            assert parsed is None, f"Should return None for {invalid_input}"

    # ==================== Date Section Identification Tests ====================

    def test_identify_date_sections_continuous_dates(
        self, thick_border_service, daily_details_workbook
    ):
        """Test identification of continuous date sections."""
        worksheet = daily_details_workbook["Daily Details"]

        # Add test data with continuous dates
        base_date = date(2025, 1, 15)
        row_data = [
            (2, base_date),
            (3, base_date),
            (4, base_date),
            (5, base_date + timedelta(days=1)),
            (6, base_date + timedelta(days=1)),
            (7, base_date + timedelta(days=2)),
        ]

        for row, date_val in row_data:
            worksheet.cell(row=row, column=1, value=date_val)
            worksheet.cell(row=row, column=2, value=f"Route{row}")

        sections = thick_border_service._identify_date_sections(worksheet, 2)

        # Should identify 3 sections
        assert len(sections) == 3
        assert sections[base_date] == (2, 4)  # Rows 2-4
        assert sections[base_date + timedelta(days=1)] == (5, 6)  # Rows 5-6
        assert sections[base_date + timedelta(days=2)] == (7, 7)  # Row 7

    def test_identify_date_sections_scattered_dates(
        self, thick_border_service, daily_details_workbook
    ):
        """Test identification with scattered date sections."""
        worksheet = daily_details_workbook["Daily Details"]

        # Add test data with scattered dates
        dates = [
            date(2025, 1, 15),
            date(2025, 1, 10),  # Earlier date
            date(2025, 1, 15),  # Back to first date
            date(2025, 1, 20),  # Later date
        ]

        for i, date_val in enumerate(dates, start=2):
            worksheet.cell(row=i, column=1, value=date_val)
            worksheet.cell(row=i, column=2, value=f"Route{i}")

        sections = thick_border_service._identify_date_sections(worksheet, 2)

        # Should identify each transition as separate section
        assert len(sections) == 4  # Each row is different from previous
        assert sections[date(2025, 1, 15)] == (2, 2)
        assert sections[date(2025, 1, 10)] == (3, 3)
        # Second occurrence of 2025-1-15 overwrites first
        assert sections[date(2025, 1, 20)] == (5, 5)

    def test_identify_date_sections_with_gaps(self, thick_border_service, daily_details_workbook):
        """Test identification with gaps in data."""
        worksheet = daily_details_workbook["Daily Details"]

        # Add test data with gaps
        worksheet.cell(row=2, column=1, value=date(2025, 1, 15))
        worksheet.cell(row=3, column=1, value=date(2025, 1, 15))
        # Row 4 is empty (gap)
        worksheet.cell(row=5, column=1, value=date(2025, 1, 16))

        sections = thick_border_service._identify_date_sections(worksheet, 2)

        # Should stop at first gap
        assert len(sections) == 1
        assert sections[date(2025, 1, 15)] == (2, 3)

    # ==================== Border Application Tests ====================

    def test_apply_thick_border_to_section_single_row(
        self, thick_border_service, daily_details_workbook
    ):
        """Test border application to single row section."""
        worksheet = daily_details_workbook["Daily Details"]
        test_date = date(2025, 1, 15)

        # Add single row of data
        worksheet.cell(row=2, column=1, value=test_date)
        worksheet.cell(row=2, column=2, value="CX1")

        thick_border_service._apply_thick_border_to_section(worksheet, 2, 2, test_date)

        # Check that all borders are thick for single row
        for col in range(1, 25):
            cell = worksheet.cell(row=2, column=col)
            assert cell.border.top.style == "thick"
            assert cell.border.bottom.style == "thick"

            if col == 1:
                assert cell.border.left.style == "thick"
            else:
                assert cell.border.left.style == "thin"

            if col == 24:
                assert cell.border.right.style == "thick"
            else:
                assert cell.border.right.style == "thin"

    def test_apply_thick_border_to_section_multiple_rows(
        self, thick_border_service, daily_details_workbook
    ):
        """Test border application to multiple row section."""
        worksheet = daily_details_workbook["Daily Details"]
        test_date = date(2025, 1, 15)

        # Add multiple rows of data
        for i in range(3):
            worksheet.cell(row=i + 2, column=1, value=test_date)
            worksheet.cell(row=i + 2, column=2, value=f"CX{i+1}")

        thick_border_service._apply_thick_border_to_section(worksheet, 2, 4, test_date)

        # Check borders for each row
        for row in range(2, 5):  # Rows 2-4
            for col in range(1, 25):  # Columns 1-24
                cell = worksheet.cell(row=row, column=col)

                # Top border - thick only on first row
                if row == 2:
                    assert cell.border.top.style == "thick"
                else:
                    assert cell.border.top.style == "thin"

                # Bottom border - thick only on last row
                if row == 4:
                    assert cell.border.bottom.style == "thick"
                else:
                    assert cell.border.bottom.style == "thin"

                # Left border - thick only on first column
                if col == 1:
                    assert cell.border.left.style == "thick"
                else:
                    assert cell.border.left.style == "thin"

                # Right border - thick only on last column
                if col == 24:
                    assert cell.border.right.style == "thick"
                else:
                    assert cell.border.right.style == "thin"

    # ==================== Append Functionality Tests ====================

    def test_apply_thick_borders_after_append_new_date(
        self, thick_border_service, daily_details_with_data
    ):
        """Test applying borders after appending new date section."""
        worksheet = daily_details_with_data["Daily Details"]

        # Get current max row
        original_max_row = worksheet.max_row

        # Add new data with different date
        new_date = date(2025, 2, 1)
        new_start_row = original_max_row + 1
        num_new_rows = 2

        for i in range(num_new_rows):
            row = new_start_row + i
            worksheet.cell(row=row, column=1, value=new_date)
            worksheet.cell(row=row, column=2, value=f"CX{row}")

        # Apply borders to new section
        thick_border_service.apply_thick_borders_after_append(
            worksheet, new_start_row, num_new_rows
        )

        # Check that new section has thick borders
        for row in range(new_start_row, new_start_row + num_new_rows):
            # Check left and right borders
            left_cell = worksheet.cell(row=row, column=1)
            assert left_cell.border.left.style == "thick"

            right_cell = worksheet.cell(row=row, column=24)
            assert right_cell.border.right.style == "thick"

        # Check top border of new section
        top_row_cell = worksheet.cell(row=new_start_row, column=1)
        assert top_row_cell.border.top.style == "thick"

        # Check bottom border of new section
        bottom_row_cell = worksheet.cell(row=new_start_row + num_new_rows - 1, column=1)
        assert bottom_row_cell.border.bottom.style == "thick"

    def test_apply_thick_borders_after_append_same_date(
        self, thick_border_service, daily_details_workbook
    ):
        """Test appending to existing date section."""
        worksheet = daily_details_workbook["Daily Details"]

        # Add initial date section
        test_date = date(2025, 1, 15)
        worksheet.cell(row=2, column=1, value=test_date)
        worksheet.cell(row=2, column=2, value="CX1")

        # Apply initial borders
        thick_border_service._apply_thick_border_to_section(worksheet, 2, 2, test_date)

        # Verify initial thick bottom border
        initial_cell = worksheet.cell(row=2, column=1)
        assert initial_cell.border.bottom.style == "thick"

        # Add more rows with same date
        new_start_row = 3
        num_new_rows = 2

        for i in range(num_new_rows):
            row = new_start_row + i
            worksheet.cell(row=row, column=1, value=test_date)
            worksheet.cell(row=row, column=2, value=f"CX{row}")

        # Apply borders after append
        thick_border_service.apply_thick_borders_after_append(
            worksheet, new_start_row, num_new_rows
        )

        # The bottom border of the original section should now be thin
        # (this is handled by _extend_previous_section)
        # But the new last row should have thick bottom border
        last_row_cell = worksheet.cell(row=new_start_row + num_new_rows - 1, column=1)
        assert last_row_cell.border.bottom.style == "thick"

    def test_extend_previous_section(self, thick_border_service, daily_details_workbook):
        """Test extending previous date section."""
        worksheet = daily_details_workbook["Daily Details"]

        # Set up a cell with thick bottom border
        cell = worksheet.cell(row=2, column=1)
        cell.border = Border(
            top=Side(style="thin"),
            bottom=Side(style="thick"),
            left=Side(style="thick"),
            right=Side(style="thin"),
        )

        # Extend the previous section
        thick_border_service._extend_previous_section(worksheet, 2)

        # Bottom border should now be thin
        assert cell.border.bottom.style == "thin"
        # Other borders should remain unchanged
        assert cell.border.top.style == "thin"
        assert cell.border.left.style == "thick"
        assert cell.border.right.style == "thin"

    # ==================== Performance Tests ====================

    def test_performance_large_single_section(self, thick_border_service, daily_details_workbook):
        """Test performance with large single date section."""
        worksheet = daily_details_workbook["Daily Details"]

        # Add large section (100 rows)
        test_date = date(2025, 1, 15)
        num_rows = 100

        for i in range(num_rows):
            row = i + 2
            worksheet.cell(row=row, column=1, value=test_date)
            worksheet.cell(row=row, column=2, value=f"CX{i+1}")

        start_time = time.time()
        thick_border_service.apply_thick_borders_to_daily_details(worksheet)
        end_time = time.time()

        execution_time = end_time - start_time

        # Should complete in reasonable time (< 5 seconds for 100 rows)
        assert execution_time < 5.0, f"Large section processing took {execution_time:.2f} seconds"

        print(f"Large section processing: {execution_time:.3f}s for {num_rows} rows")

    def test_performance_many_small_sections(self, thick_border_service, daily_details_workbook):
        """Test performance with many small date sections."""
        worksheet = daily_details_workbook["Daily Details"]

        # Add 50 single-row sections with different dates
        base_date = date(2025, 1, 1)
        num_sections = 50

        for i in range(num_sections):
            row = i + 2
            section_date = base_date + timedelta(days=i)
            worksheet.cell(row=row, column=1, value=section_date)
            worksheet.cell(row=row, column=2, value=f"CX{i+1}")

        start_time = time.time()
        thick_border_service.apply_thick_borders_to_daily_details(worksheet)
        end_time = time.time()

        execution_time = end_time - start_time

        # Should complete in reasonable time (< 3 seconds for 50 sections)
        assert execution_time < 3.0, f"Many sections processing took {execution_time:.2f} seconds"

        print(f"Many sections processing: {execution_time:.3f}s for {num_sections} sections")

    # ==================== Error Handling Tests ====================

    @patch("src.services.daily_details_thick_borders.logger")
    def test_logging_behavior(self, mock_logger, thick_border_service, daily_details_with_data):
        """Test that appropriate logging occurs."""
        worksheet = daily_details_with_data["Daily Details"]

        thick_border_service.apply_thick_borders_to_daily_details(worksheet)

        # Should log start of processing
        mock_logger.info.assert_any_call("Applying thick borders to Daily Details sheet")

        # Should log completion with section count
        completion_calls = [
            call
            for call in mock_logger.info.call_args_list
            if "Applied thick borders to" in str(call)
        ]
        assert len(completion_calls) >= 1

    def test_error_handling_malformed_worksheet(self, thick_border_service):
        """Test handling of malformed worksheet data."""
        # This is more of a safety test - openpyxl should handle most issues
        workbook = Workbook()
        worksheet = workbook.create_sheet("Test")

        # Add some problematic data
        worksheet.cell(row=2, column=1, value="not-a-date")
        worksheet.cell(row=3, column=1, value=None)
        worksheet.cell(row=4, column=1, value=123)  # Number instead of date

        # Should not crash
        try:
            thick_border_service.apply_thick_borders_to_daily_details(worksheet)
        except Exception as e:
            pytest.fail(f"Should handle malformed data gracefully, but got: {e}")

    def test_empty_date_cells(self, thick_border_service, daily_details_workbook):
        """Test handling of empty date cells in data rows."""
        worksheet = daily_details_workbook["Daily Details"]

        # Add mixed data with some empty date cells
        worksheet.cell(row=2, column=1, value=date(2025, 1, 15))
        worksheet.cell(row=2, column=2, value="CX1")

        worksheet.cell(row=3, column=1, value=None)  # Empty date
        worksheet.cell(row=3, column=2, value="CX2")

        worksheet.cell(row=4, column=1, value=date(2025, 1, 16))
        worksheet.cell(row=4, column=2, value="CX3")

        # Should handle gracefully and process what it can
        sections = thick_border_service._identify_date_sections(worksheet, 2)

        # Should stop at empty cell and not process beyond
        assert len(sections) == 1
        assert sections[date(2025, 1, 15)] == (2, 2)

    # ==================== Edge Cases ====================

    def test_same_date_different_formats(self, thick_border_service, daily_details_workbook):
        """Test handling same date in different formats."""
        worksheet = daily_details_workbook["Daily Details"]

        # Add same date in different formats
        jan_15 = date(2025, 1, 15)
        worksheet.cell(row=2, column=1, value=jan_15)
        worksheet.cell(row=3, column=1, value=datetime(2025, 1, 15, 10, 30))  # datetime
        worksheet.cell(row=4, column=1, value="01/15/2025")  # string

        sections = thick_border_service._identify_date_sections(worksheet, 2)

        # All should be grouped as same date
        assert len(sections) == 1
        assert sections[jan_15] == (2, 4)

    def test_start_row_parameter(self, thick_border_service, daily_details_workbook):
        """Test different start row parameters."""
        worksheet = daily_details_workbook["Daily Details"]

        # Add data starting from row 5
        test_date = date(2025, 1, 15)
        for i in range(3):
            worksheet.cell(row=i + 5, column=1, value=test_date)
            worksheet.cell(row=i + 5, column=2, value=f"CX{i+1}")

        # Process starting from row 5
        thick_border_service.apply_thick_borders_to_daily_details(worksheet, start_row=5)

        # Check that borders were applied starting from row 5
        cell_5_1 = worksheet.cell(row=5, column=1)
        assert cell_5_1.border.top.style == "thick"
        assert cell_5_1.border.left.style == "thick"

        # Rows before start_row should be unaffected
        cell_2_1 = worksheet.cell(row=2, column=1)
        assert cell_2_1.border.top.style != "thick" or cell_2_1.border.top.style is None

    def test_column_range_boundaries(self, thick_border_service, daily_details_workbook):
        """Test that borders are applied to correct column range."""
        worksheet = daily_details_workbook["Daily Details"]

        # Add test data
        test_date = date(2025, 1, 15)
        worksheet.cell(row=2, column=1, value=test_date)

        thick_border_service._apply_thick_border_to_section(worksheet, 2, 2, test_date)

        # Check column 1 has thick left border
        assert worksheet.cell(row=2, column=1).border.left.style == "thick"

        # Check column 24 has thick right border
        assert worksheet.cell(row=2, column=24).border.right.style == "thick"

        # Check column 25 is not affected (beyond range)
        cell_25 = worksheet.cell(row=2, column=25)
        assert cell_25.border.right.style != "thick" or cell_25.border.right.style is None
