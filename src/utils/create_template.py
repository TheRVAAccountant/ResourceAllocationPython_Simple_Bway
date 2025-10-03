"""Create template Excel files matching Google Apps Script expectations."""

import pandas as pd
from pathlib import Path
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


def create_daily_summary_log_template(output_path: str = None):
    """Create a Daily Summary Log template matching Google Apps Script structure.
    
    Creates an Excel file with sheets:
    - Daily Details (with proper column headers)
    - Vehicle Log (vehicle data)
    - Employees (driver data)
    - Route Types (route type options)
    
    Args:
        output_path: Output file path. If None, uses default name with year.
    """
    if not output_path:
        current_year = datetime.now().year
        output_path = f"Daily_Summary_Log_{current_year}.xlsx"
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Define styles
    header_fill = PatternFill(start_color="46bdc6", end_color="46bdc6", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    header_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Create Daily Details sheet
    daily_details = wb.create_sheet("Daily Details")
    daily_details_headers = [
        "Date", "Van ID", "VIN", "GeoTab Code", "Driver ID", "Name",
        "Route Type", "Route Number", "Start Time", "End Time",
        "1:40pm", "3:00pm", "4:20pm", "5:40pm", "7:00pm", "8:20pm",
        "Stops Remaining", "Packages Remaining", "Notes", "Status",
        "Unique Identifier"
    ]
    
    for col, header in enumerate(daily_details_headers, 1):
        cell = daily_details.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = header_border
        cell.alignment = header_alignment
    
    # Set column widths
    for col in range(1, len(daily_details_headers) + 1):
        daily_details.column_dimensions[chr(64 + col)].width = 15
    
    # Add sample data for Daily Details
    sample_date = date.today()
    for row in range(2, 5):
        daily_details.cell(row=row, column=1, value=sample_date)
        daily_details.cell(row=row, column=2, value=f"VAN{row-1:03d}")
        daily_details.cell(row=row, column=3, value=f"VIN{row-1:06d}")
        daily_details.cell(row=row, column=4, value=f"GT{row-1:04d}")
        daily_details.cell(row=row, column=5, value=f"EMP{row-1:03d}")
        daily_details.cell(row=row, column=6, value=f"Driver {row-1}")
        daily_details.cell(row=row, column=7, value="Standard")
        daily_details.cell(row=row, column=8, value=f"R{row-1:02d}")
    
    # Create Vehicle Log sheet
    vehicle_log = wb.create_sheet("Vehicle Log")
    vehicle_headers = ["Van ID", "VIN", "GeoTab Code", "Branded or Rental", 
                      "Location", "Status", "Priority", "Notes"]
    
    for col, header in enumerate(vehicle_headers, 1):
        cell = vehicle_log.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = header_border
        cell.alignment = header_alignment
    
    # Add sample vehicle data
    vehicle_data = [
        ["VAN001", "VIN000001", "GT0001", "Branded", "Main", "Active", "High", ""],
        ["VAN002", "VIN000002", "GT0002", "Branded", "North", "Active", "Medium", ""],
        ["VAN003", "VIN000003", "GT0003", "Rental", "South", "Active", "Low", ""],
        ["VAN004", "VIN000004", "GT0004", "Branded", "Main", "Maintenance", "High", ""],
        ["VAN005", "VIN000005", "GT0005", "Rental", "North", "Active", "Medium", ""],
    ]
    
    for row_idx, row_data in enumerate(vehicle_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            vehicle_log.cell(row=row_idx, column=col_idx, value=value)
    
    # Create Employees sheet
    employees = wb.create_sheet("Employees")
    employee_headers = ["Employee ID", "Name", "Location", "Status", 
                       "Priority", "Experience", "License Type", "Notes"]
    
    for col, header in enumerate(employee_headers, 1):
        cell = employees.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = header_border
        cell.alignment = header_alignment
    
    # Add sample employee data
    employee_data = [
        ["EMP001", "John Smith", "Main", "Active", "High", 5, "CDL", ""],
        ["EMP002", "Jane Doe", "North", "Active", "Medium", 3, "Standard", ""],
        ["EMP003", "Bob Johnson", "South", "Active", "Low", 2, "Standard", ""],
        ["EMP004", "Alice Williams", "Main", "Active", "High", 7, "CDL", ""],
        ["EMP005", "Charlie Brown", "North", "Inactive", "Medium", 1, "Standard", "On leave"],
    ]
    
    for row_idx, row_data in enumerate(employee_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            employees.cell(row=row_idx, column=col_idx, value=value)
    
    # Create Route Types sheet
    route_types = wb.create_sheet("Route Types")
    route_headers = ["Route Type", "Description", "Priority", "Max Stops", "Max Packages"]
    
    for col, header in enumerate(route_headers, 1):
        cell = route_types.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = header_border
        cell.alignment = header_alignment
    
    # Add route type data
    route_data = [
        ["Standard", "Regular delivery route", "Medium", 150, 300],
        ["Express", "Priority delivery route", "High", 100, 200],
        ["Bulk", "Large package route", "Low", 50, 100],
        ["Rural", "Extended area route", "Medium", 120, 250],
        ["Urban", "City center route", "High", 180, 350],
    ]
    
    for row_idx, row_data in enumerate(route_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            route_types.cell(row=row_idx, column=col_idx, value=value)
    
    # Save the workbook
    wb.save(output_path)
    print(f"Template created: {output_path}")
    return output_path


def create_input_files_for_upload():
    """Create the two input files expected by the Google Apps Script upload dialog.
    
    Creates:
    - Day_of_Ops.xlsx: Contains operational data (routes, schedules)
    - Daily_Routes.xlsx: Contains route assignments
    """
    # Create Day of Ops file
    day_of_ops_data = {
        "Date": [date.today()] * 5,
        "Route": [f"R{i:02d}" for i in range(1, 6)],
        "Type": ["Standard", "Express", "Bulk", "Rural", "Urban"],
        "Start Time": ["08:00", "08:30", "09:00", "07:30", "08:00"],
        "End Time": ["17:00", "16:30", "18:00", "17:30", "17:00"],
        "Stops": [150, 100, 50, 120, 180],
        "Packages": [300, 200, 100, 250, 350],
    }
    
    df_ops = pd.DataFrame(day_of_ops_data)
    df_ops.to_excel("Day_of_Ops.xlsx", index=False, sheet_name="Operations")
    print("Created: Day_of_Ops.xlsx")
    
    # Create Daily Routes file
    daily_routes_data = {
        "Van ID": [f"VAN{i:03d}" for i in range(1, 6)],
        "Driver ID": [f"EMP{i:03d}" for i in range(1, 6)],
        "Route": [f"R{i:02d}" for i in range(1, 6)],
        "Status": ["Assigned"] * 5,
        "Notes": [""] * 5,
    }
    
    df_routes = pd.DataFrame(daily_routes_data)
    df_routes.to_excel("Daily_Routes.xlsx", index=False, sheet_name="Routes")
    print("Created: Daily_Routes.xlsx")


def create_python_compatible_template(output_path: str = None):
    """Create a template compatible with the Python application.
    
    Creates an Excel file with sheets using Python naming convention:
    - Vehicles (instead of Vehicle Log)
    - Drivers (instead of Employees)
    - Daily Details (same as GAS)
    
    Args:
        output_path: Output file path. If None, uses default name.
    """
    if not output_path:
        output_path = "Resource_Allocation_Template.xlsx"
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Define styles
    header_fill = PatternFill(start_color="46bdc6", end_color="46bdc6", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    header_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Create Vehicles sheet
    vehicles = wb.create_sheet("Vehicles")
    vehicle_headers = ["Vehicle Number", "Type", "Location", "Status", "Priority", "Capacity"]
    
    for col, header in enumerate(vehicle_headers, 1):
        cell = vehicles.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = header_border
        cell.alignment = header_alignment
    
    # Add sample vehicle data
    for i in range(2, 7):
        vehicles.cell(row=i, column=1, value=f"VEH{i-1:03d}")
        vehicles.cell(row=i, column=2, value=["Standard", "Premium", "Economy"][(i-2) % 3])
        vehicles.cell(row=i, column=3, value=["Main", "North", "South"][(i-2) % 3])
        vehicles.cell(row=i, column=4, value="available")
        vehicles.cell(row=i, column=5, value=50 + ((i-2) % 3) * 10)
        vehicles.cell(row=i, column=6, value=4 + ((i-2) % 2) * 2)
    
    # Create Drivers sheet
    drivers = wb.create_sheet("Drivers")
    driver_headers = ["Employee ID", "Name", "Location", "Status", "Priority", "Experience", "License Type"]
    
    for col, header in enumerate(driver_headers, 1):
        cell = drivers.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = header_border
        cell.alignment = header_alignment
    
    # Add sample driver data
    names = ["John Smith", "Jane Doe", "Bob Johnson", "Alice Williams", "Charlie Brown"]
    for i in range(2, 7):
        drivers.cell(row=i, column=1, value=f"EMP{i-1:03d}")
        drivers.cell(row=i, column=2, value=names[i-2])
        drivers.cell(row=i, column=3, value=["Main", "North", "South"][(i-2) % 3])
        drivers.cell(row=i, column=4, value="active")
        drivers.cell(row=i, column=5, value=["Low", "Medium", "High"][(i-2) % 3])
        drivers.cell(row=i, column=6, value=(i-2) % 10)
        drivers.cell(row=i, column=7, value="Standard")
    
    # Create Daily Details sheet
    daily_details = wb.create_sheet("Daily Details")
    daily_details_headers = [
        "Date", "Vehicle Number", "Driver Name", "Route", "Start Time", 
        "End Time", "Status", "Notes"
    ]
    
    for col, header in enumerate(daily_details_headers, 1):
        cell = daily_details.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = header_border
        cell.alignment = header_alignment
    
    # Save the workbook
    wb.save(output_path)
    print(f"Template created: {output_path}")
    return output_path


if __name__ == "__main__":
    import sys
    
    if len(sys.argv) > 1 and sys.argv[1] == "--gas":
        # Create Google Apps Script compatible templates
        print("Creating Google Apps Script compatible templates...")
        create_daily_summary_log_template()
        create_input_files_for_upload()
    elif len(sys.argv) > 1 and sys.argv[1] == "--python":
        # Create Python application compatible template
        print("Creating Python application compatible template...")
        create_python_compatible_template()
    else:
        # Create both versions
        print("Creating all template versions...")
        print("\n1. Google Apps Script compatible:")
        create_daily_summary_log_template()
        create_input_files_for_upload()
        print("\n2. Python application compatible:")
        create_python_compatible_template()
        print("\nUse --gas or --python flags to create specific versions.")