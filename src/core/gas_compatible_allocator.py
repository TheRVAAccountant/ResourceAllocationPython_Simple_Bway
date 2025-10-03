"""GAS-compatible allocation engine matching exact Google Apps Script logic."""

from datetime import datetime, date
from typing import Dict, List, Optional, Tuple, Any
from pathlib import Path
import time
import pandas as pd
from loguru import logger
from collections import defaultdict

from src.models.allocation import AllocationResult, AllocationStatus
from src.services.daily_details_writer import DailyDetailsWriter
from src.services.duplicate_validator import DuplicateVehicleValidator
from src.services.unassigned_vehicles_writer import UnassignedVehiclesWriter
from src.services.allocation_output_writer import AllocationOutputWriter
from src.services.allocation_history_service import AllocationHistoryService


class GASCompatibleAllocator:
    """Allocation engine that exactly matches Google Apps Script logic.
    
    This allocator:
    1. Loads Day of Ops and Daily Routes files
    2. Filters for DSP = "BWAY" routes only
    3. Matches routes with operational vehicles by type
    4. Uses first-come-first-served allocation (no optimization)
    5. Creates Results, Daily Details, and Unassigned sheets
    """
    
    # Service Type to Van Type mapping (exact match from GAS)
    SERVICE_TYPE_TO_VAN_TYPE = {
        "Standard Parcel - Extra Large Van - US": "Extra Large",
        "Standard Parcel - Large Van": "Large",
        "Standard Parcel Step Van - US": "Step Van"
    }
    
    def __init__(self):
        """Initialize the GAS-compatible allocator."""
        self.day_of_ops_data = None
        self.vehicle_status_data = None
        self.daily_routes_data = None
        self.vehicle_log_data = None
        self.allocation_results = []
        self.assigned_van_ids = []
        self.unassigned_vehicles = []
        
        # Initialize validators and writers
        self.duplicate_validator = DuplicateVehicleValidator()
        self.duplicate_validator.initialize()
        self.unassigned_writer = UnassignedVehiclesWriter()
        self.unassigned_writer.initialize()
        self.output_writer = AllocationOutputWriter()
        self.output_writer.initialize()
        self.history_service = AllocationHistoryService()
        self.history_service.initialize()
    
    def load_day_of_ops(self, file_path: str, sheet_name: str = "Solution") -> pd.DataFrame:
        """Load Day of Ops file.
        
        Expected columns:
        - Route Code
        - Service Type  
        - DSP
        - Wave
        - Staging Location
        
        Args:
            file_path: Path to Day of Ops Excel file.
            sheet_name: Name of sheet containing data (default: "Solution").
            
        Returns:
            DataFrame with Day of Ops data.
        """
        logger.info(f"Loading Day of Ops from {file_path}, sheet: {sheet_name}")
        
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            
            # Verify required columns
            required_cols = ["Route Code", "Service Type", "DSP", "Wave", "Staging Location"]
            missing_cols = [col for col in required_cols if col not in df.columns]
            
            if missing_cols:
                raise ValueError(f"Missing required columns in Day of Ops: {missing_cols}")
            
            self.day_of_ops_data = df
            logger.info(f"Loaded {len(df)} routes from Day of Ops")
            
            return df
            
        except Exception as e:
            logger.error(f"Failed to load Day of Ops: {e}")
            raise
    
    def load_daily_routes(self, file_path: str, sheet_name: str = "Routes") -> pd.DataFrame:
        """Load Daily Routes file.
        
        Expected columns:
        - Route code (or Route Code)
        - Driver name (or Driver Name)
        
        Args:
            file_path: Path to Daily Routes Excel file.
            sheet_name: Name of sheet containing data (default: "Routes").
            
        Returns:
            DataFrame with Daily Routes data.
        """
        logger.info(f"Loading Daily Routes from {file_path}, sheet: {sheet_name}")
        
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            
            # Handle column name variations
            df.columns = df.columns.str.strip()
            
            # Normalize column names
            if "Route code" in df.columns:
                df.rename(columns={"Route code": "Route Code"}, inplace=True)
            if "Driver name" in df.columns:
                df.rename(columns={"Driver name": "Driver Name"}, inplace=True)
            
            # Verify required columns
            required_cols = ["Route Code", "Driver Name"]
            missing_cols = [col for col in required_cols if col not in df.columns]
            
            if missing_cols:
                # Try alternative names
                if "Route" in df.columns and "Route Code" not in df.columns:
                    df.rename(columns={"Route": "Route Code"}, inplace=True)
                if "Driver" in df.columns and "Driver Name" not in df.columns:
                    df.rename(columns={"Driver": "Driver Name"}, inplace=True)
                
                # Check again
                missing_cols = [col for col in required_cols if col not in df.columns]
                if missing_cols:
                    raise ValueError(f"Missing required columns in Daily Routes: {missing_cols}")
            
            self.daily_routes_data = df
            logger.info(f"Loaded {len(df)} driver assignments from Daily Routes")
            
            return df
            
        except Exception as e:
            logger.error(f"Failed to load Daily Routes: {e}")
            raise
    
    def load_vehicle_status(
        self,
        file_path: str,
        sheet_name: str = "Vehicle Status"
    ) -> pd.DataFrame:
        """Load Vehicle Status from Daily Summary file.
        
        Expected columns:
        - Van ID
        - Type (Extra Large, Large, Step Van)
        - Opnal? Y/N (operational status)
        
        Args:
            file_path: Path to Daily Summary Excel file.
            sheet_name: Name of sheet containing vehicle status.
            
        Returns:
            DataFrame with Vehicle Status data.
        """
        logger.info(f"Loading Vehicle Status from {file_path}, sheet: {sheet_name}")
        
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            
            # Verify required columns
            required_cols = ["Van ID", "Type", "Opnal? Y/N"]
            missing_cols = [col for col in required_cols if col not in df.columns]
            
            if missing_cols:
                # Try alternative names
                if "Vehicle ID" in df.columns and "Van ID" not in df.columns:
                    df.rename(columns={"Vehicle ID": "Van ID"}, inplace=True)
                if "Vehicle Type" in df.columns and "Type" not in df.columns:
                    df.rename(columns={"Vehicle Type": "Type"}, inplace=True)
                if "Operational" in df.columns and "Opnal? Y/N" not in df.columns:
                    df.rename(columns={"Operational": "Opnal? Y/N"}, inplace=True)
                
                # Check again
                missing_cols = [col for col in required_cols if col not in df.columns]
                if missing_cols:
                    raise ValueError(f"Missing required columns in Vehicle Status: {missing_cols}")
            
            self.vehicle_status_data = df
            logger.info(f"Loaded {len(df)} vehicles from Vehicle Status")
            
            return df
            
        except Exception as e:
            logger.error(f"Failed to load Vehicle Status: {e}")
            raise
    
    def load_vehicle_log(
        self,
        file_path: str,
        sheet_name: str = "Vehicle Log"
    ) -> pd.DataFrame:
        """Load Vehicle Log from Daily Summary file.
        
        Expected columns (will auto-map from actual column names):
        - Van ID
        - VIN  
        - GeoTab (actual column name in Vehicle Log sheet)
        - Branded or Rental
        
        Args:
            file_path: Path to Daily Summary Excel file.
            sheet_name: Name of sheet containing vehicle log.
            
        Returns:
            DataFrame with Vehicle Log data.
        """
        logger.info(f"Loading Vehicle Log from {file_path}, sheet: {sheet_name}")
        
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            logger.info(f"Original Vehicle Log columns: {list(df.columns)}")
            
            # Map to expected column names based on actual Vehicle Log structure
            column_mapping = {}
            col_lookup = {col.lower().strip(): col for col in df.columns}

            # Standard column remapping
            for key, target in [("van id", "Van ID"), ("vehicle id", "Van ID"), ("vin", "VIN"), ("geotab", "GeoTab"), ("geotab code", "GeoTab")]:
                if key in col_lookup:
                    column_mapping[col_lookup[key]] = target

            brand_column_source = None
            for key in ("van type", "branded or rental", "brand or rental", "brand/rental"):
                if key in col_lookup:
                    brand_column_source = col_lookup[key]
                    column_mapping[brand_column_source] = "Branded or Rental"
                    break

            # Apply column mapping
            if column_mapping:
                df.rename(columns=column_mapping, inplace=True)
                logger.info(f"Applied column mapping: {column_mapping}")

            # If brand column wasn't mapped, attempt to derive from obvious alternatives
            if "Branded or Rental" not in df.columns and brand_column_source is None:
                logger.warning("Vehicle Log is missing a Branded/Rental (Van Type) column; brand prioritization will be limited.")

            # Verify we have the required columns after mapping
            required_cols = ["Van ID", "VIN", "GeoTab", "Branded or Rental"]
            missing_cols = [col for col in required_cols if col not in df.columns]
            
            if missing_cols:
                logger.warning(f"Missing columns in Vehicle Log after mapping: {missing_cols}")
                logger.info(f"Available columns: {list(df.columns)}")
                # Still continue with available data
            
            # Filter to only the columns we need (if they exist)
            available_cols = [col for col in required_cols if col in df.columns]
            if available_cols:
                df = df[available_cols]
            
            # Log sample data to verify content
            if len(df) > 0:
                logger.info(f"Sample Vehicle Log data (first row):")
                first_row = df.iloc[0]
                for col in df.columns:
                    logger.debug(f"  {col}: {first_row[col]}")
            
            self.vehicle_log_data = df
            logger.info(f"Loaded {len(df)} vehicles from Vehicle Log with columns: {list(df.columns)}")
            
            return df
            
        except Exception as e:
            logger.warning(f"Failed to load Vehicle Log: {e}")
            # Return empty dataframe so allocation can continue
            self.vehicle_log_data = pd.DataFrame(columns=["Van ID", "VIN", "GeoTab", "Branded or Rental"])
            return self.vehicle_log_data
    
    def get_van_type(self, service_type: str) -> Optional[str]:
        """Map service type to required van type.
        
        Exactly matches GAS logic:
        - Specific service types map to specific van types
        - Nursery Routes always get "Large" vans
        
        Args:
            service_type: Service type from Day of Ops.
            
        Returns:
            Required van type or None if no match.
        """
        if pd.isna(service_type):
            return None
        
        service_type = str(service_type).strip()
        
        # Check exact mapping first
        if service_type in self.SERVICE_TYPE_TO_VAN_TYPE:
            return self.SERVICE_TYPE_TO_VAN_TYPE[service_type]
        
        # Special case for Nursery Routes
        if "Nursery Route Level" in service_type:
            return "Large"
        
        logger.debug(f"No van type mapping for service type: {service_type}")
        return None
    
    def _normalize_brand_label(self, value: Any) -> Optional[str]:
        """Normalize raw Branded/Rental values to priority buckets."""
        if value is None:
            return None
        try:
            if pd.isna(value):
                return None
        except Exception:
            pass
        text = str(value).strip().lower()
        if not text:
            return None
        if "brand" in text:
            return "branded"
        if "rent" in text:
            return "rental"
        return None
    
    def _build_brand_priority_map(self, operational_vehicles: pd.DataFrame) -> Dict[str, str]:
        """Build a mapping of van IDs to branded/rental priority labels."""
        brand_map: Dict[str, str] = {}

        def _update_from_df(df: pd.DataFrame):
            if df is None or df.empty:
                return
            colmap = {c.lower(): c for c in df.columns}
            van_col = colmap.get("van id") or colmap.get("vehicle id")
            brand_col = None
            for key in ("branded or rental", "brand or rental", "brand/rental", "van type"):
                if key in colmap:
                    brand_col = colmap[key]
                    break
            if not van_col or not brand_col:
                return
            for _, row in df.iterrows():
                van_id = str(row.get(van_col, "") or "").strip()
                if not van_id:
                    continue
                label = self._normalize_brand_label(row.get(brand_col, ""))
                if label:
                    brand_map.setdefault(van_id, label)

        _update_from_df(self.vehicle_log_data)
        _update_from_df(operational_vehicles)

        return brand_map
    
    def filter_bway_routes(self) -> pd.DataFrame:
        """Filter Day of Ops for DSP = 'BWAY' routes.
        
        Returns:
            DataFrame with BWAY routes only.
        """
        if self.day_of_ops_data is None:
            raise ValueError("Must load Day of Ops before filtering")
        
        bway_routes = self.day_of_ops_data[
            self.day_of_ops_data["DSP"] == "BWAY"
        ].copy()
        
        logger.info(f"Filtered {len(bway_routes)} BWAY routes from Day of Ops")
        return bway_routes
    
    def allocate_vehicles_to_routes(self, bway_routes: pd.DataFrame = None) -> Tuple[List[Dict], List[str]]:
        """Allocate vehicles to routes using GAS logic.
        
        Process:
        1. Filter Day of Ops for DSP = "BWAY"
        2. Filter vehicles for Operational = "Y"
        3. Group vehicles by type
        4. Assign first available vehicle of matching type to each route
        
        Returns:
            Tuple of (allocation_results, assigned_van_ids)
        """
        if self.vehicle_status_data is None:
            raise ValueError("Must load Vehicle Status before allocation")
        
        # Use provided bway_routes or filter from loaded data
        if bway_routes is None:
            if self.day_of_ops_data is None:
                raise ValueError("Must load Day of Ops or provide filtered routes")
            bway_routes = self.filter_bway_routes()
        
        logger.info(f"Allocating vehicles for {len(bway_routes)} BWAY routes")
        
        # Step 2: Filter for operational vehicles
        logger.info("Filtering vehicles for Operational = 'Y'")
        operational_vehicles = self.vehicle_status_data[
            self.vehicle_status_data["Opnal? Y/N"].astype(str).str.upper() == "Y"
        ].copy()
        logger.info(f"Found {len(operational_vehicles)} operational vehicles")
        
        # Step 3: Group vehicles by type
        brand_priority_map = self._build_brand_priority_map(operational_vehicles)
        vehicle_groups = defaultdict(lambda: {"branded": [], "rental": []})

        def _normalize_van_type(value):
            if pd.isna(value):
                return ""
            return str(value).strip()

        def _group_total(group: dict[str, list]) -> int:
            return len(group["branded"]) + len(group["rental"])

        for _, vehicle in operational_vehicles.iterrows():
            van_type = _normalize_van_type(vehicle.get("Type", ""))
            if not van_type:
                continue
            record = vehicle.to_dict()
            van_id = str(record.get("Van ID", "") or "").strip()
            brand_label = brand_priority_map.get(van_id, "rental")
            tier = "branded" if brand_label == "branded" else "rental"
            vehicle_groups[van_type][tier].append(record)

        logger.info(
            f"Vehicle groups (branded priority): {[(k, _group_total(v), len(v['branded'])) for k, v in vehicle_groups.items()]}"
        )
        
        # Step 4: Allocate vehicles to routes
        allocation_results = []
        assigned_van_ids = []
        
        logger.info(f"Starting allocation for {len(bway_routes)} routes")
        
        for idx, (_, route) in enumerate(bway_routes.iterrows()):
            service_type = route["Service Type"]
            required_van_type = self.get_van_type(service_type)
            
            if not required_van_type:
                logger.debug(f"No van type for service '{service_type}' (Route: {route['Route Code']})")
                continue
            
            group = vehicle_groups[required_van_type]
            if _group_total(group) == 0:
                logger.debug(f"No available '{required_van_type}' for route {route['Route Code']}")
                continue
            
            # Assign first available vehicle of correct type prioritizing branded assets
            if group["branded"]:
                assigned_vehicle = group["branded"].pop(0)
            else:
                assigned_vehicle = group["rental"].pop(0)
            assigned_van_ids.append(assigned_vehicle["Van ID"])
            
            # Create allocation result (matches GAS structure)
            result = {
                "Route Code": route["Route Code"],
                "Service Type": service_type,
                "DSP": route["DSP"],
                "Wave": route["Wave"],
                "Staging Location": route["Staging Location"],
                "Van ID": assigned_vehicle["Van ID"],
                "Device Name": assigned_vehicle["Van ID"],  # GAS uses Van ID as Device Name
                "Van Type": assigned_vehicle["Type"],
                "Operational": assigned_vehicle["Opnal? Y/N"],
                "Associate Name": ""  # Will be filled from Daily Routes
            }
            
            allocation_results.append(result)
            logger.debug(f"Assigned {assigned_vehicle['Van ID']} to route {route['Route Code']}")
        
        logger.info(f"Allocated {len(allocation_results)} vehicles to routes")
        if allocation_results:
            logger.debug(f"Sample allocation result: {allocation_results[0]}")
        
        self.allocation_results = allocation_results
        self.assigned_van_ids = assigned_van_ids
        
        # Validate for duplicates
        validation_result = self.duplicate_validator.validate_allocations(allocation_results)
        if validation_result.has_duplicates():
            logger.warning(f"Duplicate validation: {validation_result.get_summary()}")
            # Mark duplicates in results
            self.allocation_results = self.duplicate_validator.mark_duplicates_in_results(
                allocation_results, validation_result
            )
        
        return allocation_results, assigned_van_ids
    
    def map_driver_names(self) -> List[Dict]:
        """Map driver names from Daily Routes to allocation results.
        
        Returns:
            Updated allocation results with driver names.
        """
        return self.update_with_driver_names()
    
    def update_with_driver_names(self) -> List[Dict]:
        """Update allocation results with driver names from Daily Routes.
        
        Matches GAS updateResultsWithDailyRoutes logic.
        
        Returns:
            Updated allocation results with driver names.
        """
        if self.daily_routes_data is None:
            logger.warning("No Daily Routes data loaded")
            return self.allocation_results
        
        # Build route -> driver mapping
        route_driver_map = {}
        for _, row in self.daily_routes_data.iterrows():
            route_code = row.get("Route Code", row.get("Route code", ""))
            driver_name = row.get("Driver Name", row.get("Driver name", "N/A"))
            if route_code:
                route_driver_map[str(route_code).strip()] = str(driver_name).strip()
        
        logger.info(f"Built route-driver mapping with {len(route_driver_map)} entries")
        
        # Update allocation results with driver names
        today = datetime.now().strftime("%m/%d/%Y")
        
        for result in self.allocation_results:
            route_code = result["Route Code"]
            associate_name = route_driver_map.get(route_code, "N/A")
            result["Associate Name"] = associate_name
            
            # Generate unique identifier (GAS format)
            unique_id = f"{today}|{route_code}|{associate_name}|{result['Van ID']}"
            result["Unique Identifier"] = unique_id
            
            logger.debug(f"Updated route {route_code} with driver {associate_name}")
        
        return self.allocation_results
    
    def identify_unassigned_vehicles(self) -> pd.DataFrame:
        """Identify operational vehicles that were not assigned.
        
        Returns:
            DataFrame of unassigned operational vehicles.
        """
        if self.vehicle_status_data is None:
            raise ValueError("Vehicle Status data not loaded")
        
        # Get operational vehicles
        operational = self.vehicle_status_data[
            self.vehicle_status_data["Opnal? Y/N"].astype(str).str.upper() == "Y"
        ]
        
        # Filter out assigned vehicles
        unassigned = operational[
            ~operational["Van ID"].isin(self.assigned_van_ids)
        ]
        
        logger.info(f"Found {len(unassigned)} unassigned operational vehicles")
        self.unassigned_vehicles = unassigned
        
        return unassigned
    
    def create_allocation_result(self) -> AllocationResult:
        """Create an AllocationResult object compatible with the Python system.
        
        Returns:
            AllocationResult object with allocation data.
        """
        # Group allocations by driver
        driver_vehicles = defaultdict(list)
        for result in self.allocation_results:
            driver_name = result.get("Associate Name", "N/A")
            van_id = result["Van ID"]
            driver_vehicles[driver_name].append(van_id)
        
        # Get unallocated vehicle IDs
        if isinstance(self.unassigned_vehicles, pd.DataFrame) and not self.unassigned_vehicles.empty:
            unallocated_ids = list(self.unassigned_vehicles["Van ID"])
        else:
            unallocated_ids = []
        
        # Check for duplicate warnings in results
        warnings = []
        errors = []
        duplicate_count = 0
        
        for result in self.allocation_results:
            if result.get("Validation Status") == "DUPLICATE":
                duplicate_count += 1
                warning_msg = result.get("Validation Warning", "")
                if warning_msg and warning_msg not in warnings:
                    warnings.append(warning_msg)
        
        if duplicate_count > 0:
            warnings.insert(0, f"Found {duplicate_count} duplicate vehicle assignments")
        
        # Driver counts for metadata
        total_driver_count = 0
        if self.daily_routes_data is not None and not self.daily_routes_data.empty:
            cols = {c.lower().strip(): c for c in self.daily_routes_data.columns}
            driver_col = cols.get("driver name") or cols.get("driver")
            if driver_col:
                total_driver_count = (
                    self.daily_routes_data[driver_col]
                    .dropna()
                    .astype(str)
                    .str.strip()
                    .replace({"": pd.NA})
                    .dropna()
                    .nunique()
                )
        if total_driver_count == 0:
            total_driver_count = len(driver_vehicles)
        active_driver_count = len([v for v in driver_vehicles.values() if v])
        allocation_rate = (len(self.assigned_van_ids) / len(self.allocation_results) * 100) if self.allocation_results else 0.0
        
        # Create AllocationResult
        allocation_result = AllocationResult(
            request_id=f"GAS_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
            allocations=dict(driver_vehicles),
            unallocated_vehicles=unallocated_ids,
            status=AllocationStatus.COMPLETED,
            timestamp=datetime.now(),
            warnings=warnings,
            errors=errors,
            metadata={
                "source": "GAS_Compatible",
                "total_routes": len(self.allocation_results),
                "total_assigned": len(self.assigned_van_ids),
                "total_unassigned": len(unallocated_ids),
                "duplicate_count": duplicate_count,
                "total_drivers": int(total_driver_count),
                "active_drivers": int(active_driver_count),
                "allocation_rate": round(allocation_rate, 2),
                "detailed_results": self.allocation_results  # Store in metadata instead
            }
        )
        
        return allocation_result
    
    def record_history(
        self,
        allocation_result: AllocationResult,
        files: Optional[Dict[str, str]] = None,
        error: Optional[str] = None
    ) -> None:
        """Persist the allocation run to history with consistent defaults."""
        duplicate_conflicts = 0
        metadata = getattr(allocation_result, 'metadata', {}) or {}
        if metadata:
            duplicate_conflicts = metadata.get('duplicate_count', 0)
        if isinstance(duplicate_conflicts, (list, tuple, set)):
            duplicate_conflicts = len([item for item in duplicate_conflicts if item is not None])
        try:
            self.history_service.save_allocation(
                result=allocation_result,
                engine_name="GASCompatibleAllocator",
                files=files or {},
                duplicate_conflicts=duplicate_conflicts,
                error=error
            )
            logger.info("Allocation saved to history")
        except Exception as exc:
            logger.error(f"Failed to record allocation history: {exc}")
    
    def run_full_allocation(
        self,
        day_of_ops_file: str,
        daily_routes_file: str,
        vehicle_status_file: str,
        output_file: str = None
    ) -> AllocationResult:
        """Run the complete GAS-compatible allocation process.
        
        Args:
            day_of_ops_file: Path to Day of Ops Excel file.
            daily_routes_file: Path to Daily Routes Excel file.
            vehicle_status_file: Path to file with Vehicle Status sheet.
            output_file: Optional path for output file.
            
        Returns:
            AllocationResult with complete allocation data.
        """
        logger.info("Starting GAS-compatible allocation process")
        start_ts = time.perf_counter()
        
        # Load all data
        self.load_day_of_ops(day_of_ops_file)
        self.load_daily_routes(daily_routes_file)
        self.load_vehicle_status(vehicle_status_file)
        
        # Try to load Vehicle Log from same file as Vehicle Status
        self.load_vehicle_log(vehicle_status_file)
        
        # Perform allocation
        self.allocate_vehicles_to_routes()
        
        # Update with driver names
        self.update_with_driver_names()
        
        # Identify unassigned vehicles
        self.identify_unassigned_vehicles()
        
        # Create allocation result
        result = self.create_allocation_result()
        processing_time = time.perf_counter() - start_ts
        result.metadata["processing_time_seconds"] = round(processing_time, 4)
        
        # Save to allocation history
        self.record_history(
            allocation_result=result,
            files={
                "day_of_ops": day_of_ops_file,
                "daily_routes": daily_routes_file,
                "vehicle_status": vehicle_status_file
            }
        )
        
        # If output file specified, write results
        if output_file:
            self.write_results_to_excel(result, output_file)
        
        logger.info("GAS-compatible allocation complete")
        return result
    
    def write_results_to_excel(self, allocation_result: AllocationResult, output_file: str, create_results_file: bool = True) -> Optional[str]:
        """Write allocation results to Excel.
        
        Updates Daily Details in the Daily Summary Log and optionally creates
        a separate results file in the outputs directory.
        
        Args:
            allocation_result: The allocation results.
            output_file: Path to Daily Summary Log file.
            create_results_file: Whether to create separate results file.
        """
        logger.info(f"Writing results to {output_file}")
        
        writer = DailyDetailsWriter()
        writer.initialize()
        
        if not writer.validate():
            raise ValueError("Failed to initialize DailyDetailsWriter")
        
        allocation_date = date.today()
        
        # Prepare vehicle log dictionary from Vehicle Log sheet
        vehicle_log_dict = {}
        
        # First try to use actual Vehicle Log data
        if self.vehicle_log_data is not None and not self.vehicle_log_data.empty:
            logger.info(f"Building vehicle_log_dict from {len(self.vehicle_log_data)} Vehicle Log entries")
            logger.debug(f"Vehicle Log columns available: {list(self.vehicle_log_data.columns)}")
            
            for _, vehicle in self.vehicle_log_data.iterrows():
                # Extract scalar values from Series to avoid type issues
                # Handle cases where the value might be a Series object
                van_id_val = vehicle.get("Van ID", "")
                if isinstance(van_id_val, pd.Series):
                    van_id_val = van_id_val.iloc[0] if len(van_id_val) > 0 else ""
                van_id = str(van_id_val).strip() if pd.notna(van_id_val) else ""
                
                if not van_id:
                    logger.warning("Skipping Vehicle Log entry with missing Van ID")
                    continue
                    
                # Ensure we get string values, not Series objects
                def extract_value(val):
                    """Extract scalar value from potentially Series object."""
                    if isinstance(val, pd.Series):
                        return val.iloc[0] if len(val) > 0 else ""
                    return val
                
                vin_val = extract_value(vehicle.get("VIN", ""))
                geotab_val = extract_value(vehicle.get("GeoTab", ""))
                brand_rental_val = extract_value(vehicle.get("Branded or Rental", ""))
                
                vin_value = str(vin_val).strip() if pd.notna(vin_val) else ""
                geotab_value = str(geotab_val).strip() if pd.notna(geotab_val) else ""
                brand_rental_value = str(brand_rental_val).strip() if pd.notna(brand_rental_val) else ""
                
                vehicle_log_dict[van_id] = {
                    "vin": vin_value,
                    "geotab": geotab_value,
                    "brand_or_rental": brand_rental_value,
                    "vehicle_type": ""  # This comes from allocation, not Vehicle Log
                }
                
                # Log sample data for first few vehicles
                if len(vehicle_log_dict) <= 3:
                    logger.debug(f"Vehicle {van_id}: VIN={vin_value}, GeoTab={geotab_value}, Type={brand_rental_value}")
            
            # Log summary statistics
            vehicles_with_geotab = sum(1 for v in vehicle_log_dict.values() if v["geotab"])
            vehicles_with_type = sum(1 for v in vehicle_log_dict.values() if v["brand_or_rental"])
            logger.info(f"Vehicle Log summary: {len(vehicle_log_dict)} total, {vehicles_with_geotab} with GeoTab, {vehicles_with_type} with Type")
        # Fallback to Vehicle Status if no Vehicle Log
        elif self.vehicle_status_data is not None:
            logger.warning("No Vehicle Log data available, using limited data from Vehicle Status")
            
            # Define extract_value function for this section too
            def extract_value(val):
                """Extract scalar value from potentially Series object."""
                if isinstance(val, pd.Series):
                    return val.iloc[0] if len(val) > 0 else ""
                return val
            
            for _, vehicle in self.vehicle_status_data.iterrows():
                # Extract scalar values from Series to avoid type issues
                van_id_val = extract_value(vehicle.get("Van ID", ""))
                van_id = str(van_id_val).strip() if pd.notna(van_id_val) else ""
                if not van_id:
                    continue
                
                vin_val = extract_value(vehicle.get("VIN", ""))
                geotab_val = extract_value(vehicle.get("GeoTab Code", ""))
                type_val = extract_value(vehicle.get("Type", ""))
                
                vehicle_log_dict[van_id] = {
                    "vin": str(vin_val).strip() if pd.notna(vin_val) else "",
                    "geotab": str(geotab_val).strip() if pd.notna(geotab_val) else "",
                    "brand_or_rental": "",  # Not available in Vehicle Status
                    "vehicle_type": str(type_val).strip() if pd.notna(type_val) else ""
                }
        
        # Use the writer to append to existing or create new file
        if Path(output_file).exists():
            # Append to existing file
            success = writer.append_to_existing_file(
                file_path=output_file,
                allocation_result=allocation_result,
                allocation_date=allocation_date,
                vehicle_log_dict=vehicle_log_dict
            )
        else:
            # Create new file with all sheets
            from openpyxl import Workbook
            wb = Workbook()
            
            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
            
            # Create Daily Details sheet
            daily_details = wb.create_sheet("Daily Details")
            writer._setup_daily_details_headers(daily_details)
            
            # Save and then append
            wb.save(output_file)
            success = writer.append_to_existing_file(
                file_path=output_file,
                allocation_result=allocation_result,
                allocation_date=allocation_date,
                vehicle_log_dict=vehicle_log_dict
            )
        
        if success:
            logger.info(f"Successfully wrote results to {output_file}")
            
            # Unassigned vehicles sheet is now created in the separate results file
        else:
            logger.error(f"Failed to write results to {output_file}")
            
        # Create separate results file if requested
        results_file_path = None
        if success and create_results_file:
            try:
                results_file_path = self.create_results_output_file(allocation_result, allocation_date)
                logger.info(f"Created separate results file: {results_file_path}")
            except Exception as e:
                logger.error(f"Failed to create separate results file: {e}")
                # Don't fail the whole operation if results file creation fails
                pass
        
        return results_file_path
    
    def create_output_file(self, output_file: str, allocation_date: date = None):
        """Create output file with allocation results.
        
        Args:
            output_file: Path to output Excel file.
            allocation_date: Date for the allocation (defaults to today).
        """
        if allocation_date is None:
            allocation_date = date.today()
        
        # Create allocation result if not already done
        if not self.allocation_results:
            logger.warning("No allocation results to write")
            # Still create an empty result for consistency
            result = self.create_allocation_result()
            return result, None
        
        # Update with driver names if not done
        if self.allocation_results and not any(r.get("Associate Name") for r in self.allocation_results):
            self.update_with_driver_names()
        
        # Identify unassigned if not done
        if not hasattr(self, 'unassigned_vehicles') or self.unassigned_vehicles is None:
            self.identify_unassigned_vehicles()
        
        # Create the allocation result object
        result = self.create_allocation_result()
        
        logger.info(f"Created allocation result with {len(self.allocation_results)} detailed results")
        if self.allocation_results:
            logger.debug(f"Sample allocation data: {self.allocation_results[0]}")
        
        # Write to Excel and get results file path
        results_file_path = self.write_results_to_excel(result, output_file)
        
        return result, results_file_path
    
    def create_results_output_file(self, allocation_result: AllocationResult, allocation_date: date = None) -> str:
        """Create a separate results file in the outputs directory.
        
        Args:
            allocation_result: The allocation results.
            allocation_date: Date for the allocation (defaults to today).
            
        Returns:
            Path to the created results file.
        """
        if allocation_date is None:
            allocation_date = date.today()
            
        # Prepare vehicle log dictionary
        vehicle_log_dict = self._build_vehicle_log_dict()
        
        # Ensure unassigned vehicles have been identified
        if not isinstance(self.unassigned_vehicles, pd.DataFrame):
            logger.info("Identifying unassigned vehicles before creating results file")
            self.identify_unassigned_vehicles()
        
        # Get the unassigned vehicles DataFrame
        unassigned_df = self.unassigned_vehicles
        if not isinstance(unassigned_df, pd.DataFrame):
            logger.warning("Failed to get unassigned vehicles DataFrame, using empty DataFrame")
            unassigned_df = pd.DataFrame()
                
        # Create the results file
        results_file_path = self.output_writer.create_results_file(
            allocation_result=allocation_result,
            unassigned_vehicles=unassigned_df,
            vehicle_log_dict=vehicle_log_dict,
            allocation_date=allocation_date
        )
        
        return results_file_path
        
    def _build_vehicle_log_dict(self) -> Dict[str, Dict]:
        """Build vehicle log dictionary from available data.
        
        Returns:
            Dictionary mapping van IDs to vehicle information.
        """
        vehicle_log_dict = {}
        
        # First try to use actual Vehicle Log data
        if self.vehicle_log_data is not None and not self.vehicle_log_data.empty:
            for _, vehicle in self.vehicle_log_data.iterrows():
                van_id = str(vehicle.get("Van ID", "")).strip()
                if not van_id:
                    continue
                    
                # Extract values safely
                def extract_value(val):
                    if isinstance(val, pd.Series):
                        return val.iloc[0] if len(val) > 0 else ""
                    return val
                
                vehicle_log_dict[van_id] = {
                    "vin": str(extract_value(vehicle.get("VIN", ""))).strip() if pd.notna(extract_value(vehicle.get("VIN", ""))) else "",
                    "geotab": str(extract_value(vehicle.get("GeoTab", ""))).strip() if pd.notna(extract_value(vehicle.get("GeoTab", ""))) else "",
                    "brand_or_rental": str(extract_value(vehicle.get("Branded or Rental", ""))).strip() if pd.notna(extract_value(vehicle.get("Branded or Rental", ""))) else "",
                    "vehicle_type": ""
                }
                
        # Fallback to Vehicle Status if no Vehicle Log
        elif self.vehicle_status_data is not None:
            for _, vehicle in self.vehicle_status_data.iterrows():
                van_id = str(vehicle.get("Van ID", "")).strip()
                if not van_id:
                    continue
                    
                vehicle_log_dict[van_id] = {
                    "vin": str(vehicle.get("VIN", "")).strip() if pd.notna(vehicle.get("VIN", "")) else "",
                    "geotab": str(vehicle.get("GeoTab Code", "")).strip() if pd.notna(vehicle.get("GeoTab Code", "")) else "",
                    "brand_or_rental": "",
                    "vehicle_type": str(vehicle.get("Type", "")).strip() if pd.notna(vehicle.get("Type", "")) else ""
                }
                
        return vehicle_log_dict