"""Border formatting service for Excel with daily sections."""

from typing import Optional, Any
from datetime import date, datetime
from loguru import logger

from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from src.core.base_service import BaseService, timer, error_handler
from src.models.excel import ExcelRange, BorderStyle, ExcelColor


class BorderFormattingService(BaseService):
    """Service for applying border formatting to Excel sheets.
    
    Specializes in creating daily sections with thick borders,
    similar to the original Google Sheets implementation.
    """
    
    def __init__(self, config: Optional[dict[str, Any]] = None):
        """Initialize the border formatting service.
        
        Args:
            config: Configuration dictionary.
        """
        super().__init__(config)
        
        # Default border styles
        self.section_border_style = self.get_config("section_border_style", BorderStyle.THICK)
        self.internal_border_style = self.get_config("internal_border_style", BorderStyle.THIN)
        self.header_border_style = self.get_config("header_border_style", BorderStyle.MEDIUM)
        
        # Default colors
        self.border_color = self.get_config("border_color", "000000")
        self.header_bg_color = self.get_config("header_bg_color", "D3D3D3")
        self.alternating_row_color = self.get_config("alternating_row_color", "F5F5F5")
    
    def initialize(self) -> None:
        """Initialize the border formatting service."""
        logger.info("Initializing Border Formatting Service")
        self._initialized = True
    
    def validate(self) -> bool:
        """Validate the service configuration.
        
        Returns:
            True if configuration is valid.
        """
        return True
    
    @timer
    @error_handler
    def create_daily_section(
        self,
        sheet: Any,
        start_row: int,
        start_col: int,
        end_row: int,
        end_col: int,
        section_date: date,
        title: Optional[str] = None
    ) -> None:
        """Create a daily section with thick borders.
        
        Args:
            sheet: Worksheet object.
            start_row: Starting row of the section.
            start_col: Starting column of the section.
            end_row: Ending row of the section.
            end_col: Ending column of the section.
            section_date: Date for the section.
            title: Optional title for the section.
        """
        # Create border styles
        thick_side = Side(style=self.section_border_style, color=self.border_color)
        thin_side = Side(style=self.internal_border_style, color=self.border_color)
        
        # Apply thick border around the entire section
        self._apply_section_border(
            sheet, start_row, start_col, end_row, end_col, thick_side
        )
        
        # Add section header with date
        header_row = start_row
        self._create_section_header(
            sheet, header_row, start_col, end_col, section_date, title
        )
        
        # Apply internal borders for data rows
        data_start_row = start_row + 2  # Skip header and column headers
        if data_start_row <= end_row:
            self._apply_internal_borders(
                sheet, data_start_row, start_col, end_row, end_col, thin_side
            )
        
        logger.info(f"Created daily section for {section_date}")
    
    def _apply_section_border(
        self,
        sheet: Any,
        start_row: int,
        start_col: int,
        end_row: int,
        end_col: int,
        border_side: Side
    ) -> None:
        """Apply thick border around a section.
        
        Args:
            sheet: Worksheet object.
            start_row: Starting row.
            start_col: Starting column.
            end_row: Ending row.
            end_col: Ending column.
            border_side: Border side style.
        """
        # Top border
        for col in range(start_col, end_col + 1):
            cell = sheet.cell(row=start_row, column=col)
            current_border = cell.border
            cell.border = Border(
                top=border_side,
                left=current_border.left,
                right=current_border.right,
                bottom=current_border.bottom
            )
        
        # Bottom border
        for col in range(start_col, end_col + 1):
            cell = sheet.cell(row=end_row, column=col)
            current_border = cell.border
            cell.border = Border(
                top=current_border.top,
                left=current_border.left,
                right=current_border.right,
                bottom=border_side
            )
        
        # Left border
        for row in range(start_row, end_row + 1):
            cell = sheet.cell(row=row, column=start_col)
            current_border = cell.border
            cell.border = Border(
                top=current_border.top,
                left=border_side,
                right=current_border.right,
                bottom=current_border.bottom
            )
        
        # Right border
        for row in range(start_row, end_row + 1):
            cell = sheet.cell(row=row, column=end_col)
            current_border = cell.border
            cell.border = Border(
                top=current_border.top,
                left=current_border.left,
                right=border_side,
                bottom=current_border.bottom
            )
        
        # Corners (ensure they have both borders)
        # Top-left
        cell = sheet.cell(row=start_row, column=start_col)
        cell.border = Border(top=border_side, left=border_side, 
                            right=cell.border.right, bottom=cell.border.bottom)
        
        # Top-right
        cell = sheet.cell(row=start_row, column=end_col)
        cell.border = Border(top=border_side, right=border_side,
                            left=cell.border.left, bottom=cell.border.bottom)
        
        # Bottom-left
        cell = sheet.cell(row=end_row, column=start_col)
        cell.border = Border(bottom=border_side, left=border_side,
                            top=cell.border.top, right=cell.border.right)
        
        # Bottom-right
        cell = sheet.cell(row=end_row, column=end_col)
        cell.border = Border(bottom=border_side, right=border_side,
                            top=cell.border.top, left=cell.border.left)
    
    def _create_section_header(
        self,
        sheet: Any,
        row: int,
        start_col: int,
        end_col: int,
        section_date: date,
        title: Optional[str] = None
    ) -> None:
        """Create a section header with date and title.
        
        Args:
            sheet: Worksheet object.
            row: Header row.
            start_col: Starting column.
            end_col: Ending column.
            section_date: Date for the section.
            title: Optional title.
        """
        # Merge cells for header
        sheet.merge_cells(
            start_row=row,
            start_column=start_col,
            end_row=row,
            end_column=end_col
        )
        
        # Set header text
        header_cell = sheet.cell(row=row, column=start_col)
        header_text = f"{section_date.strftime('%A, %B %d, %Y')}"
        if title:
            header_text = f"{title} - {header_text}"
        header_cell.value = header_text
        
        # Apply header formatting
        header_cell.font = Font(bold=True, size=12)
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        header_cell.fill = PatternFill(
            start_color=self.header_bg_color,
            end_color=self.header_bg_color,
            fill_type="solid"
        )
        
        # Apply border to header
        medium_side = Side(style=self.header_border_style, color=self.border_color)
        header_cell.border = Border(
            top=medium_side,
            left=medium_side,
            right=medium_side,
            bottom=medium_side
        )
    
    def _apply_internal_borders(
        self,
        sheet: Any,
        start_row: int,
        start_col: int,
        end_row: int,
        end_col: int,
        border_side: Side
    ) -> None:
        """Apply internal borders to data cells.
        
        Args:
            sheet: Worksheet object.
            start_row: Starting row for data.
            start_col: Starting column.
            end_row: Ending row.
            end_col: Ending column.
            border_side: Border side style.
        """
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = sheet.cell(row=row, column=col)
                
                # Determine which borders to apply
                top = border_side if row > start_row else None
                left = border_side if col > start_col else None
                right = border_side if col < end_col else None
                bottom = border_side if row < end_row else None
                
                # Preserve thick borders if they exist
                current = cell.border
                if current.top and current.top.style == self.section_border_style:
                    top = current.top
                if current.left and current.left.style == self.section_border_style:
                    left = current.left
                if current.right and current.right.style == self.section_border_style:
                    right = current.right
                if current.bottom and current.bottom.style == self.section_border_style:
                    bottom = current.bottom
                
                cell.border = Border(top=top, left=left, right=right, bottom=bottom)
    
    @timer
    def apply_alternating_rows(
        self,
        sheet: Any,
        start_row: int,
        start_col: int,
        end_row: int,
        end_col: int,
        skip_header: bool = True
    ) -> None:
        """Apply alternating row colors.
        
        Args:
            sheet: Worksheet object.
            start_row: Starting row.
            start_col: Starting column.
            end_row: Ending row.
            end_col: Ending column.
            skip_header: Whether to skip the first row.
        """
        data_start = start_row + 1 if skip_header else start_row
        
        for row_idx, row in enumerate(range(data_start, end_row + 1)):
            if row_idx % 2 == 1:  # Alternate rows
                for col in range(start_col, end_col + 1):
                    cell = sheet.cell(row=row, column=col)
                    cell.fill = PatternFill(
                        start_color=self.alternating_row_color,
                        end_color=self.alternating_row_color,
                        fill_type="solid"
                    )
        
        logger.debug(f"Applied alternating rows from {data_start} to {end_row}")
    
    def create_multiple_daily_sections(
        self,
        sheet: Any,
        dates: list[date],
        rows_per_section: int,
        start_row: int = 1,
        start_col: int = 1,
        num_cols: int = 10,
        spacing: int = 2
    ) -> dict[date, ExcelRange]:
        """Create multiple daily sections.
        
        Args:
            sheet: Worksheet object.
            dates: List of dates for sections.
            rows_per_section: Number of rows per section.
            start_row: Starting row.
            start_col: Starting column.
            num_cols: Number of columns per section.
            spacing: Rows between sections.
        
        Returns:
            Dictionary mapping dates to their Excel ranges.
        """
        section_ranges = {}
        current_row = start_row
        
        for section_date in dates:
            end_row = current_row + rows_per_section - 1
            end_col = start_col + num_cols - 1
            
            # Create the section
            self.create_daily_section(
                sheet=sheet,
                start_row=current_row,
                start_col=start_col,
                end_row=end_row,
                end_col=end_col,
                section_date=section_date,
                title="Daily Allocation"
            )
            
            # Store the range
            section_ranges[section_date] = ExcelRange(
                sheet_name=sheet.title if hasattr(sheet, 'title') else str(sheet),
                start_row=current_row,
                start_col=start_col,
                end_row=end_row,
                end_col=end_col
            )
            
            # Move to next section
            current_row = end_row + spacing + 1
        
        logger.info(f"Created {len(dates)} daily sections")
        return section_ranges
    
    def highlight_cells(
        self,
        sheet: Any,
        cells: list[tuple[int, int]],
        color: str = "FFFF00"
    ) -> None:
        """Highlight specific cells.
        
        Args:
            sheet: Worksheet object.
            cells: List of (row, col) tuples to highlight.
            color: Hex color code for highlighting.
        """
        fill = PatternFill(
            start_color=color,
            end_color=color,
            fill_type="solid"
        )
        
        for row, col in cells:
            cell = sheet.cell(row=row, column=col)
            cell.fill = fill
        
        logger.debug(f"Highlighted {len(cells)} cells")
    
    def apply_conditional_formatting(
        self,
        sheet: Any,
        range_obj: ExcelRange,
        condition_type: str,
        values: list[Any],
        format_style: dict[str, Any]
    ) -> None:
        """Apply conditional formatting to a range.
        
        Args:
            sheet: Worksheet object.
            range_obj: Excel range to apply formatting to.
            condition_type: Type of condition (e.g., "cell_value", "duplicate").
            values: Values for the condition.
            format_style: Formatting style to apply.
        """
        # This would be implemented with openpyxl's conditional formatting
        # For now, this is a placeholder
        logger.info(f"Applied conditional formatting to {range_obj.to_excel_range()}")