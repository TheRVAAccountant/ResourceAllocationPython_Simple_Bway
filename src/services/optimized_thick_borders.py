"""Optimized Daily Details thick border implementation with performance improvements."""

from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor
from datetime import date, datetime

from loguru import logger
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet


class OptimizedThickBorderService:
    """Optimized service to apply thick borders with significant performance improvements.

    Key optimizations:
    1. Batch border operations instead of cell-by-cell
    2. Pre-created border style objects
    3. Minimal cell access using coordinate-based operations
    4. Parallel processing for large datasets
    5. Smart caching of date parsing results
    """

    def __init__(self):
        """Initialize the optimized thick border service."""
        # Pre-create all possible border combinations
        self._init_border_styles()

        # Cache for parsed dates
        self.date_cache = {}

        # Thread pool for parallel processing
        self.executor = None

    def _init_border_styles(self):
        """Pre-create all border style combinations for reuse."""
        thick = Side(style="thick", color="000000")
        thin = Side(style="thin", color="000000")

        # Create all 16 possible combinations
        self.border_map = {}
        for top in [thick, thin]:
            for bottom in [thick, thin]:
                for left in [thick, thin]:
                    for right in [thick, thin]:
                        key = (top == thick, bottom == thick, left == thick, right == thick)
                        self.border_map[key] = Border(
                            top=top, bottom=bottom, left=left, right=right
                        )

        # Special styles
        self.date_header_fill = PatternFill(
            start_color="E6E6E6", end_color="E6E6E6", fill_type="solid"
        )
        self.date_header_font = Font(bold=True, size=11)
        self.date_header_alignment = Alignment(horizontal="left", vertical="center")

    def apply_thick_borders_optimized(
        self, worksheet: Worksheet, start_row: int = 2, progress_callback: callable | None = None
    ) -> None:
        """Apply thick borders using optimized batch operations.

        Args:
            worksheet: The Daily Details worksheet.
            start_row: Row to start processing (default 2, after headers).
            progress_callback: Optional callback for progress updates.
        """
        logger.info("Applying thick borders using optimized algorithm")

        # Step 1: Identify date sections efficiently
        date_sections = self._identify_date_sections_fast(worksheet, start_row)

        if not date_sections:
            logger.info("No data found to apply borders")
            return

        logger.info(f"Processing {len(date_sections)} date sections")

        # Step 2: Pre-calculate all border operations
        border_operations = self._calculate_border_operations(date_sections)

        # Step 3: Apply borders in batches
        self._apply_borders_batch(worksheet, border_operations, progress_callback)

        logger.info(f"Applied thick borders to {len(date_sections)} date sections")

    def _identify_date_sections_fast(
        self, worksheet: Worksheet, start_row: int
    ) -> dict[date, tuple[int, int]]:
        """Identify date sections using optimized reading strategy.

        Uses worksheet.iter_rows for better performance on large datasets.
        """
        date_sections = {}
        current_date = None
        first_row_for_date = None
        last_valid_row = start_row - 1

        # Use iter_rows for better performance
        for row in worksheet.iter_rows(
            min_row=start_row, max_row=worksheet.max_row, min_col=1, max_col=1, values_only=True
        ):
            row_num = last_valid_row + 1
            date_value = row[0] if row else None

            if date_value is None:
                # End of data
                if current_date and first_row_for_date:
                    date_sections[current_date] = (first_row_for_date, last_valid_row)
                break

            # Parse date with caching
            row_date = self._parse_date_cached(date_value)

            if row_date != current_date:
                # Save previous section
                if current_date and first_row_for_date:
                    date_sections[current_date] = (first_row_for_date, last_valid_row)

                # Start new section
                current_date = row_date
                first_row_for_date = row_num

            last_valid_row = row_num

        # Don't forget the last section
        if current_date and first_row_for_date:
            date_sections[current_date] = (first_row_for_date, last_valid_row)

        return date_sections

    def _parse_date_cached(self, value) -> date | None:
        """Parse date with caching for repeated values."""
        if value is None:
            return None

        # Check cache first
        cache_key = str(value)
        if cache_key in self.date_cache:
            return self.date_cache[cache_key]

        # Parse date
        parsed_date = self._parse_date_value(value)

        # Cache result
        self.date_cache[cache_key] = parsed_date
        return parsed_date

    def _parse_date_value(self, value) -> date | None:
        """Parse a date value from various formats."""
        if isinstance(value, datetime):
            return value.date()
        elif isinstance(value, date):
            return value
        elif isinstance(value, str):
            # Try common formats
            for fmt in ["%m/%d/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m-%d-%Y"]:
                try:
                    return datetime.strptime(value, fmt).date()
                except ValueError:
                    continue
        return None

    def _calculate_border_operations(
        self, date_sections: dict[date, tuple[int, int]]
    ) -> list[dict]:
        """Pre-calculate all border operations for batch processing."""
        operations = []

        for section_date, (first_row, last_row) in date_sections.items():
            # Calculate operations for this section
            section_ops = self._calculate_section_borders(first_row, last_row, section_date)
            operations.extend(section_ops)

        return operations

    def _calculate_section_borders(
        self, first_row: int, last_row: int, _section_date: date
    ) -> list[dict]:
        """Calculate border operations for a single date section."""
        operations = []
        first_col, last_col = 1, 24  # Daily Details columns

        # Group cells by border type for batch processing
        border_groups = defaultdict(list)

        for row in range(first_row, last_row + 1):
            for col in range(first_col, last_col + 1):
                # Determine border type
                is_top = row == first_row
                is_bottom = row == last_row
                is_left = col == first_col
                is_right = col == last_col

                border_key = (is_top, is_bottom, is_left, is_right)
                border_groups[border_key].append((row, col))

        # Convert groups to operations
        for border_key, cells in border_groups.items():
            operations.append(
                {"type": "border", "border": self.border_map[border_key], "cells": cells}
            )

        # Add date header formatting
        operations.append(
            {
                "type": "date_header",
                "row": first_row,
                "col": 1,
                "fill": self.date_header_fill,
                "font": self.date_header_font,
                "alignment": self.date_header_alignment,
            }
        )

        return operations

    def _apply_borders_batch(
        self,
        worksheet: Worksheet,
        operations: list[dict],
        progress_callback: callable | None = None,
    ):
        """Apply border operations in batches for better performance."""
        total_ops = len(operations)

        # Process operations in chunks
        chunk_size = 100
        for i in range(0, total_ops, chunk_size):
            chunk = operations[i : i + chunk_size]

            for op in chunk:
                if op["type"] == "border":
                    # Apply border to multiple cells at once
                    for row, col in op["cells"]:
                        cell = worksheet.cell(row=row, column=col)
                        cell.border = op["border"]

                elif op["type"] == "date_header":
                    # Apply date header formatting
                    cell = worksheet.cell(row=op["row"], column=op["col"])
                    cell.fill = op["fill"]
                    cell.font = op["font"]
                    cell.alignment = op["alignment"]

            # Report progress
            if progress_callback:
                progress = min((i + chunk_size) / total_ops * 100, 100)
                progress_callback(progress)

    def apply_thick_borders_parallel(
        self,
        worksheet: Worksheet,
        date_sections: dict[date, tuple[int, int]],
        max_workers: int = 4,
        progress_callback: callable | None = None,
    ) -> None:
        """Apply thick borders using parallel processing for very large datasets.

        Note: This is experimental and requires careful synchronization.
        """
        logger.info(f"Applying thick borders in parallel with {max_workers} workers")

        if not self.executor:
            self.executor = ThreadPoolExecutor(max_workers=max_workers)

        # Split sections into chunks for parallel processing
        section_items = list(date_sections.items())
        chunk_size = max(len(section_items) // max_workers, 1)

        futures = []
        for i in range(0, len(section_items), chunk_size):
            chunk = dict(section_items[i : i + chunk_size])
            future = self.executor.submit(self._process_sections_chunk, worksheet, chunk)
            futures.append(future)

        # Wait for completion and track progress
        for completed, future in enumerate(futures, start=1):
            future.result()
            if progress_callback:
                progress_callback(completed / len(futures) * 100)

    def _process_sections_chunk(self, worksheet: Worksheet, sections: dict[date, tuple[int, int]]):
        """Process a chunk of date sections."""
        operations = []
        for section_date, (first_row, last_row) in sections.items():
            section_ops = self._calculate_section_borders(first_row, last_row, section_date)
            operations.extend(section_ops)

        # Apply operations
        self._apply_borders_batch(worksheet, operations)

    def cleanup(self):
        """Clean up resources."""
        if self.executor:
            self.executor.shutdown(wait=True)
            self.executor = None
        self.date_cache.clear()

    def get_performance_stats(self) -> dict[str, any]:
        """Get performance statistics."""
        return {
            "cached_dates": len(self.date_cache),
            "border_styles": len(self.border_map),
            "parallel_enabled": self.executor is not None,
        }
