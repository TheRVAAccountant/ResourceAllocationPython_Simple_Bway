"""Excel service for workbook operations using xlwings."""

from __future__ import annotations

from contextlib import suppress
from pathlib import Path
from typing import Any

import pandas as pd
from loguru import logger

try:
    import xlwings as xw

    XLWINGS_AVAILABLE = True
except ImportError:
    XLWINGS_AVAILABLE = False
    xw = None  # Define xw as None when not available
    logger.warning("xlwings not available - Excel integration will be limited")

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from src.core.base_service import BaseService, error_handler, timer
from src.models.allocation import AllocationResult
from src.models.excel import ExcelRange, ExcelStyle


class ExcelService(BaseService):
    """Service for Excel workbook operations.

    Provides functionality for reading, writing, and formatting
    Excel workbooks using both xlwings (for live Excel integration)
    and openpyxl (for file operations).
    """

    def __init__(self, config: dict[str, Any] | None = None):
        """Initialize the Excel service.

        Args:
            config: Configuration dictionary.
        """
        super().__init__(config)
        self.app = None
        self.workbook = None
        self.use_xlwings = self.get_config("use_xlwings", XLWINGS_AVAILABLE)
        self.excel_visible = self.get_config("excel_visible", False)
        self.display_alerts = self.get_config("display_alerts", False)
        self.template_path = self.get_config("template_path", None)

    def initialize(self) -> None:
        """Initialize the Excel service."""
        logger.info("Initializing Excel Service")

        if self.use_xlwings and XLWINGS_AVAILABLE:
            try:
                self.app = xw.App(visible=self.excel_visible, add_book=False)
                self.app.display_alerts = self.display_alerts
                logger.info("xlwings initialized successfully")
            except Exception as e:
                logger.warning(f"Failed to initialize xlwings: {e}")
                self.use_xlwings = False

        self._initialized = True

    def validate(self) -> bool:
        """Validate the service configuration.

        Returns:
            True if configuration is valid.
        """
        if self.template_path:
            template = Path(self.template_path)
            if not template.exists():
                logger.error(f"Template file not found: {self.template_path}")
                return False

        return True

    def cleanup(self) -> None:
        """Clean up Excel resources."""
        if self.workbook and self.use_xlwings:
            with suppress(Exception):
                self.workbook.close()

        if self.app and self.use_xlwings:
            with suppress(Exception):
                self.app.quit()

        super().cleanup()

    @timer
    @error_handler
    def create_workbook(self, template: str | None = None) -> Workbook | Any:
        """Create a new workbook.

        Args:
            template: Path to template file.

        Returns:
            Workbook object (xlwings Book or openpyxl Workbook).
        """
        if self.use_xlwings:
            if template and Path(template).exists():
                self.workbook = self.app.books.open(template)
            else:
                self.workbook = self.app.books.add()
            logger.info("Created xlwings workbook")
        else:
            if template and Path(template).exists():
                self.workbook = load_workbook(template)
            else:
                self.workbook = Workbook()
            logger.info("Created openpyxl workbook")

        return self.workbook

    @timer
    @error_handler
    def open_workbook(self, file_path: str) -> Workbook | Any:
        """Open an existing workbook.

        Args:
            file_path: Path to the Excel file.

        Returns:
            Workbook object.
        """
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        if self.use_xlwings:
            self.workbook = self.app.books.open(str(file_path))
        else:
            self.workbook = load_workbook(str(file_path))

        logger.info(f"Opened workbook: {file_path}")
        return self.workbook

    @timer
    @error_handler
    def save_workbook(self, file_path: str | None = None) -> None:
        """Save the workbook.

        Args:
            file_path: Path to save the file. If None, saves to current location.
        """
        if not self.workbook:
            raise ValueError("No workbook is open")

        if self.use_xlwings:
            if file_path:
                self.workbook.save(file_path)
            else:
                self.workbook.save()
        else:
            if file_path:
                self.workbook.save(file_path)
            else:
                raise ValueError("File path required for openpyxl save")

        logger.info(f"Saved workbook: {file_path or 'current location'}")

    def get_sheet(self, sheet_name: str) -> Any:
        """Get a worksheet by name.

        Args:
            sheet_name: Name of the worksheet.

        Returns:
            Worksheet object.
        """
        if not self.workbook:
            raise ValueError("No workbook is open")

        if self.use_xlwings:
            return self.workbook.sheets[sheet_name]
        else:
            return self.workbook[sheet_name]

    def get_sheet_names(self) -> list[str]:
        """Get list of all sheet names in the workbook.

        Returns:
            List of sheet names.
        """
        if not self.workbook:
            return []

        if self.use_xlwings:
            return [sheet.name for sheet in self.workbook.sheets]
        else:
            return list(self.workbook.sheetnames)

    def create_sheet(self, sheet_name: str, position: int | None = None) -> Any:
        """Create a new worksheet.

        Args:
            sheet_name: Name for the new worksheet.
            position: Position to insert the sheet.

        Returns:
            Created worksheet object.
        """
        if not self.workbook:
            raise ValueError("No workbook is open")

        if self.use_xlwings:
            if position is not None:
                sheet = self.workbook.sheets.add(sheet_name, before=position)
            else:
                sheet = self.workbook.sheets.add(sheet_name)
        else:
            sheet = self.workbook.create_sheet(sheet_name, position)

        logger.info(f"Created sheet: {sheet_name}")
        return sheet

    @timer
    def write_data(
        self,
        sheet_name: str,
        data: pd.DataFrame | list | dict,
        start_row: int = 1,
        start_col: int = 1,
        headers: bool = True,
    ) -> None:
        """Write data to a worksheet.

        Args:
            sheet_name: Name of the worksheet.
            data: Data to write (DataFrame, list, or dict).
            start_row: Starting row (1-based).
            start_col: Starting column (1-based).
            headers: Whether to include headers for DataFrame.
        """
        sheet = self.get_sheet(sheet_name)

        if isinstance(data, pd.DataFrame):
            self._write_dataframe(sheet, data, start_row, start_col, headers)
        elif isinstance(data, list):
            self._write_list(sheet, data, start_row, start_col)
        elif isinstance(data, dict):
            self._write_dict(sheet, data, start_row, start_col)
        else:
            raise ValueError(f"Unsupported data type: {type(data)}")

        logger.info(f"Wrote data to {sheet_name}")

    def _write_dataframe(
        self, sheet: Any, df: pd.DataFrame, start_row: int, start_col: int, headers: bool
    ) -> None:
        """Write DataFrame to worksheet."""
        if self.use_xlwings:
            sheet.range((start_row, start_col)).value = df if headers else df.values
        else:
            # Write headers
            if headers:
                for col_idx, col_name in enumerate(df.columns, start=start_col):
                    sheet.cell(row=start_row, column=col_idx, value=str(col_name))
                start_row += 1

            # Write data
            for row_idx, row_data in enumerate(df.values, start=start_row):
                for col_idx, value in enumerate(row_data, start=start_col):
                    sheet.cell(row=row_idx, column=col_idx, value=value)

    def _write_list(self, sheet: Any, data: list, start_row: int, start_col: int) -> None:
        """Write list to worksheet."""
        if self.use_xlwings:
            sheet.range((start_row, start_col)).value = data
        else:
            for row_idx, row_data in enumerate(data, start=start_row):
                if isinstance(row_data, list | tuple):
                    for col_idx, value in enumerate(row_data, start=start_col):
                        sheet.cell(row=row_idx, column=col_idx, value=value)
                else:
                    sheet.cell(row=row_idx, column=start_col, value=row_data)

    def _write_dict(self, sheet: Any, data: dict, start_row: int, start_col: int) -> None:
        """Write dictionary to worksheet."""
        for row_idx, (key, value) in enumerate(data.items(), start=start_row):
            if self.use_xlwings:
                sheet.range((row_idx, start_col)).value = key
                sheet.range((row_idx, start_col + 1)).value = value
            else:
                sheet.cell(row=row_idx, column=start_col, value=key)
                sheet.cell(row=row_idx, column=start_col + 1, value=value)

    def read_data(
        self, sheet_name: str, range_str: str | None = None, as_dataframe: bool = True
    ) -> pd.DataFrame | list:
        """Read data from a worksheet.

        Args:
            sheet_name: Name of the worksheet.
            range_str: Range to read (e.g., 'A1:C10'). If None, reads all data.
            as_dataframe: Whether to return as DataFrame.

        Returns:
            Data as DataFrame or list.
        """
        sheet = self.get_sheet(sheet_name)

        if self.use_xlwings:
            data = sheet.range(range_str).value if range_str else sheet.used_range.value
        else:
            data = []
            if range_str:
                excel_range = ExcelRange.from_excel_range(sheet_name, range_str)
                for row in sheet.iter_rows(
                    min_row=excel_range.start_row,
                    max_row=excel_range.end_row,
                    min_col=excel_range.start_col,
                    max_col=excel_range.end_col,
                    values_only=True,
                ):
                    data.append(list(row))
            else:
                for row in sheet.values:
                    if any(row):  # Skip empty rows
                        data.append(list(row))

        if as_dataframe and data:
            df = pd.DataFrame(data[1:], columns=data[0])
            return df

        return data

    def apply_style(self, excel_range: ExcelRange, style: ExcelStyle) -> None:
        """Apply style to a range of cells.

        Args:
            excel_range: Range to apply style to.
            style: Style to apply.
        """
        sheet = self.get_sheet(excel_range.sheet_name)

        if self.use_xlwings:
            self._apply_style_xlwings(sheet, excel_range, style)
        else:
            self._apply_style_openpyxl(sheet, excel_range, style)

        logger.debug(f"Applied style to {excel_range.to_excel_range()}")

    def _apply_style_xlwings(self, sheet: Any, excel_range: ExcelRange, style: ExcelStyle) -> None:
        """Apply style using xlwings."""
        range_obj = sheet.range(excel_range.to_excel_range())

        if style.font:
            range_obj.font.name = style.font.name
            range_obj.font.size = style.font.size
            range_obj.font.bold = style.font.bold
            range_obj.font.italic = style.font.italic

            if style.font.color and style.font.color.rgb:
                range_obj.font.color = tuple(
                    int(style.font.color.rgb[i : i + 2], 16) for i in (0, 2, 4)
                )

        if style.alignment:
            range_obj.api.HorizontalAlignment = self._get_xlwings_alignment(
                style.alignment.horizontal
            )
            range_obj.api.VerticalAlignment = self._get_xlwings_alignment(
                style.alignment.vertical, vertical=True
            )
            range_obj.api.WrapText = style.alignment.wrap_text

        if style.number_format:
            range_obj.number_format = style.number_format

    def _apply_style_openpyxl(self, sheet: Any, excel_range: ExcelRange, style: ExcelStyle) -> None:
        """Apply style using openpyxl."""
        for row in range(excel_range.start_row, excel_range.end_row + 1):
            for col in range(excel_range.start_col, excel_range.end_col + 1):
                cell = sheet.cell(row=row, column=col)

                if style.font:
                    cell.font = Font(
                        name=style.font.name,
                        size=style.font.size,
                        bold=style.font.bold,
                        italic=style.font.italic,
                        underline="single" if style.font.underline else None,
                        strike=style.font.strike,
                        color=style.font.color.rgb if style.font.color else None,
                    )

                if style.alignment:
                    cell.alignment = Alignment(
                        horizontal=style.alignment.horizontal,
                        vertical=style.alignment.vertical,
                        text_rotation=style.alignment.text_rotation,
                        wrap_text=style.alignment.wrap_text,
                        shrink_to_fit=style.alignment.shrink_to_fit,
                        indent=style.alignment.indent,
                    )

                if style.fill and style.fill.fg_color:
                    cell.fill = PatternFill(
                        start_color=style.fill.fg_color.rgb,
                        end_color=style.fill.bg_color.rgb if style.fill.bg_color else None,
                        fill_type=style.fill.pattern_type,
                    )

                if style.number_format:
                    cell.number_format = style.number_format

    def _get_xlwings_alignment(self, alignment: str, vertical: bool = False) -> int:
        """Get xlwings alignment constant."""
        if vertical:
            mapping = {
                "top": -4160,
                "center": -4108,
                "bottom": -4107,
                "justify": -4130,
                "distributed": -4117,
            }
        else:
            mapping = {
                "left": -4131,
                "center": -4108,
                "right": -4152,
                "justify": -4130,
                "fill": 5,
                "distributed": -4117,
            }

        return mapping.get(alignment, -4131 if not vertical else -4107)

    def write_allocation_result(
        self, sheet_name: str, result: AllocationResult, start_row: int = 1
    ) -> None:
        """Write allocation result to worksheet.

        Args:
            sheet_name: Name of the worksheet.
            result: Allocation result to write.
            start_row: Starting row for the data.
        """
        # Prepare allocation data
        allocation_data = []
        for driver_id, vehicles in result.allocations.items():
            for vehicle_id in vehicles:
                allocation_data.append(
                    {
                        "Driver ID": driver_id,
                        "Vehicle ID": vehicle_id,
                        "Status": "Allocated",
                        "Timestamp": result.timestamp,
                    }
                )

        # Add unallocated vehicles
        for vehicle_id in result.unallocated_vehicles:
            allocation_data.append(
                {
                    "Driver ID": "Unallocated",
                    "Vehicle ID": vehicle_id,
                    "Status": "Unallocated",
                    "Timestamp": result.timestamp,
                }
            )

        # Convert to DataFrame and write
        df = pd.DataFrame(allocation_data)
        self.write_data(sheet_name, df, start_row=start_row)

        logger.info(f"Wrote allocation result to {sheet_name}")
