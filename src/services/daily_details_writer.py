"""Daily Details writer service for GAS-compatible output."""

from datetime import datetime, date
from pathlib import Path
from typing import Optional, Any, Union, List, Dict
from loguru import logger
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from src.core.base_service import BaseService
from src.models.allocation import AllocationResult
from src.services.daily_details_thick_borders import DailyDetailsThickBorderService


class DailyDetailsWriter(BaseService):
    """Handles writing allocation results to Daily Details sheet format.
    
    This service ensures compatibility with Google Apps Script by:
    1. Creating/updating Daily Details sheet with exact 24-column structure
    2. Appending to existing files without overwriting data
    3. Generating unique identifiers to prevent duplicates
    4. Creating dated Results and Unassigned sheets
    """
    
    # Define the exact 24 columns for Daily Details sheet
    DAILY_DETAILS_COLUMNS = [
        "Date",
        "Route #",
        "Name",
        "Asset ID",
        "Van ID",
        "VIN",
        "GeoTab Code",
        "Type",
        "Vehicle Type",
        "Route Type",
        "Rescue",
        "Delivery Pace 1:40pm",
        "Delivery Pace 3:40pm",
        "Delivery Pace 5:40pm",
        "Delivery Pace 7:40pm",
        "Delivery Pace 9:40pm",
        "RTS TIME",
        "Pkg. Delivered",
        "Pkg. Returned",
        "Route Notes",
        "Week Number",
        "Unique Identifier",
        "Vehicle Inspection",
        "Route Completion"
    ]
    
    # Results sheet columns (11 columns matching GAS)
    RESULTS_COLUMNS = [
        "Route Code",
        "Service Type",
        "DSP",
        "Wave",
        "Staging Location",
        "Van ID",
        "Device Name",
        "Van Type",
        "Operational",
        "Associate Name",
        "Unique Identifier"
    ]
    
    def __init__(self, excel_service=None):
        """Initialize the Daily Details writer.
        
        Args:
            excel_service: Reference to Excel service for file operations.
        """
        super().__init__()
        self.excel_service = excel_service
        self.header_fill = PatternFill(start_color="46BDC6", end_color="46BDC6", fill_type="solid")
        self.header_font = Font(bold=True, color="000000")
        self.header_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        self.header_alignment = Alignment(horizontal="center", vertical="center")

        # Brand priority color fills for Daily Details Van ID column
        self.branded_fill = PatternFill(start_color="D4F5DC", end_color="D4F5DC", fill_type="solid")
        self.rental_fill = PatternFill(start_color="F3F3F3", end_color="F3F3F3", fill_type="solid")
        
        # Initialize thick border service
        self.thick_border_service = DailyDetailsThickBorderService()
    
    def initialize(self) -> None:
        """Initialize the Daily Details writer service.
        
        Sets up the service and validates dependencies.
        """
        if self._initialized:
            return
            
        logger.info("Initializing DailyDetailsWriter")
        
        # Validate dependencies
        try:
            import openpyxl
            import pandas
        except ImportError as e:
            logger.error(f"Missing required dependency: {e}")
            raise
        
        self._initialized = True
        logger.info("DailyDetailsWriter initialized successfully")
    
    def validate(self) -> bool:
        """Validate the writer configuration and state.
        
        Returns:
            True if the service is valid and ready to use.
        """
        if not self._initialized:
            logger.warning("DailyDetailsWriter not initialized")
            return False
        
        # Validate column configurations
        if len(self.DAILY_DETAILS_COLUMNS) != 24:
            logger.error(f"Invalid DAILY_DETAILS_COLUMNS: expected 24, got {len(self.DAILY_DETAILS_COLUMNS)}")
            return False
        
        if len(self.RESULTS_COLUMNS) != 11:
            logger.error(f"Invalid RESULTS_COLUMNS: expected 11, got {len(self.RESULTS_COLUMNS)}")
            return False
        
        return True
    
    def append_to_existing_file(
        self,
        file_path: str,
        allocation_result: AllocationResult,
        allocation_date: date,
        vehicle_log_dict: Optional[Dict] = None,
        skip_results_sheet: bool = True
    ) -> bool:
        """Append allocation results to an existing Daily Summary Log file.
        
        This method:
        1. Opens the existing file
        2. Finds or creates the Daily Details sheet
        3. Checks for duplicates using unique identifiers
        4. Appends new allocation rows
        5. Creates dated Results and Unassigned sheets
        
        Args:
            file_path: Path to the existing Daily Summary Log file.
            allocation_result: The allocation results to append.
            allocation_date: Date of the allocation.
            vehicle_log_dict: Optional dictionary of vehicle details (VIN, GeoTab, etc.).
            
        Returns:
            True if successfully appended, False otherwise.
        """
        try:
            # Load the existing workbook
            logger.info(f"Opening existing file: {file_path}")
            workbook = load_workbook(file_path)
            
            # Find or create Daily Details sheet
            if "Daily Details" in workbook.sheetnames:
                daily_details = workbook["Daily Details"]
                logger.info("Found existing Daily Details sheet")
            else:
                daily_details = workbook.create_sheet("Daily Details", 0)
                logger.info("Created new Daily Details sheet")
                self._setup_daily_details_headers(daily_details)
            
            # Verify headers match expected structure
            if not self._verify_headers(daily_details):
                self._update_headers(daily_details)
            
            # Get existing unique identifiers to check for duplicates
            existing_ids = self._get_existing_unique_ids(daily_details)
            
            # Prepare new rows to append
            new_rows = self._prepare_allocation_rows(
                allocation_result,
                allocation_date,
                vehicle_log_dict,
                existing_ids
            )
            
            if new_rows:
                # Find the last populated row
                last_row = self._get_last_populated_row(daily_details)
                
                # Append new rows
                start_row = last_row + 1
                for row_idx, row_data in enumerate(new_rows, start=start_row):
                    for col_idx, value in enumerate(row_data, start=1):
                        daily_details.cell(row=row_idx, column=col_idx, value=value)
                
                logger.info(f"Appended {len(new_rows)} rows to Daily Details")
                
                # Apply formatting to new rows
                self._apply_row_formatting(daily_details, start_row, len(new_rows))
                
                # Apply thick borders to the new date sections
                self.thick_border_service.apply_thick_borders_after_append(
                    daily_details,
                    start_row,
                    len(new_rows)
                )
            else:
                logger.warning("No new rows to append (all duplicates)")
            
            # Create dated Results sheet (unless skipped for separate file)
            if not skip_results_sheet:
                results_sheet_name = self._create_dated_sheet_name(allocation_date, "Results")
                if results_sheet_name not in workbook.sheetnames:
                    self._create_results_sheet(workbook, results_sheet_name, allocation_result, allocation_date)
            
            # Create Unassigned sheet if needed
            if allocation_result.unallocated_vehicles:
                unassigned_sheet_name = self._create_dated_sheet_name(allocation_date, "Available & Unassigned")
                if unassigned_sheet_name not in workbook.sheetnames:
                    self._create_unassigned_sheet(workbook, unassigned_sheet_name, allocation_result, allocation_date)
            
            # Save the workbook
            workbook.save(file_path)
            logger.info(f"Successfully saved updates to {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to append to existing file: {e}")
            return False
    
    def _verify_headers(self, sheet) -> bool:
        """Verify that the sheet has the correct headers.
        
        Args:
            sheet: The worksheet to verify.
            
        Returns:
            True if headers match, False otherwise.
        """
        current_headers = []
        for col in range(1, 25):  # Check first 24 columns
            cell = sheet.cell(row=1, column=col)
            if cell.value:
                current_headers.append(str(cell.value).strip())
        
        return current_headers == self.DAILY_DETAILS_COLUMNS
    
    def _update_headers(self, sheet):
        """Update sheet headers to match the expected structure.
        
        Args:
            sheet: The worksheet to update.
        """
        logger.info("Updating Daily Details headers to match GAS format")
        
        for col_idx, header in enumerate(self.DAILY_DETAILS_COLUMNS, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.header_border
            cell.alignment = self.header_alignment
    
    def _setup_daily_details_headers(self, sheet):
        """Set up the Daily Details sheet with proper headers and formatting.
        
        Args:
            sheet: The worksheet to set up.
        """
        # Add all 24 column headers
        for col_idx, header in enumerate(self.DAILY_DETAILS_COLUMNS, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.header_border
            cell.alignment = self.header_alignment
        
        # Set column widths
        column_widths = {
            "Date": 12,
            "Route #": 10,
            "Name": 20,
            "Asset ID": 10,
            "Van ID": 10,
            "VIN": 17,
            "GeoTab Code": 12,
            "Type": 10,
            "Vehicle Type": 12,
            "Route Type": 12,
            "Rescue": 8,
            "RTS TIME": 10,
            "Pkg. Delivered": 12,
            "Pkg. Returned": 12,
            "Route Notes": 20,
            "Week Number": 12,
            "Unique Identifier": 25,
            "Vehicle Inspection": 15,
            "Route Completion": 15
        }
        
        for col_idx, header in enumerate(self.DAILY_DETAILS_COLUMNS, start=1):
            width = column_widths.get(header, 15)  # Default width of 15
            sheet.column_dimensions[get_column_letter(col_idx)].width = width
    
    def apply_thick_borders_to_entire_sheet(self, worksheet):
        """Apply thick borders to the entire Daily Details sheet.
        
        This method should be called when you want to apply thick borders
        to all existing data in the sheet, not just newly added rows.
        
        Args:
            worksheet: The Daily Details worksheet.
        """
        self.thick_border_service.apply_thick_borders_to_daily_details(worksheet)
    
    def _get_existing_unique_ids(self, sheet) -> set:
        """Get all existing unique identifiers from the sheet.
        
        Args:
            sheet: The worksheet to check.
            
        Returns:
            Set of existing unique identifiers.
        """
        unique_ids = set()
        unique_id_col = self.DAILY_DETAILS_COLUMNS.index("Unique Identifier") + 1
        
        for row in range(2, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=unique_id_col).value
            if cell_value:
                unique_ids.add(str(cell_value))
        
        logger.info(f"Found {len(unique_ids)} existing unique identifiers")
        return unique_ids
    
    def _get_last_populated_row(self, sheet) -> int:
        """Find the last row with data in the sheet.
        
        Args:
            sheet: The worksheet to check.
            
        Returns:
            The last populated row number.
        """
        last_row = 1  # Start after header
        
        # Check the Date column (column 1) for the last populated row
        for row in range(sheet.max_row, 1, -1):
            if sheet.cell(row=row, column=1).value:
                last_row = row
                break
        
        logger.info(f"Last populated row: {last_row}")
        return last_row
    
    def generate_unique_identifier(
        self,
        allocation_date: date,
        route: str,
        device: str,
        van_id: str
    ) -> str:
        """Generate a unique identifier for an allocation record.
        
        Format: Date|Route|Device|VanID
        
        Args:
            allocation_date: The allocation date.
            route: Route code.
            device: Device/Asset ID.
            van_id: Van ID.
            
        Returns:
            Unique identifier string.
        """
        date_str = allocation_date.strftime("%m/%d/%Y")
        return f"{date_str}|{route}|{device}|{van_id}"

    def _normalize_brand_label(self, value: Any) -> Optional[str]:
        """Normalize raw brand/rental values to simplified labels."""
        if value is None:
            return None
        try:
            if pd.isna(value):  # type: ignore[arg-type]
                return None
        except Exception:
            pass
        text = str(value).strip().lower()
        if not text:
            return None
        if "brand" in text:
            return "branded"
        if "rent" in text:
            return "rental"
        return None

    def _normalize_brand_display(self, value: Any) -> str:
        """Return a human-readable brand label when possible."""
        label = self._normalize_brand_label(value)
        if label == "branded":
            return "Branded"
        if label == "rental":
            return "Rental"
        if value is None:
            return ""
        try:
            if pd.isna(value):  # type: ignore[arg-type]
                return ""
        except Exception:
            pass
        return str(value).strip()
    
    def _prepare_allocation_rows(
        self,
        allocation_result: AllocationResult,
        allocation_date: date,
        vehicle_log_dict: Optional[Dict] = None,
        existing_ids: Optional[set] = None
    ) -> List[List[Any]]:
        """Prepare allocation data rows for Daily Details sheet.
        
        Args:
            allocation_result: The allocation results.
            allocation_date: Date of allocation.
            vehicle_log_dict: Optional vehicle details dictionary.
            existing_ids: Set of existing unique IDs to check for duplicates.
            
        Returns:
            List of row data to append.
        """
        if existing_ids is None:
            existing_ids = set()
        
        rows = []
        week_number = allocation_date.isocalendar()[1]
        
        # Check if we have detailed results in metadata (from GAS allocator)
        detailed_results = allocation_result.metadata.get("detailed_results", [])
        
        if detailed_results:
            # Use detailed results from GAS allocator which includes route codes
            for result in detailed_results:
                vehicle_id = result.get("Van ID", "")
                route_code = result.get("Route Code", "")
                driver_name = result.get("Associate Name", "")
                service_type = result.get("Service Type", "")
                
                # Generate unique identifier using actual route code
                unique_id = self.generate_unique_identifier(
                    allocation_date,
                    route_code,
                    vehicle_id,  # Device ID
                    vehicle_id   # Van ID
                )
                
                # Skip if duplicate
                if unique_id in existing_ids:
                    logger.debug(f"Skipping duplicate: {unique_id}")
                    continue
                
                # Create row matching the 24-column structure
                row = [None] * 24
                
                # Populate known fields (indices are 0-based)
                row[0] = allocation_date  # Date
                row[1] = route_code  # Route # - Now using actual route code
                row[2] = driver_name  # Name - Now using actual driver name
                row[3] = vehicle_id  # Asset ID
                row[4] = vehicle_id  # Van ID
                
                # Look up vehicle details if available
                if vehicle_log_dict and vehicle_id in vehicle_log_dict:
                    vehicle_info = vehicle_log_dict[vehicle_id]
                    vin_value = vehicle_info.get("vin", "")
                    geotab_value = vehicle_info.get("geotab", "")
                    brand_rental_value = vehicle_info.get("brand_or_rental", "")
                    
                    row[5] = vin_value  # VIN
                    row[6] = geotab_value  # GeoTab Code
                    row[7] = self._normalize_brand_display(brand_rental_value)  # Type (Branded or Rental)
                    # Vehicle Type comes from allocation result
                    row[8] = result.get("Van Type", "")  # Vehicle Type
                    
                    # Log if GeoTab or Type are empty
                    if not geotab_value or not brand_rental_value:
                        logger.debug(f"Vehicle {vehicle_id}: VIN={vin_value}, GeoTab={geotab_value}, Type={brand_rental_value}")
                else:
                    # Use data from result if available
                    logger.warning(f"Vehicle {vehicle_id} not found in vehicle_log_dict")
                    row[7] = ""
                    row[8] = result.get("Van Type", "")  # Vehicle Type
                
                # Route Type is the Service Type itself (no transformation)
                row[9] = service_type  # Route Type - direct pass-through from Day of Ops
                
                row[10] = ""  # Rescue
                # Delivery Pace columns (11-15) left empty for form updates
                row[11] = ""  # Delivery Pace 1:40pm
                row[12] = ""  # Delivery Pace 3:40pm
                row[13] = ""  # Delivery Pace 5:40pm
                row[14] = ""  # Delivery Pace 7:40pm
                row[15] = ""  # Delivery Pace 9:40pm
                row[16] = ""  # RTS TIME
                row[17] = ""  # Pkg. Delivered
                row[18] = ""  # Pkg. Returned
                row[19] = ""  # Route Notes
                row[20] = week_number  # Week Number
                row[21] = unique_id  # Unique Identifier
                row[22] = ""  # Vehicle Inspection
                row[23] = ""  # Route Completion
                
                rows.append(row)
        else:
            # Fallback to basic allocation data (for non-GAS allocators)
            for driver_id, vehicles in allocation_result.allocations.items():
                for idx, vehicle_id in enumerate(vehicles):
                    # Generate synthetic route code as fallback
                    route_code = f"R{len(rows)+1:03d}"
                    
                    # Generate unique identifier
                    unique_id = self.generate_unique_identifier(
                        allocation_date,
                        route_code,
                        vehicle_id,
                        vehicle_id
                    )
                    
                    # Skip if duplicate
                    if unique_id in existing_ids:
                        logger.debug(f"Skipping duplicate: {unique_id}")
                        continue
                    
                    # Create row matching the 24-column structure
                    row = [None] * 24
                    
                    # Populate known fields (indices are 0-based)
                    row[0] = allocation_date  # Date
                    row[1] = route_code  # Route #
                    row[2] = driver_id  # Name
                    row[3] = vehicle_id  # Asset ID
                    row[4] = vehicle_id  # Van ID
                    
                    # Look up vehicle details if available
                    if vehicle_log_dict and vehicle_id in vehicle_log_dict:
                        vehicle_info = vehicle_log_dict[vehicle_id]
                        row[5] = vehicle_info.get("vin", "")  # VIN
                        row[6] = vehicle_info.get("geotab", "")  # GeoTab Code
                        row[7] = self._normalize_brand_display(vehicle_info.get("brand_or_rental", ""))  # Type
                        row[8] = vehicle_info.get("vehicle_type", "")  # Vehicle Type
                    else:
                        row[7] = ""
                        row[8] = ""
                    
                    row[9] = ""  # Route Type (needs service type from allocation)
                    row[10] = ""  # Rescue
                    # Delivery Pace columns (11-15) left empty for form updates
                    row[11] = ""  # Delivery Pace 1:40pm
                    row[12] = ""  # Delivery Pace 3:40pm
                    row[13] = ""  # Delivery Pace 5:40pm
                    row[14] = ""  # Delivery Pace 7:40pm
                    row[15] = ""  # Delivery Pace 9:40pm
                    row[16] = ""  # RTS TIME
                    row[17] = ""  # Pkg. Delivered
                    row[18] = ""  # Pkg. Returned
                    row[19] = ""  # Route Notes
                    row[20] = week_number  # Week Number
                    row[21] = unique_id  # Unique Identifier
                    row[22] = ""  # Vehicle Inspection
                    row[23] = ""  # Route Completion
                    
                    rows.append(row)
        
        logger.info(f"Prepared {len(rows)} new rows for Daily Details")
        return rows
    
    def _apply_row_formatting(self, sheet, start_row: int, num_rows: int):
        """Apply formatting to newly added rows.
        
        Args:
            sheet: The worksheet.
            start_row: First row to format.
            num_rows: Number of rows to format.
        """
        # Apply thin borders to all data cells
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        for row in range(start_row, start_row + num_rows):
            for col in range(1, 25):  # 24 columns
                cell = sheet.cell(row=row, column=col)
                cell.border = thin_border
                
                # Center align certain columns
                if col in [1, 2, 4, 5, 11, 12, 13, 14, 15, 16, 21]:  # Date, Route #, IDs, Times
                    cell.alignment = Alignment(horizontal="center")
                
                # Format date column
                if col == 1 and cell.value:
                    cell.number_format = "mm/dd/yyyy"
            brand_label = self._normalize_brand_label(sheet.cell(row=row, column=8).value)
            if brand_label == "branded":
                sheet.cell(row=row, column=5).fill = self.branded_fill
            elif brand_label == "rental":
                sheet.cell(row=row, column=5).fill = self.rental_fill
    
    def _create_dated_sheet_name(self, allocation_date: date, suffix: str) -> str:
        """Create a sheet name with MM-DD-YY format.
        
        Args:
            allocation_date: The date for the sheet.
            suffix: Sheet type suffix (e.g., "Results", "Available & Unassigned").
            
        Returns:
            Formatted sheet name.
        """
        date_str = allocation_date.strftime("%m-%d-%y")
        return f"{date_str} {suffix}"
    
    def _create_results_sheet(
        self,
        workbook,
        sheet_name: str,
        allocation_result: AllocationResult,
        allocation_date: date
    ):
        """Create a Results sheet with allocation results.
        
        Args:
            workbook: The workbook to add the sheet to.
            sheet_name: Name for the new sheet.
            allocation_result: The allocation results.
            allocation_date: Date of allocation.
        """
        sheet = workbook.create_sheet(sheet_name)
        
        # Add headers
        for col_idx, header in enumerate(self.RESULTS_COLUMNS, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.header_border
            cell.alignment = self.header_alignment
        
        # Add data rows
        row_num = 2
        
        # Check if we have detailed results in metadata (from GAS allocator)
        detailed_results = allocation_result.metadata.get("detailed_results", [])
        
        if detailed_results:
            # Use detailed results from GAS allocator
            for result in detailed_results:
                unique_id = result.get("Unique Identifier", "")
                if not unique_id:
                    # Generate if not present
                    unique_id = self.generate_unique_identifier(
                        allocation_date,
                        result.get("Route Code", ""),
                        result.get("Van ID", ""),
                        result.get("Van ID", "")
                    )
                
                sheet.cell(row=row_num, column=1, value=result.get("Route Code", ""))
                sheet.cell(row=row_num, column=2, value=result.get("Service Type", ""))
                sheet.cell(row=row_num, column=3, value=result.get("DSP", ""))
                sheet.cell(row=row_num, column=4, value=result.get("Wave", ""))
                sheet.cell(row=row_num, column=5, value=result.get("Staging Location", ""))
                sheet.cell(row=row_num, column=6, value=result.get("Van ID", ""))
                sheet.cell(row=row_num, column=7, value=result.get("Device Name", ""))
                sheet.cell(row=row_num, column=8, value=result.get("Van Type", ""))
                sheet.cell(row=row_num, column=9, value=result.get("Operational", ""))
                sheet.cell(row=row_num, column=10, value=result.get("Associate Name", ""))
                sheet.cell(row=row_num, column=11, value=unique_id)
                
                row_num += 1
        else:
            # Fallback for basic allocation results
            for driver_id, vehicles in allocation_result.allocations.items():
                for vehicle_id in vehicles:
                    unique_id = self.generate_unique_identifier(
                        allocation_date,
                        f"R{row_num-1:03d}",
                        vehicle_id,
                        vehicle_id
                    )
                    
                    sheet.cell(row=row_num, column=1, value=f"R{row_num-1:03d}")  # Route Code
                    sheet.cell(row=row_num, column=2, value="Standard")  # Service Type
                    sheet.cell(row=row_num, column=3, value="BWAY")  # DSP
                    sheet.cell(row=row_num, column=4, value="8:00 AM")  # Wave
                    sheet.cell(row=row_num, column=5, value="STG.G.1")  # Staging Location
                    sheet.cell(row=row_num, column=6, value=vehicle_id)  # Van ID
                    sheet.cell(row=row_num, column=7, value=vehicle_id)  # Device Name
                    sheet.cell(row=row_num, column=8, value="Standard")  # Van Type
                    sheet.cell(row=row_num, column=9, value="Y")  # Operational
                    sheet.cell(row=row_num, column=10, value=driver_id)  # Associate Name
                    sheet.cell(row=row_num, column=11, value=unique_id)  # Unique Identifier
                    
                    row_num += 1
        
        logger.info(f"Created Results sheet '{sheet_name}' with {row_num-2} rows")
    
    def _create_unassigned_sheet(
        self,
        workbook,
        sheet_name: str,
        allocation_result: AllocationResult,
        allocation_date: date
    ):
        """Create an Unassigned Vans sheet.
        
        Args:
            workbook: The workbook to add the sheet to.
            sheet_name: Name for the new sheet.
            allocation_result: The allocation results.
            allocation_date: Date of allocation.
        """
        sheet = workbook.create_sheet(sheet_name)
        
        # Headers for unassigned sheet (15 columns as per GAS)
        headers = [
            "Van ID", "Year", "Make", "Model", "Style",
            "Type", "License Tag Number", "License Tag State",
            "Ownership", "VIN", "Van Type", "Issue",
            "Date GROUNDED", "Date UNGROUNDED", "Opnal? Y/N"
        ]
        
        # Add headers
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.header_border
            cell.alignment = self.header_alignment
        
        # Add unassigned vehicles
        for row_idx, vehicle_id in enumerate(allocation_result.unallocated_vehicles, start=2):
            sheet.cell(row=row_idx, column=1, value=vehicle_id)  # Van ID
            sheet.cell(row=row_idx, column=15, value="Y")  # Operational
        
        logger.info(f"Created Unassigned sheet '{sheet_name}' with {len(allocation_result.unallocated_vehicles)} vehicles")