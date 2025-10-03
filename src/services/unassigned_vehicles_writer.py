"""Service for creating and managing unassigned vehicles Excel sheets."""

from datetime import date, datetime
from pathlib import Path
from typing import Any

import pandas as pd
from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.core.base_service import BaseService


class UnassignedVehiclesWriter(BaseService):
    """Manages creation and update of unassigned vehicles sheet."""

    # Column definitions for unassigned vehicles sheet
    COLUMNS = [
        ("Van ID", 12),
        ("Vehicle Type", 15),
        ("Operational Status", 12),
        ("Last Known Location", 18),
        ("Days Since Last Assignment", 20),
        ("VIN", 20),
        ("GeoTab Code", 15),
        ("Branded or Rental", 15),
        ("Notes", 25),
        ("Unassigned Date", 12),
        ("Unassigned Time", 12),
    ]

    # Header formatting
    HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Border styles
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def __init__(self, config: dict | None = None):
        """Initialize the unassigned vehicles writer.

        Args:
            config: Optional configuration dictionary.
        """
        super().__init__(config)
        self.historical_data_path = self.get_config("historical_data_path", None)
        self.enable_alternating_rows = self.get_config("enable_alternating_rows", True)

    def initialize(self) -> None:
        """Initialize the unassigned vehicles writer service.

        This method is called to set up the writer for use.
        """
        logger.info("Initializing UnassignedVehiclesWriter")

        # Verify historical data path if configured
        if self.historical_data_path:
            path = Path(self.historical_data_path)
            if not path.exists():
                logger.warning(f"Historical data path does not exist: {self.historical_data_path}")

        # Mark as initialized
        self._initialized = True
        logger.info(
            f"UnassignedVehiclesWriter initialized with "
            f"enable_alternating_rows={self.enable_alternating_rows}"
        )

    def validate(self) -> bool:
        """Validate the service configuration and state.

        Returns:
            True if the service is valid and ready to use.
        """
        try:
            # Check if we have valid configuration
            if not isinstance(self.enable_alternating_rows, bool):
                logger.error(
                    f"Invalid enable_alternating_rows value: {self.enable_alternating_rows}"
                )
                return False

            # If historical data path is set, verify it's a valid path string
            if self.historical_data_path is not None and not isinstance(
                self.historical_data_path, str
            ):
                logger.error(
                    f"Invalid historical_data_path type: {type(self.historical_data_path)}"
                )
                return False

            # Service is valid
            logger.debug("UnassignedVehiclesWriter validation successful")
            return True

        except Exception as e:
            logger.error(f"Error during validation: {e}")
            return False

    def create_unassigned_sheet(
        self,
        workbook: Workbook,
        unassigned_vehicles: pd.DataFrame,
        vehicle_log_dict: dict[str, dict],
        allocation_date: date,
        historical_assignments: pd.DataFrame | None = None,
    ) -> Worksheet:
        """
        Create or update unassigned vehicles sheet.

        Args:
            workbook: The workbook to add sheet to.
            unassigned_vehicles: DataFrame of unassigned vehicles.
            vehicle_log_dict: Dictionary mapping van ID to vehicle details.
            allocation_date: Date of the allocation.
            historical_assignments: Optional historical assignment data.

        Returns:
            The created worksheet.
        """
        # Generate sheet name
        sheet_name = allocation_date.strftime("%m-%d-%y") + " Available & Unassigned"

        # Remove existing sheet if present
        if sheet_name in workbook.sheetnames:
            logger.info(f"Removing existing sheet: {sheet_name}")
            del workbook[sheet_name]

        # Create new sheet
        logger.info(f"Creating unassigned vehicles sheet: {sheet_name}")
        worksheet = workbook.create_sheet(sheet_name)

        # Set up headers
        self._setup_headers(worksheet)

        # Write data
        self._write_unassigned_data(
            worksheet,
            unassigned_vehicles,
            vehicle_log_dict,
            allocation_date,
            historical_assignments,
        )

        # Apply formatting
        self.format_unassigned_sheet(worksheet)

        return worksheet

    def _setup_headers(self, worksheet: Worksheet) -> None:
        """Set up column headers."""
        for col_idx, (header, width) in enumerate(self.COLUMNS, start=1):
            cell = worksheet.cell(row=1, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER

            # Set column width
            worksheet.column_dimensions[get_column_letter(col_idx)].width = width

    def _write_unassigned_data(
        self,
        worksheet: Worksheet,
        unassigned_vehicles: pd.DataFrame,
        vehicle_log_dict: dict[str, dict],
        allocation_date: date,
        historical_assignments: pd.DataFrame | None = None,
    ) -> None:
        """Write unassigned vehicle data to worksheet."""
        current_row = 2
        timestamp = datetime.now()

        for _, vehicle in unassigned_vehicles.iterrows():
            van_id = vehicle.get("Van ID", "")
            if not van_id:
                continue

            # Get vehicle details from log
            vehicle_details = vehicle_log_dict.get(van_id, {})

            # Calculate days since last assignment
            days_since_assignment = (
                self.calculate_days_since_assignment(van_id, historical_assignments)
                if historical_assignments is not None
                else 0
            )

            # Write row data
            row_data = [
                van_id,  # Van ID
                vehicle.get("Type", "Unknown"),  # Vehicle Type
                vehicle.get("Opnal? Y/N", "N"),  # Operational Status
                vehicle.get("Location", ""),  # Last Known Location
                days_since_assignment,  # Days Since Last Assignment
                vehicle_details.get("vin", ""),  # VIN
                vehicle_details.get("geotab", ""),  # GeoTab Code
                vehicle_details.get("brand_or_rental", ""),  # Branded or Rental
                "",  # Notes (empty for now)
                allocation_date.strftime("%m/%d/%Y"),  # Unassigned Date
                timestamp.strftime("%H:%M:%S"),  # Unassigned Time
            ]

            for col_idx, value in enumerate(row_data, start=1):
                cell = worksheet.cell(row=current_row, column=col_idx, value=value)
                cell.border = self.THIN_BORDER

                # Apply alternating row colors
                if self.enable_alternating_rows and current_row % 2 == 0:
                    cell.fill = PatternFill(
                        start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
                    )

                # Format specific columns
                if col_idx == 5 or col_idx in [10, 11]:  # Days Since Last Assignment
                    cell.alignment = Alignment(horizontal="center")

            current_row += 1

        logger.info(f"Wrote {current_row - 2} unassigned vehicles to sheet")

    def format_unassigned_sheet(self, worksheet: Worksheet) -> None:
        """Apply formatting to unassigned vehicles sheet."""
        # Get data range
        max_row = worksheet.max_row
        max_col = len(self.COLUMNS)

        if max_row < 2:
            return  # No data to format

        # Apply AutoFilter
        data_range = f"A1:{get_column_letter(max_col)}{max_row}"
        worksheet.auto_filter.ref = data_range
        logger.debug(f"Applied AutoFilter to range: {data_range}")

        # Freeze top row
        worksheet.freeze_panes = "A2"

        # Set print settings
        worksheet.page_setup.orientation = "landscape"
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = False

        # Set margins
        worksheet.page_margins.left = 0.5
        worksheet.page_margins.right = 0.5
        worksheet.page_margins.top = 0.75
        worksheet.page_margins.bottom = 0.75
        worksheet.page_margins.header = 0.3
        worksheet.page_margins.footer = 0.3

    def calculate_days_since_assignment(
        self, vehicle_id: str, historical_data: pd.DataFrame
    ) -> int:
        """
        Calculate days since vehicle was last assigned.

        Args:
            vehicle_id: The vehicle ID to check.
            historical_data: DataFrame with historical assignment data.

        Returns:
            Number of days since last assignment, or 0 if never assigned.
        """
        if historical_data is None or historical_data.empty:
            return 0

        try:
            # Filter for this vehicle's assignments
            vehicle_assignments = historical_data[historical_data["Van ID"] == vehicle_id]

            if vehicle_assignments.empty:
                return 0

            # Get most recent assignment date
            # Assume there's a date column in historical data
            if "Date" in vehicle_assignments.columns:
                last_date = pd.to_datetime(vehicle_assignments["Date"]).max()

                # Calculate days difference
                days_diff = (datetime.now().date() - last_date.date()).days
                return max(0, days_diff)

            return 0

        except Exception as e:
            logger.warning(f"Error calculating days since assignment for {vehicle_id}: {e}")
            return 0

    def create_unassigned_summary(self, unassigned_vehicles: pd.DataFrame) -> dict[str, Any]:
        """
        Create a summary of unassigned vehicles.

        Args:
            unassigned_vehicles: DataFrame of unassigned vehicles.

        Returns:
            Summary dictionary with statistics.
        """
        total_unassigned = len(unassigned_vehicles)

        # Group by vehicle type
        type_counts = {}
        if "Type" in unassigned_vehicles.columns:
            type_counts = unassigned_vehicles["Type"].value_counts().to_dict()

        # Get operational vs non-operational
        operational_count = 0
        if "Opnal? Y/N" in unassigned_vehicles.columns:
            operational_count = len(
                unassigned_vehicles[
                    unassigned_vehicles["Opnal? Y/N"].astype(str).str.upper() == "Y"
                ]
            )

        summary = {
            "total_unassigned": total_unassigned,
            "operational_unassigned": operational_count,
            "non_operational_unassigned": total_unassigned - operational_count,
            "by_type": type_counts,
            "timestamp": datetime.now().isoformat(),
        }

        logger.info(
            f"Unassigned summary: {total_unassigned} total, " f"{operational_count} operational"
        )

        return summary

    def export_unassigned_to_csv(
        self,
        unassigned_vehicles: pd.DataFrame,
        vehicle_log_dict: dict[str, dict],
        output_path: str,
        allocation_date: date,
    ) -> None:
        """
        Export unassigned vehicles to CSV file.

        Args:
            unassigned_vehicles: DataFrame of unassigned vehicles.
            vehicle_log_dict: Dictionary mapping van ID to vehicle details.
            output_path: Path for CSV output.
            allocation_date: Date of the allocation.
        """
        # Prepare export data
        export_data = []

        for _, vehicle in unassigned_vehicles.iterrows():
            van_id = vehicle.get("Van ID", "")
            if not van_id:
                continue

            vehicle_details = vehicle_log_dict.get(van_id, {})

            export_data.append(
                {
                    "Van ID": van_id,
                    "Vehicle Type": vehicle.get("Type", "Unknown"),
                    "Operational Status": vehicle.get("Opnal? Y/N", "N"),
                    "VIN": vehicle_details.get("vin", ""),
                    "GeoTab Code": vehicle_details.get("geotab", ""),
                    "Branded or Rental": vehicle_details.get("brand_or_rental", ""),
                    "Unassigned Date": allocation_date.strftime("%m/%d/%Y"),
                }
            )

        # Create DataFrame and export
        export_df = pd.DataFrame(export_data)
        export_df.to_csv(output_path, index=False)

        logger.info(f"Exported {len(export_df)} unassigned vehicles to {output_path}")
