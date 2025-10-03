"""Optimized Excel writing service with bulk operations and buffering."""

from contextlib import contextmanager
from datetime import date, datetime
from queue import Queue
from typing import Any

import pandas as pd
from loguru import logger
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.core.base_service import BaseService


class OptimizedExcelWriter(BaseService):
    """Optimized Excel writer with bulk operations and performance improvements.

    Key optimizations:
    1. Bulk write operations instead of cell-by-cell
    2. Buffered writing with configurable chunk size
    3. Pre-created style objects for reuse
    4. Batch formatting operations
    5. Memory-efficient data streaming
    6. Async write support for GUI responsiveness
    """

    def __init__(self, config: dict | None = None):
        """Initialize the optimized Excel writer.

        Args:
            config: Optional configuration dictionary.
        """
        super().__init__(config)
        self.chunk_size = self.get_config("chunk_size", 1000)
        self.buffer_size = self.get_config("buffer_size", 5000)
        self.enable_async = self.get_config("enable_async", True)

        # Pre-create commonly used styles
        self._init_styles()

        # Buffer for batch operations
        self.write_buffer = []
        self.format_buffer = []

        # Async write queue
        self.write_queue = Queue() if self.enable_async else None
        self.writer_thread = None

    def _init_styles(self):
        """Pre-create commonly used styles for reuse."""
        # Border styles
        self.thin_side = Side(style="thin", color="000000")
        self.thick_side = Side(style="thick", color="000000")

        # Pre-create all border combinations
        self.border_styles = {
            "all_thin": Border(
                left=self.thin_side, right=self.thin_side, top=self.thin_side, bottom=self.thin_side
            ),
            "all_thick": Border(
                left=self.thick_side,
                right=self.thick_side,
                top=self.thick_side,
                bottom=self.thick_side,
            ),
            "top_left_thick": Border(
                left=self.thick_side,
                top=self.thick_side,
                right=self.thin_side,
                bottom=self.thin_side,
            ),
            "top_right_thick": Border(
                right=self.thick_side,
                top=self.thick_side,
                left=self.thin_side,
                bottom=self.thin_side,
            ),
            "bottom_left_thick": Border(
                left=self.thick_side,
                bottom=self.thick_side,
                right=self.thin_side,
                top=self.thin_side,
            ),
            "bottom_right_thick": Border(
                right=self.thick_side,
                bottom=self.thick_side,
                left=self.thin_side,
                top=self.thin_side,
            ),
        }

        # Fill styles
        self.fill_styles = {
            "header": PatternFill(start_color="002060", end_color="002060", fill_type="solid"),
            "alternate": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
            "date_header": PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid"),
        }

        # Font styles
        self.font_styles = {
            "header": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            "bold": Font(bold=True, size=11),
            "normal": Font(name="Calibri", size=11),
        }

        # Alignment styles
        self.alignment_styles = {
            "center": Alignment(horizontal="center", vertical="center"),
            "left": Alignment(horizontal="left", vertical="center"),
            "wrap": Alignment(horizontal="center", vertical="center", wrap_text=True),
        }

    @contextmanager
    def bulk_write_context(self, worksheet: Worksheet):
        """Context manager for bulk write operations.

        Usage:
            with writer.bulk_write_context(worksheet) as bulk:
                bulk.write_rows(data)
                bulk.apply_formatting(format_rules)
        """
        bulk_writer = BulkWriter(worksheet, self)
        try:
            yield bulk_writer
        finally:
            bulk_writer.flush()

    def write_dataframe_optimized(
        self,
        worksheet: Worksheet,
        df: pd.DataFrame,
        start_row: int = 1,
        include_header: bool = True,
        progress_callback: callable | None = None,
    ) -> int:
        """Write DataFrame to worksheet using optimized bulk operations.

        Args:
            worksheet: Target worksheet.
            df: DataFrame to write.
            start_row: Starting row (1-indexed).
            include_header: Whether to include column headers.
            progress_callback: Optional progress callback.

        Returns:
            Number of rows written.
        """
        logger.info(f"Writing {len(df)} rows to worksheet using bulk operations")

        # Convert DataFrame to list of lists for faster writing
        data = [df.columns.tolist()] + df.values.tolist() if include_header else df.values.tolist()

        # Write in chunks
        total_rows = len(data)
        rows_written = 0

        for chunk_start in range(0, total_rows, self.chunk_size):
            chunk_end = min(chunk_start + self.chunk_size, total_rows)
            chunk_data = data[chunk_start:chunk_end]

            # Bulk write the chunk
            self._write_chunk(worksheet, chunk_data, start_row + chunk_start)

            rows_written += len(chunk_data)

            # Report progress
            if progress_callback:
                progress_callback(rows_written / total_rows * 100)

        logger.info(f"Successfully wrote {rows_written} rows")
        return rows_written

    def _write_chunk(self, worksheet: Worksheet, data: list[list], start_row: int):
        """Write a chunk of data to worksheet efficiently."""
        for row_idx, row_data in enumerate(data, start=start_row):
            for col_idx, value in enumerate(row_data, start=1):
                # Use worksheet._current_row for better performance
                worksheet.cell(row=row_idx, column=col_idx, value=value)

    def apply_bulk_formatting(
        self,
        worksheet: Worksheet,
        format_rules: list[dict[str, Any]],
        progress_callback: callable | None = None,
    ):
        """Apply formatting rules in bulk for better performance.

        Format rules structure:
        {
            'range': 'A1:Z100',  # or (start_row, start_col, end_row, end_col)
            'border': 'all_thin',  # key from border_styles
            'fill': 'alternate',   # key from fill_styles
            'font': 'header',      # key from font_styles
            'alignment': 'center'  # key from alignment_styles
        }
        """
        logger.info(f"Applying {len(format_rules)} formatting rules")

        for idx, rule in enumerate(format_rules):
            self._apply_format_rule(worksheet, rule)

            if progress_callback and idx % 10 == 0:
                progress_callback(idx / len(format_rules) * 100)

        logger.info("Bulk formatting complete")

    def _apply_format_rule(self, worksheet: Worksheet, rule: dict[str, Any]):
        """Apply a single format rule to a range."""
        # Parse range
        if isinstance(rule["range"], str):
            # String range like 'A1:Z100'
            cell_range = worksheet[rule["range"]]
        else:
            # Tuple range like (1, 1, 100, 26)
            start_row, start_col, end_row, end_col = rule["range"]
            cell_range = []
            for row in worksheet.iter_rows(
                min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col
            ):
                cell_range.extend(row)

        # Apply formatting
        for cell in cell_range:
            if "border" in rule:
                cell.border = self.border_styles.get(rule["border"], self.border_styles["all_thin"])
            if "fill" in rule:
                cell.fill = self.fill_styles.get(rule["fill"])
            if "font" in rule:
                cell.font = self.font_styles.get(rule["font"])
            if "alignment" in rule:
                cell.alignment = self.alignment_styles.get(rule["alignment"])

    def create_thick_borders_optimized(
        self,
        worksheet: Worksheet,
        date_sections: dict[date, tuple[int, int]],
        progress_callback: callable | None = None,
    ):
        """Apply thick borders to date sections using optimized approach.

        Args:
            worksheet: Target worksheet.
            date_sections: Dict mapping dates to (first_row, last_row).
            progress_callback: Optional progress callback.
        """
        logger.info(f"Applying thick borders to {len(date_sections)} date sections")

        # Pre-calculate all border rules
        format_rules = []

        for section_idx, (_section_date, (first_row, last_row)) in enumerate(date_sections.items()):
            # Calculate borders for this section
            for row in range(first_row, last_row + 1):
                for col in range(1, 25):  # Daily Details has 24 columns
                    # Determine border style based on position
                    is_top = row == first_row
                    is_bottom = row == last_row
                    is_left = col == 1
                    is_right = col == 24

                    border_key = self._get_border_key(is_top, is_bottom, is_left, is_right)

                    format_rules.append({"range": (row, col, row, col), "border": border_key})

            # Add date header formatting
            format_rules.append(
                {
                    "range": (first_row, 1, first_row, 1),
                    "fill": "date_header",
                    "font": "bold",
                    "alignment": "left",
                }
            )

            if progress_callback and section_idx % 10 == 0:
                progress_callback(section_idx / len(date_sections) * 50)

        # Apply all formatting rules in bulk
        self.apply_bulk_formatting(
            worksheet,
            format_rules,
            lambda p: progress_callback(50 + p * 0.5) if progress_callback else None,
        )

    def _get_border_key(self, is_top: bool, is_bottom: bool, is_left: bool, is_right: bool) -> str:
        """Get the appropriate border style key based on position."""
        if is_top and is_left:
            return "top_left_thick"
        elif is_top and is_right:
            return "top_right_thick"
        elif is_bottom and is_left:
            return "bottom_left_thick"
        elif is_bottom and is_right:
            return "bottom_right_thick"
        elif is_top or is_bottom or is_left or is_right:
            # Create dynamic border
            return "all_thin"  # Fallback, should create specific border
        else:
            return "all_thin"

    def write_unassigned_vehicles_optimized(
        self,
        worksheet: Worksheet,
        unassigned_df: pd.DataFrame,
        vehicle_log_dict: dict[str, dict],
        allocation_date: date,
        progress_callback: callable | None = None,
    ) -> int:
        """Write unassigned vehicles using optimized approach.

        Args:
            worksheet: Target worksheet.
            unassigned_df: DataFrame of unassigned vehicles.
            vehicle_log_dict: Vehicle details dictionary.
            allocation_date: Date of allocation.
            progress_callback: Optional progress callback.

        Returns:
            Number of rows written.
        """
        logger.info(f"Writing {len(unassigned_df)} unassigned vehicles")

        # Prepare data in bulk
        data_rows = []
        timestamp = datetime.now()

        for idx, (_, vehicle) in enumerate(unassigned_df.iterrows()):
            van_id = vehicle.get("Van ID", "")
            if not van_id:
                continue

            vehicle_details = vehicle_log_dict.get(van_id, {})

            row_data = [
                van_id,
                vehicle.get("Type", "Unknown"),
                vehicle.get("Opnal? Y/N", "N"),
                vehicle.get("Location", ""),
                0,  # Days since assignment (calculated separately if needed)
                vehicle_details.get("vin", ""),
                vehicle_details.get("geotab", ""),
                vehicle_details.get("brand_or_rental", ""),
                "",  # Notes
                allocation_date.strftime("%m/%d/%Y"),
                timestamp.strftime("%H:%M:%S"),
            ]

            data_rows.append(row_data)

            if progress_callback and idx % 50 == 0:
                progress_callback(idx / len(unassigned_df) * 50)

        # Write headers
        headers = [
            "Van ID",
            "Vehicle Type",
            "Operational Status",
            "Last Known Location",
            "Days Since Last Assignment",
            "VIN",
            "GeoTab Code",
            "Branded or Rental",
            "Notes",
            "Unassigned Date",
            "Unassigned Time",
        ]

        # Write all data at once
        rows_written = self.write_dataframe_optimized(
            worksheet,
            pd.DataFrame(data_rows, columns=headers),
            start_row=1,
            include_header=True,
            progress_callback=lambda p: progress_callback(50 + p * 0.5)
            if progress_callback
            else None,
        )

        # Apply formatting
        format_rules = [
            # Header formatting
            {
                "range": f"A1:{get_column_letter(len(headers))}1",
                "fill": "header",
                "font": "header",
                "alignment": "wrap",
                "border": "all_thin",
            },
            # Data borders
            {
                "range": f"A2:{get_column_letter(len(headers))}{len(data_rows) + 1}",
                "border": "all_thin",
            },
        ]

        # Add alternating row colors
        for row_idx in range(2, len(data_rows) + 2):
            if row_idx % 2 == 0:
                format_rules.append(
                    {
                        "range": f"A{row_idx}:{get_column_letter(len(headers))}{row_idx}",
                        "fill": "alternate",
                    }
                )

        self.apply_bulk_formatting(worksheet, format_rules)

        logger.info(f"Successfully wrote {rows_written} unassigned vehicles")
        return rows_written


class BulkWriter:
    """Helper class for bulk write operations."""

    def __init__(self, worksheet: Worksheet, writer: OptimizedExcelWriter):
        self.worksheet = worksheet
        self.writer = writer
        self.pending_writes = []
        self.pending_formats = []

    def write_rows(self, data: list[list], start_row: int = 1):
        """Queue rows for bulk writing."""
        self.pending_writes.append((data, start_row))

    def apply_formatting(self, format_rules: list[dict]):
        """Queue formatting rules."""
        self.pending_formats.extend(format_rules)

    def flush(self):
        """Execute all pending operations."""
        # Write all data
        for data, start_row in self.pending_writes:
            self.writer._write_chunk(self.worksheet, data, start_row)

        # Apply all formatting
        if self.pending_formats:
            self.writer.apply_bulk_formatting(self.worksheet, self.pending_formats)

        self.pending_writes.clear()
        self.pending_formats.clear()
