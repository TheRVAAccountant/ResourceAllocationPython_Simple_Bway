"""Daily Details thick border implementation for date-based sections."""

from datetime import date, datetime

from loguru import logger
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet


class DailyDetailsThickBorderService:
    """Service to apply thick borders to Daily Details sheet for date-based sections."""

    def __init__(self):
        """Initialize the thick border service."""
        # Define border styles
        self.thick_side = Side(style="thick", color="000000")
        self.thin_side = Side(style="thin", color="000000")

        # Define date header style
        self.date_header_fill = PatternFill(
            start_color="E6E6E6", end_color="E6E6E6", fill_type="solid"
        )
        self.date_header_font = Font(bold=False, size=11)  # Removed bold for consistency
        self.date_header_alignment = Alignment(horizontal="left", vertical="center")

    def apply_thick_borders_to_daily_details(
        self, worksheet: Worksheet, start_row: int = 2
    ) -> None:
        """
        Apply thick borders to Daily Details sheet, grouping rows by date.

        Each date section will have:
        - Thick border around the entire date section
        - Date displayed prominently in first column
        - All rows for that date contained within the border

        Args:
            worksheet: The Daily Details worksheet.
            start_row: Row to start processing (default 2, after headers).
        """
        logger.info("Applying thick borders to Daily Details sheet")

        # Get all data rows grouped by date
        date_sections = self._identify_date_sections(worksheet, start_row)

        if not date_sections:
            logger.info("No data found to apply borders")
            return

        # Apply thick borders to each date section
        for section_date, (first_row, last_row) in date_sections.items():
            self._apply_thick_border_to_section(worksheet, first_row, last_row, section_date)
            logger.debug(f"Applied thick border to {section_date}: rows {first_row}-{last_row}")

        logger.info(f"Applied thick borders to {len(date_sections)} date sections")

    def _identify_date_sections(
        self, worksheet: Worksheet, start_row: int
    ) -> dict[date, tuple[int, int]]:
        """
        Identify rows grouped by date in the Daily Details sheet.

        Args:
            worksheet: The worksheet to analyze.
            start_row: Starting row for data.

        Returns:
            Dictionary mapping dates to (first_row, last_row) tuples.
        """
        date_sections = {}
        current_date = None
        first_row_for_date = None

        # Find the last row with data
        max_row = worksheet.max_row

        for row_num in range(start_row, max_row + 1):
            # Get date from first column
            date_cell = worksheet.cell(row=row_num, column=1)

            if date_cell.value is None:
                # Empty row, might be end of data
                if current_date and first_row_for_date:
                    # Save the previous section
                    date_sections[current_date] = (first_row_for_date, row_num - 1)
                break

            # Parse the date value
            row_date = self._parse_date_value(date_cell.value)

            if row_date != current_date:
                # New date section
                if current_date and first_row_for_date:
                    # Save the previous section
                    date_sections[current_date] = (first_row_for_date, row_num - 1)

                # Start new section
                current_date = row_date
                first_row_for_date = row_num

        # Don't forget the last section
        if current_date and first_row_for_date:
            date_sections[current_date] = (first_row_for_date, max_row)

        return date_sections

    def _parse_date_value(self, value) -> date | None:
        """
        Parse a date value from various formats.

        Args:
            value: The cell value to parse.

        Returns:
            Parsed date or None if not parseable.
        """
        if value is None:
            return None

        # If it's already a date/datetime object
        if isinstance(value, datetime):
            return value.date()
        elif isinstance(value, date):
            return value

        # Try to parse string formats
        if isinstance(value, str):
            try:
                # Try common date formats
                for fmt in ["%m/%d/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m-%d-%Y"]:
                    try:
                        return datetime.strptime(value, fmt).date()
                    except ValueError:
                        continue
            except Exception:
                pass

        return None

    def _apply_thick_border_to_section(
        self, worksheet: Worksheet, first_row: int, last_row: int, _section_date: date
    ) -> None:
        """
        Apply thick border around a date section.

        Args:
            worksheet: The worksheet to modify.
            first_row: First row of the section.
            last_row: Last row of the section.
            section_date: Date for this section.
        """
        # Determine the column range (1 to 24 for Daily Details)
        first_col = 1
        last_col = 24

        # Apply thick borders to the perimeter
        for row in range(first_row, last_row + 1):
            for col in range(first_col, last_col + 1):
                cell = worksheet.cell(row=row, column=col)

                # Determine which borders to apply
                top = self.thick_side if row == first_row else self.thin_side
                bottom = self.thick_side if row == last_row else self.thin_side
                left = self.thick_side if col == first_col else self.thin_side
                right = self.thick_side if col == last_col else self.thin_side

                # Apply the border
                cell.border = Border(top=top, bottom=bottom, left=left, right=right)

        # Apply consistent formatting to all date cells in the section
        # This ensures visual consistency within each date group
        for row in range(first_row, last_row + 1):
            date_cell = worksheet.cell(row=row, column=1)
            if date_cell.value:
                # Apply consistent formatting to all date cells in the group
                date_cell.fill = self.date_header_fill
                date_cell.font = self.date_header_font
                date_cell.alignment = self.date_header_alignment

    def apply_thick_borders_after_append(
        self, worksheet: Worksheet, new_data_start_row: int, num_new_rows: int
    ) -> None:
        """
        Apply thick borders after appending new data to Daily Details.

        This method is optimized for when new rows are added at the bottom.
        It will:
        1. Check if the new rows continue an existing date section
        2. Update borders accordingly
        3. Create new date sections as needed

        Args:
            worksheet: The Daily Details worksheet.
            new_data_start_row: First row of newly added data.
            num_new_rows: Number of rows added.
        """
        logger.info(
            f"Applying thick borders to {num_new_rows} new rows starting at row {new_data_start_row}"
        )

        # Get the date from the row above (if exists) to check continuity
        previous_date = None
        if new_data_start_row > 2:
            prev_cell = worksheet.cell(row=new_data_start_row - 1, column=1)
            previous_date = self._parse_date_value(prev_cell.value)

        # Process the new rows
        current_date = None
        section_start_row = new_data_start_row

        for row_num in range(new_data_start_row, new_data_start_row + num_new_rows):
            date_cell = worksheet.cell(row=row_num, column=1)
            row_date = self._parse_date_value(date_cell.value)

            if row_date != current_date:
                # New date section within the added rows
                if current_date and section_start_row < row_num:
                    # Apply border to previous section
                    self._apply_thick_border_to_section(
                        worksheet, section_start_row, row_num - 1, current_date
                    )

                # Check if this continues the previous section
                if row_date == previous_date and row_num == new_data_start_row:
                    # Need to update the previous section's bottom border
                    self._extend_previous_section(worksheet, new_data_start_row - 1)
                    section_start_row = new_data_start_row
                else:
                    section_start_row = row_num

                current_date = row_date

        # Apply border to the last section
        if current_date and section_start_row <= new_data_start_row + num_new_rows - 1:
            self._apply_thick_border_to_section(
                worksheet, section_start_row, new_data_start_row + num_new_rows - 1, current_date
            )

    def _extend_previous_section(self, worksheet: Worksheet, previous_last_row: int) -> None:
        """
        Extend the previous date section by removing its bottom thick border.

        Args:
            worksheet: The worksheet to modify.
            previous_last_row: The last row of the previous section.
        """
        # Remove thick bottom border from previous section's last row
        for col in range(1, 25):
            cell = worksheet.cell(row=previous_last_row, column=col)
            current_border = cell.border

            # Replace thick bottom with thin
            cell.border = Border(
                top=current_border.top,
                bottom=self.thin_side,  # Change from thick to thin
                left=current_border.left,
                right=current_border.right,
            )
