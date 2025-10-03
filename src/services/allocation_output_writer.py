"""Service for writing allocation results to separate output files."""

import os
from datetime import date, datetime
from pathlib import Path
from typing import Dict, Optional

import pandas as pd
from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.core.base_service import BaseService
from src.models.allocation import AllocationResult


class AllocationOutputWriter(BaseService):
    """Service to write allocation results to separate output files in the outputs directory."""

    # Results sheet columns (11 columns matching GAS format)
    RESULTS_COLUMNS = [
        "Route Code",
        "Van ID",
        "Driver Name",
        "Employee ID",
        "Service Type",
        "Vehicle Type",
        "VIN",
        "Staging Location",
        "Wave Time",
        "Route Type",
        "Notes"
    ]

    # Unassigned vehicles sheet columns
    UNASSIGNED_COLUMNS = [
        "Van ID",
        "Vehicle Type",
        "VIN",
        "GeoTab Code",
        "Type",
        "Staging Location",
        "Last Driver",
        "Last Route",
        "Days Since Last Assignment",
        "Operational Status",
        "Notes"
    ]

    def __init__(self, output_dir: str = "outputs"):
        """Initialize the allocation output writer.
        
        Args:
            output_dir: Directory where output files will be saved.
        """
        super().__init__()
        self.output_dir = Path(output_dir)
        self._initialized = False
        
    def initialize(self) -> None:
        """Initialize the service and create output directory if needed."""
        if self._initialized:
            return
            
        logger.info(f"Initializing AllocationOutputWriter with output directory: {self.output_dir}")
        
        # Create output directory if it doesn't exist
        try:
            self.output_dir.mkdir(exist_ok=True)
            logger.info(f"Output directory ready: {self.output_dir}")
        except Exception as e:
            logger.error(f"Failed to create output directory: {e}")
            raise
            
        self._initialized = True
        
    def validate(self) -> bool:
        """Validate that the service is properly configured."""
        if not self._initialized:
            return False
            
        # Check if output directory exists and is writable
        if not self.output_dir.exists():
            logger.error(f"Output directory does not exist: {self.output_dir}")
            return False
            
        if not os.access(self.output_dir, os.W_OK):
            logger.error(f"Output directory is not writable: {self.output_dir}")
            return False
            
        return True
        
    def create_results_file(
        self,
        allocation_result: AllocationResult,
        unassigned_vehicles: pd.DataFrame,
        vehicle_log_dict: Dict[str, Dict],
        allocation_date: Optional[date] = None,
        file_suffix: Optional[str] = None
    ) -> str:
        """Create a new Excel file with allocation results.
        
        Args:
            allocation_result: The allocation results to write.
            unassigned_vehicles: DataFrame of unassigned vehicles.
            vehicle_log_dict: Dictionary of vehicle information from Vehicle Log.
            allocation_date: Date of allocation (defaults to today).
            file_suffix: Optional suffix for filename (e.g., "_v2").
            
        Returns:
            Path to the created file.
        """
        if not self.validate():
            raise ValueError("Service not properly initialized or validated")
            
        # Use provided date or today
        alloc_date = allocation_date or date.today()
        
        # Generate filename
        base_filename = f"Allocation_Results_{alloc_date.strftime('%Y-%m-%d')}"
        if file_suffix:
            base_filename += file_suffix
        filename = f"{base_filename}.xlsx"
        file_path = self.output_dir / filename
        
        # Check if file already exists and create versioned name if needed
        version = 1
        while file_path.exists():
            version += 1
            filename = f"{base_filename}_v{version}.xlsx"
            file_path = self.output_dir / filename
            
        logger.info(f"Creating allocation results file: {file_path}")
        
        try:
            # Create workbook
            workbook = Workbook()
            
            # Remove default sheet
            if "Sheet" in workbook.sheetnames:
                workbook.remove(workbook["Sheet"])
                
            # Create Results sheet
            self._create_results_sheet(workbook, allocation_result, vehicle_log_dict, alloc_date)
            
            # Create Unassigned Vehicles sheet
            self._create_unassigned_sheet(workbook, unassigned_vehicles, vehicle_log_dict, alloc_date)
            
            # Save workbook
            workbook.save(file_path)
            logger.info(f"Successfully created allocation results file: {file_path}")
            
            return str(file_path)
            
        except Exception as e:
            logger.error(f"Failed to create allocation results file: {e}")
            raise
            
    def _create_results_sheet(
        self,
        workbook: Workbook,
        allocation_result: AllocationResult,
        vehicle_log_dict: Dict[str, Dict],
        allocation_date: date
    ) -> None:
        """Create the Results sheet with allocation data."""
        sheet = workbook.create_sheet("Results")
        
        # Set up header formatting
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        for col_num, header in enumerate(self.RESULTS_COLUMNS, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            
        # Apply borders to headers
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        for col in range(1, len(self.RESULTS_COLUMNS) + 1):
            sheet.cell(row=1, column=col).border = thin_border
            
        # Write allocation data
        row_num = 2
        metadata = allocation_result.metadata or {}
        detailed_results = metadata.get("detailed_results", [])
        
        # Log debug info about the data we're about to write
        logger.debug(f"Writing Results sheet - metadata keys: {list(metadata.keys())}")
        logger.debug(f"Number of detailed_results: {len(detailed_results)}")
        if detailed_results and len(detailed_results) > 0:
            logger.debug(f"First detailed result: {detailed_results[0]}")
        
        # Add current date/time to Notes
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        for result in detailed_results:
            # Get vehicle info from vehicle log - Handle field name variations
            van_id = result.get("Van ID", result.get("van_id", ""))
            vehicle_info = vehicle_log_dict.get(van_id, {})
            
            # Prepare row data - Handle field name variations from GAS allocator
            row_data = [
                result.get("Route Code", result.get("route_code", "")),
                van_id,
                result.get("Associate Name", result.get("driver_name", "")),
                result.get("employee_id", ""),  # Not provided by GAS allocator
                result.get("Service Type", result.get("service_type", "")),
                result.get("Van Type", result.get("vehicle_type", "")),
                vehicle_info.get("vin", vehicle_info.get("VIN", "")),  # Handle both cases
                result.get("Staging Location", result.get("staging_location", "")),
                result.get("Wave", result.get("wave_time", "")),
                result.get("route_type", ""),  # Not provided by GAS allocator
                f"Allocated on {current_time}"
            ]
            
            # Write row
            for col_num, value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                
            row_num += 1
            
        # Auto-fit columns
        for col_num in range(1, len(self.RESULTS_COLUMNS) + 1):
            column_letter = get_column_letter(col_num)
            # Set reasonable column widths
            if col_num == 1:  # Route Code
                sheet.column_dimensions[column_letter].width = 12
            elif col_num == 2:  # Van ID
                sheet.column_dimensions[column_letter].width = 10
            elif col_num == 3:  # Driver Name
                sheet.column_dimensions[column_letter].width = 20
            elif col_num == 5:  # Service Type
                sheet.column_dimensions[column_letter].width = 30
            elif col_num == 7:  # VIN
                sheet.column_dimensions[column_letter].width = 20
            elif col_num == 11:  # Notes
                sheet.column_dimensions[column_letter].width = 25
            else:
                sheet.column_dimensions[column_letter].width = 15
                
        # Freeze header row
        sheet.freeze_panes = "A2"
        
        # Add autofilter
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(self.RESULTS_COLUMNS))}{row_num-1}"
        
        logger.info(f"Created Results sheet with {row_num-2} allocation records")
        
    def _create_unassigned_sheet(
        self,
        workbook: Workbook,
        unassigned_vehicles: pd.DataFrame,
        vehicle_log_dict: Dict[str, Dict],
        allocation_date: date
    ) -> None:
        """Create the Unassigned Vehicles sheet."""
        sheet = workbook.create_sheet("Unassigned Vehicles")
        
        # Set up header formatting
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")  # Orange
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        for col_num, header in enumerate(self.UNASSIGNED_COLUMNS, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            
        # Apply borders
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        for col in range(1, len(self.UNASSIGNED_COLUMNS) + 1):
            sheet.cell(row=1, column=col).border = thin_border
            
        # Write unassigned vehicle data
        row_num = 2
        
        logger.debug(f"Writing unassigned vehicles - DataFrame shape: {unassigned_vehicles.shape}")
        logger.debug(f"Unassigned vehicles columns: {list(unassigned_vehicles.columns)}")
        
        if not unassigned_vehicles.empty:
            logger.debug(f"First unassigned vehicle: {unassigned_vehicles.iloc[0].to_dict()}")
            for _, vehicle in unassigned_vehicles.iterrows():
                van_id = str(vehicle.get("Van ID", ""))
                vehicle_info = vehicle_log_dict.get(van_id, {})
                
                # Prepare row data
                row_data = [
                    van_id,
                    str(vehicle.get("Type", "")),  # Use "Type" not "Vehicle Type"
                    vehicle_info.get("vin", vehicle_info.get("VIN", "")),  # Handle both cases
                    vehicle_info.get("geotab", vehicle_info.get("GeoTab", "")),  # Handle both cases
                    vehicle_info.get("brand_or_rental", vehicle_info.get("Branded or Rental", "")),  # Handle both cases
                    str(vehicle.get("Staging Location", "")),  # Get from vehicle data
                    "",  # Last Driver (would need historical data)
                    "",  # Last Route (would need historical data)
                    "",  # Days Since Last Assignment
                    "Available",  # Operational Status
                    f"Unassigned on {allocation_date.strftime('%Y-%m-%d')}"
                ]
                
                # Write row
                for col_num, value in enumerate(row_data, 1):
                    cell = sheet.cell(row=row_num, column=col_num, value=value)
                    cell.border = thin_border
                    
                # Apply alternating row colors
                if row_num % 2 == 0:
                    row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                    for col in range(1, len(self.UNASSIGNED_COLUMNS) + 1):
                        sheet.cell(row=row_num, column=col).fill = row_fill
                        
                row_num += 1
                
        # Auto-fit columns
        for col_num in range(1, len(self.UNASSIGNED_COLUMNS) + 1):
            column_letter = get_column_letter(col_num)
            # Set reasonable column widths
            if col_num == 3:  # VIN
                sheet.column_dimensions[column_letter].width = 20
            elif col_num == 6:  # Staging Location
                sheet.column_dimensions[column_letter].width = 18
            elif col_num == 11:  # Notes
                sheet.column_dimensions[column_letter].width = 30
            else:
                sheet.column_dimensions[column_letter].width = 15
                
        # Freeze header row
        sheet.freeze_panes = "A2"
        
        # Add autofilter
        if row_num > 2:
            sheet.auto_filter.ref = f"A1:{get_column_letter(len(self.UNASSIGNED_COLUMNS))}{row_num-1}"
            
        logger.info(f"Created Unassigned Vehicles sheet with {row_num-2} vehicles")