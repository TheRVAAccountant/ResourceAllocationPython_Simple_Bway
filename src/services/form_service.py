"""Form service for Excel-based data entry."""

from typing import Optional, Any, Union
from datetime import datetime, date
from pathlib import Path
from loguru import logger
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Protection, PatternFill, Font, Alignment

from src.core.base_service import BaseService, timer, error_handler
from src.models.excel import ExcelValidation
from src.services.excel_service import ExcelService


class FormField:
    """Represents a form field in Excel."""
    
    def __init__(
        self,
        name: str,
        label: str,
        field_type: str,
        required: bool = False,
        default: Any = None,
        validation: Optional[ExcelValidation] = None,
        options: Optional[list] = None,
        help_text: Optional[str] = None
    ):
        """Initialize form field.
        
        Args:
            name: Field name.
            label: Display label.
            field_type: Field type (text, number, date, dropdown, checkbox).
            required: Whether field is required.
            default: Default value.
            validation: Excel validation rules.
            options: Options for dropdown.
            help_text: Help text for field.
        """
        self.name = name
        self.label = label
        self.field_type = field_type
        self.required = required
        self.default = default
        self.validation = validation
        self.options = options
        self.help_text = help_text
        self.cell_reference = None


class ExcelForm:
    """Represents an Excel-based form."""
    
    def __init__(self, form_id: str, title: str, description: Optional[str] = None):
        """Initialize Excel form.
        
        Args:
            form_id: Form identifier.
            title: Form title.
            description: Form description.
        """
        self.form_id = form_id
        self.title = title
        self.description = description
        self.fields: list[FormField] = []
        self.layout = "vertical"  # vertical or horizontal
        self.protected = True
        self.submission_handler = None
    
    def add_field(self, field: FormField):
        """Add field to form.
        
        Args:
            field: Form field to add.
        """
        self.fields.append(field)
    
    def get_field(self, name: str) -> Optional[FormField]:
        """Get field by name.
        
        Args:
            name: Field name.
        
        Returns:
            Form field or None.
        """
        for field in self.fields:
            if field.name == name:
                return field
        return None


class FormService(BaseService):
    """Service for creating and managing Excel forms.
    
    Provides functionality for creating data entry forms
    in Excel with validation and submission handling.
    """
    
    def __init__(self, config: Optional[dict[str, Any]] = None):
        """Initialize the form service.
        
        Args:
            config: Configuration dictionary.
        """
        super().__init__(config)
        
        self.excel_service = None
        self.forms: dict[str, ExcelForm] = {}
        self.submissions: list[dict] = []
        
        # Configuration
        self.form_template_path = self.get_config("form_template_path")
        self.enable_protection = self.get_config("enable_form_protection", True)
        self.auto_validate = self.get_config("auto_validate_forms", True)
    
    def initialize(self) -> None:
        """Initialize the form service."""
        logger.info("Initializing Form Service")
        
        # Initialize Excel service
        self.excel_service = ExcelService(self.config)
        self.excel_service.initialize()
        
        # Load default forms
        self._load_default_forms()
        
        self._initialized = True
    
    def validate(self) -> bool:
        """Validate the service configuration.
        
        Returns:
            True if configuration is valid.
        """
        return True
    
    def _load_default_forms(self):
        """Load default form definitions."""
        # Vehicle entry form
        vehicle_form = ExcelForm(
            "vehicle_entry",
            "Vehicle Entry Form",
            "Enter vehicle information"
        )
        
        vehicle_form.add_field(FormField(
            "vehicle_number",
            "Vehicle Number",
            "text",
            required=True,
            validation=ExcelValidation(
                type="textLength",
                operator="between",
                formula1="3",
                formula2="20",
                error_message="Vehicle number must be 3-20 characters"
            )
        ))
        
        vehicle_form.add_field(FormField(
            "vehicle_type",
            "Vehicle Type",
            "dropdown",
            required=True,
            options=["Standard", "Premium", "Economy", "Luxury", "SUV", "Truck", "Van"]
        ))
        
        vehicle_form.add_field(FormField(
            "location",
            "Location",
            "dropdown",
            required=True,
            options=["Main", "North", "South", "East", "West"]
        ))
        
        vehicle_form.add_field(FormField(
            "status",
            "Status",
            "dropdown",
            required=True,
            default="available",
            options=["available", "in_use", "maintenance", "retired"]
        ))
        
        vehicle_form.add_field(FormField(
            "capacity",
            "Capacity",
            "number",
            validation=ExcelValidation(
                type="whole",
                operator="greaterThan",
                formula1="0",
                error_message="Capacity must be positive"
            )
        ))
        
        self.forms["vehicle_entry"] = vehicle_form
        
        # Driver entry form
        driver_form = ExcelForm(
            "driver_entry",
            "Driver Entry Form",
            "Enter driver information"
        )
        
        driver_form.add_field(FormField(
            "employee_id",
            "Employee ID",
            "text",
            required=True
        ))
        
        driver_form.add_field(FormField(
            "name",
            "Driver Name",
            "text",
            required=True
        ))
        
        driver_form.add_field(FormField(
            "location",
            "Location",
            "dropdown",
            required=True,
            options=["Main", "North", "South", "East", "West"]
        ))
        
        driver_form.add_field(FormField(
            "experience_years",
            "Years of Experience",
            "number",
            validation=ExcelValidation(
                type="whole",
                operator="between",
                formula1="0",
                formula2="50",
                error_message="Experience must be 0-50 years"
            )
        ))
        
        driver_form.add_field(FormField(
            "license_type",
            "License Type",
            "dropdown",
            options=["Standard", "CDL-A", "CDL-B", "Chauffeur"]
        ))
        
        self.forms["driver_entry"] = driver_form
    
    @timer
    @error_handler
    def create_form_workbook(self, form_id: str, output_path: str) -> bool:
        """Create an Excel workbook with a form.
        
        Args:
            form_id: Form identifier.
            output_path: Path to save workbook.
        
        Returns:
            True if created successfully.
        """
        form = self.forms.get(form_id)
        if not form:
            logger.error(f"Form not found: {form_id}")
            return False
        
        try:
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Form"
            
            # Add title
            ws["A1"] = form.title
            ws["A1"].font = Font(size=16, bold=True)
            ws.merge_cells("A1:D1")
            
            # Add description
            if form.description:
                ws["A2"] = form.description
                ws["A2"].font = Font(size=12, italic=True)
                ws.merge_cells("A2:D2")
            
            # Add fields
            current_row = 4
            for field in form.fields:
                # Label
                ws[f"A{current_row}"] = field.label
                ws[f"A{current_row}"].font = Font(bold=True)
                
                if field.required:
                    ws[f"A{current_row}"].value += " *"
                
                # Input cell
                input_cell = f"C{current_row}"
                field.cell_reference = input_cell
                
                # Set default value
                if field.default:
                    ws[input_cell] = field.default
                
                # Apply validation
                if field.field_type == "dropdown" and field.options:
                    dv = DataValidation(
                        type="list",
                        formula1=f'"{",".join(field.options)}"',
                        showDropDown=True
                    )
                    dv.add(ws[input_cell])
                    ws.add_data_validation(dv)
                
                elif field.validation:
                    dv = DataValidation(
                        type=field.validation.type,
                        operator=field.validation.operator,
                        formula1=field.validation.formula1,
                        formula2=field.validation.formula2,
                        showErrorMessage=True,
                        errorTitle="Validation Error",
                        error=field.validation.error_message
                    )
                    dv.add(ws[input_cell])
                    ws.add_data_validation(dv)
                
                # Add help text
                if field.help_text:
                    ws[f"D{current_row}"] = field.help_text
                    ws[f"D{current_row}"].font = Font(size=10, italic=True, color="808080")
                
                # Style input cell
                ws[input_cell].fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                ws[input_cell].alignment = Alignment(horizontal="left", vertical="center")
                
                current_row += 2
            
            # Add submit button area
            ws[f"A{current_row + 1}"] = "Click to Submit"
            ws[f"A{current_row + 1}"].font = Font(bold=True, color="FFFFFF")
            ws[f"A{current_row + 1}"].fill = PatternFill(start_color="007BFF", end_color="007BFF", fill_type="solid")
            ws[f"A{current_row + 1}"].alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells(f"A{current_row + 1}:C{current_row + 1}")
            
            # Add instructions
            ws[f"A{current_row + 3}"] = "Instructions:"
            ws[f"A{current_row + 3}"].font = Font(bold=True)
            ws[f"A{current_row + 4}"] = "1. Fill in all required fields (marked with *)"
            ws[f"A{current_row + 5}"] = "2. Save the file when complete"
            ws[f"A{current_row + 6}"] = "3. Submit using the button above"
            
            # Protect sheet if enabled
            if self.enable_protection and form.protected:
                # Unlock input cells
                for field in form.fields:
                    if field.cell_reference:
                        ws[field.cell_reference].protection = Protection(locked=False)
                
                # Protect sheet
                ws.protection.sheet = True
                ws.protection.password = "resource_allocation"
            
            # Adjust column widths
            ws.column_dimensions["A"].width = 20
            ws.column_dimensions["B"].width = 5
            ws.column_dimensions["C"].width = 30
            ws.column_dimensions["D"].width = 40
            
            # Save workbook
            wb.save(output_path)
            logger.info(f"Form workbook created: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to create form workbook: {e}")
            return False
    
    @timer
    @error_handler
    def read_form_submission(self, file_path: str, form_id: str) -> Optional[dict]:
        """Read form submission from Excel file.
        
        Args:
            file_path: Path to Excel file.
            form_id: Form identifier.
        
        Returns:
            Submission data or None.
        """
        form = self.forms.get(form_id)
        if not form:
            logger.error(f"Form not found: {form_id}")
            return None
        
        try:
            # Open workbook
            self.excel_service.open_workbook(file_path)
            sheet = self.excel_service.get_sheet("Form")
            
            # Read field values
            submission = {
                "form_id": form_id,
                "timestamp": datetime.now(),
                "file_path": file_path,
                "data": {}
            }
            
            for field in form.fields:
                if field.cell_reference:
                    # Extract row and column from cell reference
                    import re
                    match = re.match(r"([A-Z]+)(\d+)", field.cell_reference)
                    if match:
                        col_letter = match.group(1)
                        row_num = int(match.group(2))
                        
                        # Read value
                        if self.excel_service.use_xlwings:
                            value = sheet.range(field.cell_reference).value
                        else:
                            col_num = ord(col_letter) - ord("A") + 1
                            value = sheet.cell(row=row_num, column=col_num).value
                        
                        submission["data"][field.name] = value
            
            # Validate submission
            if self.auto_validate:
                validation_errors = self.validate_submission(submission, form)
                if validation_errors:
                    submission["validation_errors"] = validation_errors
                    submission["is_valid"] = False
                else:
                    submission["is_valid"] = True
            
            # Store submission
            self.submissions.append(submission)
            
            logger.info(f"Form submission read: {form_id}")
            return submission
            
        except Exception as e:
            logger.error(f"Failed to read form submission: {e}")
            return None
    
    def validate_submission(self, submission: dict, form: ExcelForm) -> list[str]:
        """Validate form submission.
        
        Args:
            submission: Submission data.
            form: Form definition.
        
        Returns:
            List of validation errors.
        """
        errors = []
        data = submission.get("data", {})
        
        for field in form.fields:
            value = data.get(field.name)
            
            # Check required fields
            if field.required and (value is None or value == ""):
                errors.append(f"{field.label} is required")
            
            # Type validation
            if value is not None:
                if field.field_type == "number":
                    try:
                        float(value)
                    except:
                        errors.append(f"{field.label} must be a number")
                
                elif field.field_type == "date":
                    try:
                        if isinstance(value, str):
                            datetime.strptime(value, "%Y-%m-%d")
                    except:
                        errors.append(f"{field.label} must be a valid date")
                
                elif field.field_type == "dropdown" and field.options:
                    if value not in field.options:
                        errors.append(f"{field.label} must be one of: {', '.join(field.options)}")
        
        return errors
    
    def add_form(self, form: ExcelForm):
        """Add form definition.
        
        Args:
            form: Form to add.
        """
        self.forms[form.form_id] = form
        logger.info(f"Form added: {form.form_id}")
    
    def get_form(self, form_id: str) -> Optional[ExcelForm]:
        """Get form by ID.
        
        Args:
            form_id: Form identifier.
        
        Returns:
            Form or None.
        """
        return self.forms.get(form_id)
    
    def get_submissions(self, form_id: Optional[str] = None) -> list[dict]:
        """Get form submissions.
        
        Args:
            form_id: Optional form ID filter.
        
        Returns:
            List of submissions.
        """
        if form_id:
            return [s for s in self.submissions if s.get("form_id") == form_id]
        return self.submissions
    
    def export_submissions(self, output_path: str, form_id: Optional[str] = None) -> bool:
        """Export submissions to Excel.
        
        Args:
            output_path: Output file path.
            form_id: Optional form ID filter.
        
        Returns:
            True if exported successfully.
        """
        try:
            submissions = self.get_submissions(form_id)
            
            if not submissions:
                logger.warning("No submissions to export")
                return False
            
            # Convert to DataFrame
            data = []
            for submission in submissions:
                row = {
                    "Form ID": submission["form_id"],
                    "Timestamp": submission["timestamp"],
                    "File Path": submission["file_path"],
                    "Valid": submission.get("is_valid", "Unknown")
                }
                row.update(submission.get("data", {}))
                data.append(row)
            
            df = pd.DataFrame(data)
            
            # Export to Excel
            df.to_excel(output_path, index=False)
            
            logger.info(f"Exported {len(submissions)} submissions to: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to export submissions: {e}")
            return False