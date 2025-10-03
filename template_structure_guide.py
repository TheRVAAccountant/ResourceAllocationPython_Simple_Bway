#!/usr/bin/env python3
"""
Template Structure Guide for Daily Summary Log Output

This module provides the exact structure definitions for creating output files
that match the Daily Summary Log 2025.xlsx template format.
"""

from typing import Dict, List, Any, Optional
from dataclasses import dataclass
from datetime import datetime, time
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

@dataclass
class ColumnDefinition:
    """Definition of a column in the output template."""
    name: str
    data_type: str  # 'datetime', 'str', 'float', 'int', 'time'
    excel_format: Optional[str] = None
    width: Optional[float] = None
    required: bool = True

@dataclass
class SheetTemplate:
    """Template definition for a sheet."""
    name: str
    columns: List[ColumnDefinition]
    header_style: Dict[str, Any]
    data_validation: Dict[str, Any]

class TemplateStructure:
    """Complete structure definition for the Daily Summary Log template."""
    
    # Color definitions from the template
    HEADER_COLOR = "FF46BDC6"  # Teal background for headers
    HEADER_FONT_COLOR = "FF000000"  # Black text
    
    # Font definitions
    HEADER_FONT = Font(
        bold=True,
        size=11,
        name='Calibri',
        color=HEADER_FONT_COLOR
    )
    
    DATA_FONT = Font(
        bold=False,
        size=11,
        name='Calibri'
    )
    
    # Fill definitions
    HEADER_FILL = PatternFill(
        start_color=HEADER_COLOR,
        end_color=HEADER_COLOR,
        fill_type='solid'
    )
    
    # Border definitions
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    @classmethod
    def get_daily_details_template(cls) -> SheetTemplate:
        """Get the Daily Details sheet template."""
        columns = [
            ColumnDefinition("Date", "datetime", "mm/dd/yyyy", 12.0),
            ColumnDefinition("Route #", "str", None, 10.0),
            ColumnDefinition("Name", "str", None, 20.0),
            ColumnDefinition("Asset ID", "str", None, 12.0),
            ColumnDefinition("Van ID", "str", None, 10.0),
            ColumnDefinition("VIN", "str", None, 17.0),
            ColumnDefinition("GeoTab Code", "str", None, 12.0),
            ColumnDefinition("Type", "str", None, 15.0),
            ColumnDefinition("Vehicle Type", "str", None, 15.0),
            ColumnDefinition("Route Type", "str", None, 15.0),
            ColumnDefinition("Rescue", "str", None, 10.0, False),
            ColumnDefinition("Delivery Pace 1:40pm", "float", "0.00", 15.0, False),
            ColumnDefinition("Delivery Pace 3:40pm", "float", "0.00", 15.0, False),
            ColumnDefinition("Delivery Pace 5:40pm", "float", "0.00", 15.0, False),
            ColumnDefinition("Delivery Pace 7:40pm", "float", "0.00", 15.0, False),
            ColumnDefinition("Delivery Pace 9:40pm", "float", "0.00", 15.0, False),
            ColumnDefinition("RTS TIME", "time", "h:mm am/pm", 12.0, False),
            ColumnDefinition("Pkg. Delivered", "int", "0", 12.0, False),
            ColumnDefinition("Pkg. Returned", "int", "0", 12.0, False),
            ColumnDefinition("Route Notes", "str", None, 25.0, False),
            ColumnDefinition("Week Number", "float", "0", 12.0, False),
            ColumnDefinition("Unique Identifier", "str", None, 20.0),
            ColumnDefinition("Vehicle Inspection", "str", None, 18.0, False),
            ColumnDefinition("Route Completion", "str", None, 18.0, False),
        ]
        
        header_style = {
            'font': cls.HEADER_FONT,
            'fill': cls.HEADER_FILL,
            'border': cls.THIN_BORDER,
            'alignment': Alignment(horizontal='center', vertical='center')
        }
        
        return SheetTemplate(
            name="Daily Details",
            columns=columns,
            header_style=header_style,
            data_validation={}
        )
    
    @classmethod
    def get_available_unassigned_template(cls, date_str: str) -> SheetTemplate:
        """Get the Available & Unassigned sheet template."""
        columns = [
            ColumnDefinition("Van ID", "str", None, 10.0),
            ColumnDefinition("Year", "float", "0", 8.0),
            ColumnDefinition("Make", "str", None, 12.0),
            ColumnDefinition("Model", "str", None, 15.0),
            ColumnDefinition("Style", "str", None, 15.0),
            ColumnDefinition("Type", "str", None, 12.0),
            ColumnDefinition("License Tag Number", "str", None, 18.0),
            ColumnDefinition("License Tag State", "str", None, 15.0),
            ColumnDefinition("Ownership", "str", None, 12.0),
            ColumnDefinition("VIN", "str", None, 17.0),
            ColumnDefinition("Van Type", "str", None, 15.0),
            ColumnDefinition("Issue", "str", None, 20.0, False),
            ColumnDefinition("Date GROUNDED", "datetime", "mm/dd/yyyy", 15.0, False),
            ColumnDefinition("Date UNGROUNDED", "datetime", "mm/dd/yyyy", 15.0, False),
            ColumnDefinition("Opnal? Y/N", "str", None, 12.0, False),
        ]
        
        header_style = {
            'font': cls.HEADER_FONT,
            'fill': cls.HEADER_FILL,
            'border': cls.THIN_BORDER,
            'alignment': Alignment(horizontal='center', vertical='center')
        }
        
        return SheetTemplate(
            name=f"{date_str} - Available & Unassign",
            columns=columns,
            header_style=header_style,
            data_validation={}
        )
    
    @classmethod
    def get_results_template(cls, date_str: str) -> SheetTemplate:
        """Get the Results sheet template."""
        columns = [
            ColumnDefinition("Route Code", "str", None, 12.0),
            ColumnDefinition("Service Type", "str", None, 35.0),
            ColumnDefinition("DSP", "str", None, 8.0),
            ColumnDefinition("Wave", "time", "h:mm am/pm", 10.0),
            ColumnDefinition("Staging Location", "str", None, 15.0),
            ColumnDefinition("Van ID", "str", None, 10.0),
            ColumnDefinition("Device Name", "str", None, 15.0),
            ColumnDefinition("Van Type", "str", None, 15.0),
            ColumnDefinition("Operational", "str", None, 12.0),
            ColumnDefinition("Associate Name", "str", None, 20.0),
            ColumnDefinition("Unique Identifier", "str", None, 20.0),
        ]
        
        header_style = {
            'font': cls.HEADER_FONT,
            'fill': cls.HEADER_FILL,
            'border': cls.THIN_BORDER,
            'alignment': Alignment(horizontal='center', vertical='center')
        }
        
        return SheetTemplate(
            name=f"{date_str} - Results",
            columns=columns,
            header_style=header_style,
            data_validation={}
        )

class TemplateWriter:
    """Utility class to write data using the template structure."""
    
    def __init__(self, workbook_path: str):
        self.workbook_path = workbook_path
        self.workbook = openpyxl.Workbook()
        # Remove the default sheet
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])
    
    def create_sheet_from_template(self, template: SheetTemplate, data: List[Dict[str, Any]] = None) -> None:
        """Create a sheet from template with optional data."""
        worksheet = self.workbook.create_sheet(title=template.name)
        
        # Set up headers
        for col_idx, column_def in enumerate(template.columns, 1):
            cell = worksheet.cell(row=1, column=col_idx, value=column_def.name)
            
            # Apply header styling
            if template.header_style:
                for style_attr, style_value in template.header_style.items():
                    setattr(cell, style_attr, style_value)
            
            # Set column width
            if column_def.width:
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                worksheet.column_dimensions[col_letter].width = column_def.width
        
        # Add data if provided
        if data:
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, column_def in enumerate(template.columns, 1):
                    value = row_data.get(column_def.name)
                    cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                    
                    # Apply number formatting
                    if column_def.excel_format and value is not None:
                        cell.number_format = column_def.excel_format
                    
                    # Apply data font
                    cell.font = TemplateStructure.DATA_FONT
    
    def save(self) -> None:
        """Save the workbook."""
        self.workbook.save(self.workbook_path)
        print(f"Template workbook saved to: {self.workbook_path}")

def create_sample_template():
    """Create a sample template file for testing."""
    from datetime import datetime, time
    
    # Sample data for Daily Details
    daily_details_data = [
        {
            "Date": datetime(2025, 8, 4),
            "Route #": "CX1",
            "Name": "John Doe",
            "Asset ID": "BW100",
            "Van ID": "BW100",
            "VIN": "1FTBW1CM5GKA12345",
            "GeoTab Code": "BW100-GT",
            "Type": "Standard",
            "Vehicle Type": "Large Van",
            "Route Type": "Standard Delivery",
            "Unique Identifier": "CX1-20250804-001"
        },
        {
            "Date": datetime(2025, 8, 4),
            "Route #": "CX2",
            "Name": "Jane Smith", 
            "Asset ID": "BW101",
            "Van ID": "BW101",
            "VIN": "1FTBW1CM5GKA67890",
            "GeoTab Code": "BW101-GT",
            "Type": "Standard",
            "Vehicle Type": "Large Van",
            "Route Type": "Standard Delivery",
            "Unique Identifier": "CX2-20250804-001"
        }
    ]
    
    # Sample data for Results
    results_data = [
        {
            "Route Code": "CX1",
            "Service Type": "Standard Parcel - Large Van - US",
            "DSP": "BWAY",
            "Wave": time(10, 45),
            "Staging Location": "STG.G.1",
            "Van ID": "BW100",
            "Device Name": "BW100-DEV",
            "Van Type": "Large",
            "Operational": "Yes",
            "Associate Name": "John Doe",
            "Unique Identifier": "CX1-20250804-001"
        }
    ]
    
    # Sample data for Available & Unassigned
    available_data = [
        {
            "Van ID": "BW102",
            "Year": 2023.0,
            "Make": "FORD",
            "Model": "TRANSIT",
            "Style": "TRANSIT 150",
            "Type": "Large",
            "License Tag Number": "ABC123",
            "License Tag State": "VA",
            "Ownership": "Company",
            "VIN": "1FTBW1CM5GKA11111",
            "Van Type": "Large",
            "Opnal? Y/N": "Y"
        }
    ]
    
    # Create workbook
    date_str = "08-04-25"
    writer = TemplateWriter("sample_template_output.xlsx")
    
    # Create sheets
    daily_template = TemplateStructure.get_daily_details_template()
    writer.create_sheet_from_template(daily_template, daily_details_data)
    
    results_template = TemplateStructure.get_results_template(date_str)
    writer.create_sheet_from_template(results_template, results_data)
    
    available_template = TemplateStructure.get_available_unassigned_template(date_str)
    writer.create_sheet_from_template(available_template, available_data)
    
    writer.save()

if __name__ == "__main__":
    print("Creating sample template output file...")
    create_sample_template()
    print("\nTemplate structure definitions are ready for use.")
    print("Key classes: TemplateStructure, TemplateWriter, SheetTemplate, ColumnDefinition")