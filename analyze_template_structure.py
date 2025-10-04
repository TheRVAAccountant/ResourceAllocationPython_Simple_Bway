#!/usr/bin/env python3
"""
Excel Template Structure Analyzer

This script analyzes the Daily Summary Log 2025.xlsx file to understand its structure
and provide detailed insights for replicating the format in output files.
"""

import json
import warnings
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

# Suppress pandas warnings for cleaner output
warnings.filterwarnings("ignore", category=UserWarning)


class ExcelStructureAnalyzer:
    """Analyzes Excel file structure and formatting."""

    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.workbook = None
        self.analysis_results = {}

    def analyze(self) -> dict[str, Any]:
        """Perform comprehensive analysis of the Excel file."""
        print(f"\n{'='*80}")
        print(f"ANALYZING EXCEL FILE: {self.file_path.name}")
        print(f"{'='*80}")

        if not self.file_path.exists():
            raise FileNotFoundError(f"File not found: {self.file_path}")

        # Load workbook with openpyxl for formatting analysis
        self.workbook = openpyxl.load_workbook(self.file_path, data_only=False)

        # Get basic file info
        self.analysis_results = {
            "file_path": str(self.file_path),
            "file_size_mb": self.file_path.stat().st_size / (1024 * 1024),
            "last_modified": datetime.fromtimestamp(self.file_path.stat().st_mtime).isoformat(),
            "sheet_names": self.workbook.sheetnames,
            "total_sheets": len(self.workbook.sheetnames),
            "sheets_analysis": {},
        }

        print(f"File Size: {self.analysis_results['file_size_mb']:.2f} MB")
        print(f"Last Modified: {self.analysis_results['last_modified']}")
        print(f"Total Sheets: {self.analysis_results['total_sheets']}")
        print(f"Sheet Names: {', '.join(self.analysis_results['sheet_names'])}")

        # Analyze each sheet
        for sheet_name in self.workbook.sheetnames:
            print(f"\n{'-'*60}")
            print(f"ANALYZING SHEET: {sheet_name}")
            print(f"{'-'*60}")

            sheet_analysis = self._analyze_sheet(sheet_name)
            self.analysis_results["sheets_analysis"][sheet_name] = sheet_analysis

        return self.analysis_results

    def _analyze_sheet(self, sheet_name: str) -> dict[str, Any]:
        """Analyze a specific sheet."""
        worksheet = self.workbook[sheet_name]

        # Get sheet dimensions
        max_row = worksheet.max_row
        max_col = worksheet.max_column

        # Read data with pandas for easier analysis
        try:
            df = pd.read_excel(self.file_path, sheet_name=sheet_name, header=None)
            data_rows = len(df.dropna(how="all"))
            data_cols = len(df.dropna(how="all", axis=1).columns)
        except Exception as e:
            print(f"Warning: Could not read sheet with pandas: {e}")
            df = pd.DataFrame()
            data_rows = max_row
            data_cols = max_col

        sheet_info = {
            "name": sheet_name,
            "dimensions": {
                "max_row": max_row,
                "max_col": max_col,
                "data_rows": data_rows,
                "data_cols": data_cols,
            },
            "headers": [],
            "sample_data": [],
            "formatting": {},
            "merged_cells": [],
            "data_types": {},
            "special_patterns": [],
        }

        print(f"Dimensions: {max_row} rows × {max_col} columns")
        print(f"Data Rows: {data_rows}, Data Columns: {data_cols}")

        # Analyze headers (first row)
        self._analyze_headers(worksheet, sheet_info)

        # Analyze sample data
        self._analyze_sample_data(worksheet, sheet_info, max_rows=10)

        # Analyze formatting
        self._analyze_formatting(worksheet, sheet_info)

        # Analyze merged cells
        self._analyze_merged_cells(worksheet, sheet_info)

        # Analyze data patterns
        self._analyze_data_patterns(worksheet, sheet_info)

        return sheet_info

    def _analyze_headers(self, worksheet, sheet_info: dict) -> None:
        """Analyze header row."""
        headers = []
        header_formatting = {}

        print("\nHEADERS (Row 1):")
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col)
            value = cell.value
            if value is not None:
                headers.append(str(value).strip())

                # Get cell formatting
                header_formatting[f"col_{col}"] = {
                    "value": str(value).strip(),
                    "font": {
                        "bold": cell.font.bold if cell.font else False,
                        "italic": cell.font.italic if cell.font else False,
                        "size": cell.font.size if cell.font else None,
                        "color": cell.font.color.rgb if cell.font and cell.font.color else None,
                    },
                    "fill": {
                        "color": cell.fill.start_color.rgb
                        if cell.fill and hasattr(cell.fill, "start_color")
                        else None
                    },
                    "alignment": {
                        "horizontal": cell.alignment.horizontal if cell.alignment else None,
                        "vertical": cell.alignment.vertical if cell.alignment else None,
                    },
                }
                print(f"  Column {col}: '{value}' | Bold: {cell.font.bold if cell.font else False}")
            else:
                headers.append("")

        sheet_info["headers"] = headers
        sheet_info["formatting"]["headers"] = header_formatting

    def _analyze_sample_data(self, worksheet, sheet_info: dict, max_rows: int = 10) -> None:
        """Analyze sample data rows."""
        sample_data = []
        data_types = {}

        print(f"\nSAMPLE DATA (First {max_rows} rows):")

        for row in range(2, min(max_rows + 2, worksheet.max_row + 1)):
            row_data = []
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                value = cell.value
                row_data.append(value)

                # Track data types
                col_name = f"col_{col}"
                if col_name not in data_types:
                    data_types[col_name] = set()

                if value is not None:
                    data_types[col_name].add(type(value).__name__)

            sample_data.append(row_data)

            # Print row data for review
            non_empty_values = [str(v) for v in row_data if v is not None and str(v).strip()]
            if non_empty_values:
                print(
                    f"  Row {row}: {' | '.join(non_empty_values[:5])}{'...' if len(non_empty_values) > 5 else ''}"
                )

        sheet_info["sample_data"] = sample_data

        # Convert sets to lists for JSON serialization
        for col, types in data_types.items():
            data_types[col] = list(types)
        sheet_info["data_types"] = data_types

        print("\nDATA TYPES BY COLUMN:")
        for col, types in data_types.items():
            if types:
                print(f"  {col}: {', '.join(types)}")

    def _analyze_formatting(self, worksheet, sheet_info: dict) -> None:
        """Analyze cell formatting patterns."""
        formatting_patterns = {
            "bold_cells": [],
            "colored_cells": [],
            "bordered_cells": [],
            "number_formats": {},
        }

        print("\nFORMATTING ANALYSIS:")

        # Check first 20 rows for formatting patterns
        for row in range(1, min(21, worksheet.max_row + 1)):
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                cell_ref = f"{cell.coordinate}"

                # Check for bold text
                if cell.font and cell.font.bold:
                    formatting_patterns["bold_cells"].append(cell_ref)

                # Check for colored fills
                if cell.fill and hasattr(cell.fill, "start_color") and cell.fill.start_color.rgb:
                    color = cell.fill.start_color.rgb
                    if color not in ["00000000", "FFFFFFFF"]:  # Skip default colors
                        formatting_patterns["colored_cells"].append(
                            {"cell": cell_ref, "color": color}
                        )

                # Check for borders
                if cell.border and any(
                    [
                        cell.border.left.style if cell.border.left else None,
                        cell.border.right.style if cell.border.right else None,
                        cell.border.top.style if cell.border.top else None,
                        cell.border.bottom.style if cell.border.bottom else None,
                    ]
                ):
                    formatting_patterns["bordered_cells"].append(cell_ref)

                # Check number formats
                if cell.number_format and cell.number_format != "General":
                    if cell.number_format not in formatting_patterns["number_formats"]:
                        formatting_patterns["number_formats"][cell.number_format] = []
                    formatting_patterns["number_formats"][cell.number_format].append(cell_ref)

        sheet_info["formatting"]["patterns"] = formatting_patterns

        print(f"  Bold cells: {len(formatting_patterns['bold_cells'])}")
        print(f"  Colored cells: {len(formatting_patterns['colored_cells'])}")
        print(f"  Bordered cells: {len(formatting_patterns['bordered_cells'])}")
        print(f"  Number formats: {list(formatting_patterns['number_formats'].keys())}")

    def _analyze_merged_cells(self, worksheet, sheet_info: dict) -> None:
        """Analyze merged cell ranges."""
        merged_ranges = []

        for merged_range in worksheet.merged_cells.ranges:
            merged_ranges.append(
                {
                    "range": str(merged_range),
                    "top_left": merged_range.top[0].coordinate,
                    "size": {
                        "rows": merged_range.max_row - merged_range.min_row + 1,
                        "cols": merged_range.max_col - merged_range.min_col + 1,
                    },
                }
            )

        sheet_info["merged_cells"] = merged_ranges

        print("\nMERGED CELLS:")
        if merged_ranges:
            for merge in merged_ranges:
                print(f"  {merge['range']} ({merge['size']['rows']}×{merge['size']['cols']})")
        else:
            print("  No merged cells found")

    def _analyze_data_patterns(self, worksheet, sheet_info: dict) -> None:
        """Analyze data organization patterns."""
        patterns = []

        # Look for section headers (rows with specific formatting)
        section_headers = []
        for row in range(1, min(50, worksheet.max_row + 1)):
            cell_a = worksheet.cell(row=row, column=1)

            # Check if this might be a section header
            if cell_a.value and cell_a.font and cell_a.font.bold and isinstance(cell_a.value, str):
                # Check if it spans multiple columns or has special formatting
                is_section_header = False

                # Check for merged cells in this row
                for merged_range in worksheet.merged_cells.ranges:
                    if merged_range.min_row == row:
                        is_section_header = True
                        break

                # Check for colored background
                if (
                    cell_a.fill
                    and hasattr(cell_a.fill, "start_color")
                    and cell_a.fill.start_color.rgb not in ["00000000", "FFFFFFFF"]
                ):
                    is_section_header = True

                if is_section_header:
                    section_headers.append(
                        {
                            "row": row,
                            "text": str(cell_a.value),
                            "formatting": {
                                "bold": cell_a.font.bold if cell_a.font else False,
                                "fill_color": cell_a.fill.start_color.rgb
                                if cell_a.fill and hasattr(cell_a.fill, "start_color")
                                else None,
                            },
                        }
                    )

        if section_headers:
            patterns.append(
                {
                    "type": "section_headers",
                    "description": "Rows that appear to be section headers",
                    "instances": section_headers,
                }
            )

        sheet_info["special_patterns"] = patterns

        print("\nSPECIAL PATTERNS:")
        if patterns:
            for pattern in patterns:
                print(f"  {pattern['type']}: {pattern['description']}")
                if pattern["type"] == "section_headers":
                    for header in pattern["instances"]:
                        print(f"    Row {header['row']}: '{header['text']}'")
        else:
            print("  No special patterns detected")

    def generate_report(self) -> str:
        """Generate a comprehensive text report."""
        report = []
        report.append("=" * 80)
        report.append("EXCEL STRUCTURE ANALYSIS REPORT")
        report.append(f"File: {self.analysis_results['file_path']}")
        report.append(f"Generated: {datetime.now().isoformat()}")
        report.append("=" * 80)

        # File summary
        report.append("\nFILE SUMMARY:")
        report.append(f"  Size: {self.analysis_results['file_size_mb']:.2f} MB")
        report.append(f"  Total Sheets: {self.analysis_results['total_sheets']}")
        report.append(f"  Sheet Names: {', '.join(self.analysis_results['sheet_names'])}")

        # Detailed sheet analysis
        for sheet_name, sheet_data in self.analysis_results["sheets_analysis"].items():
            report.append(f"\n{'='*60}")
            report.append(f"SHEET: {sheet_name}")
            report.append(f"{'='*60}")

            # Dimensions
            dims = sheet_data["dimensions"]
            report.append("\nDIMENSIONS:")
            report.append(f"  Total: {dims['max_row']} rows × {dims['max_col']} columns")
            report.append(f"  Data: {dims['data_rows']} rows × {dims['data_cols']} columns")

            # Headers
            report.append("\nHEADERS:")
            for i, header in enumerate(sheet_data["headers"], 1):
                if header:
                    report.append(f"  Column {i}: '{header}'")

            # Sample data
            report.append("\nSAMPLE DATA:")
            for i, row in enumerate(sheet_data["sample_data"][:5], 2):
                non_empty = [str(v) for v in row if v is not None and str(v).strip()]
                if non_empty:
                    report.append(f"  Row {i}: {' | '.join(non_empty[:5])}")

            # Formatting summary
            fmt = sheet_data["formatting"]
            if "patterns" in fmt:
                patterns = fmt["patterns"]
                report.append("\nFORMATTING:")
                report.append(f"  Bold cells: {len(patterns.get('bold_cells', []))}")
                report.append(f"  Colored cells: {len(patterns.get('colored_cells', []))}")
                report.append(f"  Bordered cells: {len(patterns.get('bordered_cells', []))}")
                if patterns.get("number_formats"):
                    report.append(f"  Number formats: {list(patterns['number_formats'].keys())}")

            # Merged cells
            if sheet_data["merged_cells"]:
                report.append("\nMERGED CELLS:")
                for merge in sheet_data["merged_cells"]:
                    report.append(
                        f"  {merge['range']} ({merge['size']['rows']}×{merge['size']['cols']})"
                    )

            # Special patterns
            if sheet_data["special_patterns"]:
                report.append("\nSPECIAL PATTERNS:")
                for pattern in sheet_data["special_patterns"]:
                    report.append(f"  {pattern['type']}: {len(pattern.get('instances', []))} found")

        return "\n".join(report)

    def save_analysis(self, output_file: str | None = None) -> None:
        """Save analysis results to JSON file."""
        if output_file is None:
            output_file = self.file_path.stem + "_structure_analysis.json"

        with open(output_file, "w", encoding="utf-8") as f:
            json.dump(self.analysis_results, f, indent=2, default=str)

        print(f"\nAnalysis saved to: {output_file}")


def main():
    """Main analysis function."""
    file_path = (
        "/Users/jeroncrooks/CascadeProjects/ResourceAllocationPython/Daily Summary Log 2025.xlsx"
    )

    try:
        analyzer = ExcelStructureAnalyzer(file_path)
        results = analyzer.analyze()

        # Generate and display report
        report = analyzer.generate_report()
        print("\n" + report)

        # Save detailed analysis
        analyzer.save_analysis("template_structure_analysis.json")

        # Special focus on "Daily Details" sheet if it exists
        if "Daily Details" in results["sheets_analysis"]:
            print(f"\n{'='*80}")
            print("SPECIAL FOCUS: Daily Details Sheet")
            print(f"{'='*80}")

            daily_details = results["sheets_analysis"]["Daily Details"]

            print("\nThis appears to be the main output format template.")
            print("Key characteristics:")
            print(f"  - Headers: {len([h for h in daily_details['headers'] if h])}")
            print(f"  - Data rows: {daily_details['dimensions']['data_rows']}")
            print(f"  - Merged cells: {len(daily_details['merged_cells'])}")
            print(f"  - Special patterns: {len(daily_details['special_patterns'])}")

            if daily_details["headers"]:
                print("\nColumn structure:")
                for i, header in enumerate(daily_details["headers"], 1):
                    if header:
                        data_type = daily_details["data_types"].get(f"col_{i}", ["unknown"])
                        print(f"  {i:2d}. {header:<20} | Type: {', '.join(data_type)}")

        print(f"\n{'='*80}")
        print("ANALYSIS COMPLETE")
        print(f"{'='*80}")
        print("Use the generated JSON file and this report to understand the")
        print("structure that should be replicated in your output files.")

    except Exception as e:
        print(f"Error analyzing file: {e}")
        import traceback

        traceback.print_exc()


if __name__ == "__main__":
    main()
