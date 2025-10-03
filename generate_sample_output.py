"""Generate a sample output file to demonstrate the current output structure."""

from datetime import datetime, date
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


def generate_sample_output():
    """Generate a sample allocation output file showing current structure."""
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet if exists
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Create Allocations sheet
    allocations = wb.create_sheet("Allocations")
    
    # Define styles for headers
    header_fill = PatternFill(start_color="46bdc6", end_color="46bdc6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Thick border style for daily sections (matching GAS)
    thick_border = Border(
        left=Side(style="thick"),
        right=Side(style="thick"),
        top=Side(style="thick"),
        bottom=Side(style="thick")
    )
    
    # Row 1: Title with thick borders (if borders enabled)
    title_cell = allocations.cell(row=1, column=1, value="Vehicle Allocation")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")
    allocations.merge_cells('A1:D1')
    
    # Row 2: Date
    date_cell = allocations.cell(row=2, column=1, value=f"Date: {date.today()}")
    date_cell.font = Font(bold=True)
    allocations.merge_cells('A2:D2')
    
    # Row 3: Headers
    headers = ["Driver ID", "Vehicle ID", "Status", "Timestamp"]
    for col, header in enumerate(headers, 1):
        cell = allocations.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = header_border
        cell.alignment = header_alignment
    
    # Sample allocation data
    sample_data = [
        # Allocated vehicles
        ["EMP001", "VAN001", "Allocated", datetime.now()],
        ["EMP001", "VAN002", "Allocated", datetime.now()],
        ["EMP002", "VAN003", "Allocated", datetime.now()],
        ["EMP003", "VAN004", "Allocated", datetime.now()],
        ["EMP003", "VAN005", "Allocated", datetime.now()],
        ["EMP004", "VAN006", "Allocated", datetime.now()],
        ["EMP005", "VAN007", "Allocated", datetime.now()],
        # Unallocated vehicles
        ["Unallocated", "VAN008", "Unallocated", datetime.now()],
        ["Unallocated", "VAN009", "Unallocated", datetime.now()],
        ["Unallocated", "VAN010", "Unallocated", datetime.now()],
    ]
    
    # Write data
    for row_idx, row_data in enumerate(sample_data, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = allocations.cell(row=row_idx, column=col_idx)
            if isinstance(value, datetime):
                cell.value = value.strftime("%Y-%m-%d %H:%M:%S")
            else:
                cell.value = value
            
            # Add borders
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                bottom=Side(style="thin")
            )
    
    # Apply thick borders around the entire section (if borders enabled)
    # This simulates what happens when borders_checkbox is checked
    for row in range(1, 14):
        allocations.cell(row=row, column=1).border = Border(
            left=Side(style="thick"),
            right=Side(style="thin"),
            top=Side(style="thick") if row == 1 else Side(style="thin"),
            bottom=Side(style="thick") if row == 13 else Side(style="thin")
        )
        allocations.cell(row=row, column=4).border = Border(
            left=Side(style="thin"),
            right=Side(style="thick"),
            top=Side(style="thick") if row == 1 else Side(style="thin"),
            bottom=Side(style="thick") if row == 13 else Side(style="thin")
        )
    
    for col in range(2, 4):
        allocations.cell(row=1, column=col).border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thick"),
            bottom=Side(style="thin")
        )
        allocations.cell(row=13, column=col).border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thick")
        )
    
    # Adjust column widths
    allocations.column_dimensions['A'].width = 15
    allocations.column_dimensions['B'].width = 15
    allocations.column_dimensions['C'].width = 12
    allocations.column_dimensions['D'].width = 20
    
    # Add summary section
    summary_row = 15
    allocations.cell(row=summary_row, column=1, value="Summary").font = Font(bold=True, size=12)
    allocations.merge_cells(f'A{summary_row}:D{summary_row}')
    
    summary_data = [
        ["Total Drivers:", 5],
        ["Total Vehicles:", 10],
        ["Allocated Vehicles:", 7],
        ["Unallocated Vehicles:", 3],
        ["Allocation Rate:", "70%"],
    ]
    
    for idx, (label, value) in enumerate(summary_data, summary_row + 1):
        allocations.cell(row=idx, column=1, value=label).font = Font(bold=True)
        allocations.cell(row=idx, column=2, value=value)
    
    # Save with dynamic filename
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"Allocation_Results_{timestamp}.xlsx"
    wb.save(filename)
    
    print(f"Sample output file created: {filename}")
    print("\nCurrent Output File Structure:")
    print("=" * 50)
    print("Sheet: 'Allocations'")
    print("-" * 50)
    print("Row 1: Title (merged A1:D1)")
    print("       'Vehicle Allocation'")
    print("Row 2: Date (merged A2:D2)")
    print(f"       'Date: {date.today()}'")
    print("Row 3: Headers")
    print("       Driver ID | Vehicle ID | Status | Timestamp")
    print("Row 4+: Allocation Data")
    print("       - Allocated vehicles with driver assignments")
    print("       - Unallocated vehicles marked as 'Unallocated'")
    print("\nFeatures:")
    print("- Thick borders around daily sections (when enabled)")
    print("- Summary statistics at bottom")
    print("- Dynamic filename with timestamp")
    print("- Formatted headers with teal background")
    
    return filename


def show_detailed_structure():
    """Show detailed description of the output file structure."""
    
    print("\n" + "=" * 70)
    print("DETAILED OUTPUT FILE STRUCTURE")
    print("=" * 70)
    
    print("\n📁 File Naming:")
    print("  Pattern: Allocation_Results_YYYY-MM-DD_HH-MM.xlsx")
    print("  Example: Allocation_Results_2025-01-04_14-30.xlsx")
    
    print("\n📊 Sheet Structure:")
    print("\n1. ALLOCATIONS SHEET")
    print("   " + "-" * 40)
    print("   Layout:")
    print("   • Row 1: Title 'Vehicle Allocation' (merged cells)")
    print("   • Row 2: Date stamp")
    print("   • Row 3: Column headers")
    print("   • Row 4+: Allocation data")
    print("\n   Columns:")
    print("   • A: Driver ID (e.g., 'EMP001' or 'Unallocated')")
    print("   • B: Vehicle ID (e.g., 'VAN001')")
    print("   • C: Status ('Allocated' or 'Unallocated')")
    print("   • D: Timestamp (when allocation was made)")
    
    print("\n   Formatting:")
    print("   • Headers: Teal background (#46bdc6), white text, bold")
    print("   • Borders: Thin borders on all cells")
    print("   • Daily sections: Thick borders when enabled")
    print("   • Title: Large bold text, centered")
    
    print("\n   Summary Section (optional):")
    print("   • Total Drivers")
    print("   • Total Vehicles")
    print("   • Allocated Vehicles")
    print("   • Unallocated Vehicles")
    print("   • Allocation Rate (%)")
    
    print("\n🔄 Comparison with Google Apps Script Output:")
    print("   " + "-" * 40)
    print("   GAS Output:")
    print("   • Updates existing Daily Summary Log")
    print("   • Adds to 'Daily Details' sheet")
    print("   • Creates 'Results' and 'Unassigned Vans' sheets")
    print("   • Uses thick borders for daily sections")
    print("\n   Python Output:")
    print("   • Creates new file with timestamp")
    print("   • Single 'Allocations' sheet")
    print("   • Optional thick borders matching GAS style")
    print("   • Cleaner, focused output structure")
    
    print("\n✨ Key Features:")
    print("   • Simple, clear structure")
    print("   • No overwrites (timestamp in filename)")
    print("   • Professional formatting")
    print("   • Easy to import into other systems")
    print("   • Compatible with Excel filters and pivot tables")
    
    print("\n" + "=" * 70)


if __name__ == "__main__":
    # Generate sample file
    filename = generate_sample_output()
    
    # Show detailed structure
    show_detailed_structure()
    
    print(f"\n✅ Sample file '{filename}' has been created in the current directory.")
    print("   Open it in Excel to see the actual formatting and structure.")