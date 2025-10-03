#!/usr/bin/env python3
"""
Focused Excel Template Structure Analysis

This script provides a focused analysis of the key sheet types in Daily Summary Log 2025.xlsx
to understand the output format structure.
"""

import pandas as pd
import openpyxl
from pathlib import Path
from datetime import datetime

def analyze_key_sheets():
    """Analyze the key sheet types in the template."""
    file_path = "/Users/jeroncrooks/CascadeProjects/ResourceAllocationPython/Daily Summary Log 2025.xlsx"
    workbook = openpyxl.load_workbook(file_path, data_only=False)
    
    print("="*80)
    print("FOCUSED STRUCTURE ANALYSIS - Daily Summary Log 2025.xlsx")
    print("="*80)
    
    # 1. Daily Details Sheet (Main output format)
    print("\n1. DAILY DETAILS SHEET (Main Output Format)")
    print("-" * 50)
    analyze_daily_details(workbook)
    
    # 2. Available & Unassigned Sheets
    print("\n2. AVAILABLE & UNASSIGNED SHEETS")
    print("-" * 50)
    analyze_available_unassigned(workbook)
    
    # 3. Results Sheets
    print("\n3. RESULTS SHEETS")
    print("-" * 50)
    analyze_results_sheets(workbook)
    
    # 4. Configuration/Reference Sheets
    print("\n4. CONFIGURATION/REFERENCE SHEETS")
    print("-" * 50)
    analyze_config_sheets(workbook)

def analyze_daily_details(workbook):
    """Analyze the Daily Details sheet structure."""
    ws = workbook['Daily Details']
    
    print(f"Sheet Dimensions: {ws.max_row} rows × {ws.max_column} columns")
    
    # Headers analysis
    headers = []
    for col in range(1, 25):  # First 24 columns have data
        cell = ws.cell(row=1, column=col)
        if cell.value:
            headers.append({
                'col': col,
                'name': str(cell.value).strip(),
                'bold': cell.font.bold if cell.font else False,
                'data_type': 'varies'  # Will be determined by sample data
            })
    
    print(f"\nColumn Structure ({len(headers)} columns):")
    for h in headers:
        print(f"  {h['col']:2d}. {h['name']:<25} | Bold: {h['bold']}")
    
    # Sample data analysis
    print(f"\nSample Data (First 5 rows):")
    for row in range(2, 7):
        sample_row = []
        for col in range(1, 6):  # First 5 columns
            cell = ws.cell(row=row, column=col)
            if cell.value:
                sample_row.append(str(cell.value)[:20])
        if sample_row:
            print(f"  Row {row}: {' | '.join(sample_row)}")
    
    # Formatting analysis
    print(f"\nFormatting Characteristics:")
    header_cell = ws.cell(row=1, column=1)
    print(f"  - Header row is bold: {header_cell.font.bold if header_cell.font else False}")
    print(f"  - Header background color: {header_cell.fill.start_color.rgb if header_cell.fill and hasattr(header_cell.fill, 'start_color') else 'Default'}")
    
    # Check for data validation or special formatting
    special_formats = set()
    for row in range(1, min(21, ws.max_row + 1)):
        for col in range(1, min(25, ws.max_column + 1)):
            cell = ws.cell(row=row, column=col)
            if cell.number_format != 'General':
                special_formats.add(cell.number_format)
    
    if special_formats:
        print(f"  - Special number formats: {list(special_formats)}")

def analyze_available_unassigned(workbook):
    """Analyze Available & Unassigned sheet pattern."""
    # Find a recent Available & Unassigned sheet
    target_sheet = None
    for sheet_name in workbook.sheetnames:
        if 'Available & Unassign' in sheet_name:
            target_sheet = sheet_name
            break
    
    if not target_sheet:
        print("No Available & Unassigned sheet found")
        return
    
    ws = workbook[target_sheet]
    print(f"Analyzing: {target_sheet}")
    print(f"Dimensions: {ws.max_row} rows × {ws.max_column} columns")
    
    # Headers
    headers = []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        if cell.value:
            headers.append(str(cell.value).strip())
    
    print(f"\nColumns ({len(headers)}):")
    for i, header in enumerate(headers, 1):
        print(f"  {i:2d}. {header}")
    
    # Data characteristics
    data_rows = 0
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value:
            data_rows += 1
    
    print(f"\nData rows: {data_rows}")

def analyze_results_sheets(workbook):
    """Analyze Results sheet pattern."""
    # Find a recent Results sheet
    target_sheet = None
    for sheet_name in workbook.sheetnames:
        if 'Results' in sheet_name and 'Available' not in sheet_name:
            target_sheet = sheet_name
            break
    
    if not target_sheet:
        print("No Results sheet found")
        return
    
    ws = workbook[target_sheet]
    print(f"Analyzing: {target_sheet}")
    print(f"Dimensions: {ws.max_row} rows × {ws.max_column} columns")
    
    # Headers
    headers = []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        if cell.value:
            headers.append(str(cell.value).strip())
    
    print(f"\nColumns ({len(headers)}):")
    for i, header in enumerate(headers, 1):
        print(f"  {i:2d}. {header}")
    
    # Check for time formatting
    wave_col = None
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        if cell.value and 'Wave' in str(cell.value):
            wave_col = col
            break
    
    if wave_col:
        sample_cell = ws.cell(row=2, column=wave_col)
        print(f"\nWave column format: {sample_cell.number_format}")
        print(f"Sample Wave value: {sample_cell.value}")

def analyze_config_sheets(workbook):
    """Analyze configuration and reference sheets."""
    config_sheets = ['Employees', 'Route Types', 'Vehicle Status', 'System Configuration']
    
    for sheet_name in config_sheets:
        if sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            print(f"\n{sheet_name}:")
            print(f"  Dimensions: {ws.max_row} rows × {ws.max_column} columns")
            
            # Get first few column headers
            headers = []
            for col in range(1, min(6, ws.max_column + 1)):
                cell = ws.cell(row=1, column=col)
                if cell.value:
                    headers.append(str(cell.value).strip())
            
            if headers:
                print(f"  Key columns: {', '.join(headers)}")

def generate_structure_summary():
    """Generate a summary of the expected output structure."""
    print("\n" + "="*80)
    print("OUTPUT STRUCTURE REQUIREMENTS SUMMARY")
    print("="*80)
    
    print("""
KEY FINDINGS:

1. DAILY DETAILS SHEET (Main output format):
   - Primary sheet for detailed daily allocation results
   - 24 main columns with specific headers
   - Bold headers with colored background
   - Date format: datetime objects
   - Route codes: CX1, CX2, etc.
   - Employee names in full format
   - Asset/Van IDs: BW prefix format
   - Delivery pace columns with time formats
   - Package delivery/return counts

2. AVAILABLE & UNASSIGNED SHEETS:
   - 15 columns tracking vehicle availability
   - Vehicle details: Year, Make, Model, VIN, etc.
   - Status tracking: Grounded dates, operational status
   - Format: MM-DD-YY - Available & Unassign

3. RESULTS SHEETS:
   - 11 columns with allocation results
   - Route assignments with timing
   - Wave times in h:mm am/pm format
   - Service type descriptions
   - Format: MM-DD-YY - Results

4. FORMATTING STANDARDS:
   - Headers: Bold text with colored backgrounds
   - Date formats: MM/DD/YYYY and datetime objects
   - Time formats: h:mm am/pm
   - Borders: Used in Daily Details for data separation
   - Number formats: Preserve Excel formatting for times/dates

5. NAMING CONVENTIONS:
   - Sheet names: "MM-DD-YY - Description" format
   - Route codes: CX prefix with numbers
   - Vehicle IDs: BW prefix with numbers
   - Staging locations: STG.G.X format

IMPLEMENTATION REQUIREMENTS:
- Preserve exact column headers and order
- Apply proper formatting (bold headers, colors, borders)
- Use correct data types (datetime, time, string, float)
- Follow naming conventions for consistency
- Maintain Excel number formats for proper display
""")

if __name__ == "__main__":
    analyze_key_sheets()
    generate_structure_summary()